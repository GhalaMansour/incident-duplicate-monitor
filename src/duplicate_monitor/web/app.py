"""Live-monitor web dashboard.

FastAPI server that renders the duplicate review experience. The left
panel lists detected groups; the main area renders each SR in a group
as a column with fields as rows so reviewers can compare side by side.

Run modes (see ``python -m duplicate_monitor --help``):

* ``web``  — the dashboard only (assumes the poller runs elsewhere).
* ``both`` — the poller in a background thread plus the dashboard.
"""

from __future__ import annotations

import io
import json
import os
import pickle
import time
from datetime import datetime
from pathlib import Path
from typing import Optional

from fastapi import FastAPI, File, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, Response
from fastapi.staticfiles import StaticFiles

from duplicate_monitor.core.config import CFG, PACKAGE_DIR
from duplicate_monitor.storage import db

_HERE = Path(__file__).resolve().parent
_ROOT = PACKAGE_DIR

app = FastAPI(title="Kidana Live Monitor", docs_url=None, redoc_url=None)

# Mount static files (logo, etc.)
_STATIC = _HERE / "static"
if _STATIC.exists():
    app.mount("/static", StaticFiles(directory=str(_STATIC)), name="static")


@app.on_event("startup")
async def _auto_scan_on_start():
    """On startup: scan immediately if stale, then start periodic background loop."""
    import asyncio
    import threading as _th

    await asyncio.sleep(3)  # let uvicorn finish binding first
    global _SCAN_STATE, _bg_scan_maximo  # noqa: F821 — defined below
    if not CFG.has_maximo_credentials:
        return

    # ── 0. Proactively seed the row cache from the last saved scan ──────
    # This MUST happen before the quick-scan loop starts.  Without it the very
    # first quick scan (T+15 s) would start with an empty cache, run detect()
    # on only 200 rows, and save ~8 groups — overwriting the good full-scan pkl
    # with a tiny result that then perpetuates itself on every restart.
    import logging as _logging

    _wlog = _logging.getLogger("live_monitor.web")
    try:
        from duplicate_monitor import scanner as _sc

        _sc._seed_cache_from_disk()
        _wlog.info(
            "Startup: pre-seeded row cache (%d SRs, ref=%d)",
            len(_sc._rows_cache),
            _sc._full_scan_row_count,
        )
    except Exception as _seed_err:
        _wlog.warning("Startup: pre-seed failed: %s", _seed_err)

    # ── 1. Startup scan ONLY if no saved data exists at all ─────────────
    # Never auto-overwrite a good existing result. The user controls when
    # a full re-scan happens by pressing "سحب من Maximo". Background loops
    # only run quick-scan (adds new SRs; never replaces the full dataset).
    scan = _load_scan()
    no_data = scan is None or scan.get("sr_count", 0) < 10
    if no_data and not _SCAN_STATE.get("running"):
        _th.Thread(target=_bg_scan_maximo, daemon=True, name="startup_scan").start()

    # ── 2. Quick scan loop — catches new Maximo SRs within seconds ──────
    # Quick scan only ADDS new SRs / groups; it never replaces the full
    # dataset, so the 217 groups the user fetched will never disappear.
    def _quick_scan_loop():
        import time as _time, os as _os

        while True:
            try:
                interval = max(
                    10, int(_os.environ.get("LM_QUICK_SCAN_SEC", str(CFG.quick_scan_sec)))
                )
            except Exception:
                interval = CFG.quick_scan_sec
            _time.sleep(interval)
            if CFG.has_maximo_credentials:
                _bg_quick_scan_maximo()

    _th.Thread(target=_quick_scan_loop, daemon=True, name="quick_scan_scheduler").start()

    # ── 3. Periodic full re-scan — DISABLED ─────────────────────────────
    # Full scans now only happen when the user presses the button.
    # This guarantees groups never disappear between sessions.


# ── decisions ─────────────────────────────────────────────────────────────────
_DEC_PATH = _HERE / "live_decisions.json"


def _load_dec() -> dict:
    if _DEC_PATH.exists():
        try:
            return json.loads(_DEC_PATH.read_text("utf-8"))
        except:
            pass
    return {}


def _save_dec(d: dict):
    _DEC_PATH.write_text(json.dumps(d, ensure_ascii=False, indent=2), "utf-8")


# ── scan ──────────────────────────────────────────────────────────────────────
def _load_scan() -> Optional[dict]:
    if not CFG.scan_pkl.exists():
        return None
    try:
        return pickle.loads(CFG.scan_pkl.read_bytes())
    except:
        return None


_HTML_TAG_RE = __import__("re").compile(r"<[^>]+>")
_HTML_COMMENT_RE = __import__("re").compile(r"<!--.*?-->", __import__("re").DOTALL)


def _clean_html(s: str) -> str:
    """Strip HTML tags/comments from Maximo long-description text."""
    if not s:
        return ""
    s = _HTML_COMMENT_RE.sub("", str(s))
    s = _HTML_TAG_RE.sub("\n", s)
    # collapse whitespace
    s = "\n".join(line.strip() for line in s.splitlines() if line.strip())
    # decode common entities
    s = (
        s.replace("&nbsp;", " ")
        .replace("&amp;", "&")
        .replace("&lt;", "<")
        .replace("&gt;", ">")
        .replace("&quot;", '"')
    )
    return s.strip()


def _tier(score: int) -> str:
    if score >= 8:
        return "confirmed"
    if score >= 5:
        return "possible"
    return "weak"


def _tier_ar(score: int) -> str:
    if score >= 8:
        return " مؤكد"
    if score >= 5:
        return " محتمل"
    return " ضعيف"


# ── API ───────────────────────────────────────────────────────────────────────
@app.get("/api/scan")
def api_scan():
    scan = _load_scan()
    health = db.health_summary()
    decisions = _load_dec()
    if not scan:
        return JSONResponse(
            {"ready": False, "health": health},
            headers={"Cache-Control": "no-store"},
        )

    groups = []

    def _member_view(m: dict, idx: int) -> dict:
        # detect() members store:  reporter = Arabic name (REPORTED NAME col),
        #                          reported_by = username (Reported By col)
        # raw_rows store:          reporter = username (no reported_by key),
        #                          reported_name = Arabic caller name
        # Detect format is identified by the presence of 'reported_by'.
        _detect_fmt = "reported_by" in m
        _username = m.get("reported_by", "") if _detect_fmt else (m.get("reporter") or "")
        _arabic_name = (m.get("reporter") or "") if _detect_fmt else (m.get("reported_name") or "")
        return {
            "sr": m.get("sr", ""),
            "loc": m.get("loc", ""),
            "loc_ar": m.get("loc_ar", ""),
            "asset": m.get("asset", ""),
            "asset_ar": m.get("asset_ar", ""),
            "fault": m.get("fault_full") or m.get("fault_orig") or m.get("fault") or "",
            "summary": m.get("fault_full") or m.get("fault_orig") or m.get("fault") or "",
            "reported": m.get("reported", "")[:16],
            "status": m.get("status", ""),
            "status_desc": m.get("status_desc", ""),
            "detail": _clean_html(m.get("detail") or ""),
            "reporter": _username,
            "reporter_display": m.get("reporter_display") or "",
            "workzone": m.get("workzone") or "",
            "priority": m.get("priority") or "",
            "priority_desc": m.get("priority_desc") or "",
            "site": m.get("site") or m.get("siteid") or "",
            "caller_name": m.get("caller_name") or "",
            "caller_phone": m.get("caller_phone") or "",
            "caller_email": m.get("caller_email") or "",
            "caller_party": m.get("caller_party") or "",
            "reporter_phone": m.get("reporter_phone") or "",
            "reporter_email": m.get("reporter_email") or "",
            "source": m.get("source") or "",
            "source_desc": m.get("source_desc") or "",
            "ownergroup": m.get("ownergroup") or "",
            "assigned_ownergroup": m.get("assigned_ownergroup") or "",
            "lat": m.get("lat") or "",
            "lon": m.get("lon") or "",
            # Kidana custom fields
            "region": m.get("region") or "",
            "block": m.get("block") or "",
            "requestor_no": m.get("requestor_no") or "",
            "reported_name": _arabic_name,
            "contract": m.get("contract") or "",
            "contractor": m.get("contractor") or m.get("party") or "",
            "party": m.get("party") or "",
            "resp_time": m.get("resp_time") or "",
            "resp_esc": m.get("resp_esc") or "",
            "statusdate": (m.get("statusdate") or "")[:16],
            "actstart": (m.get("actstart") or m.get("targetstart") or "")[:16],
            "actfinish": (m.get("actfinish") or "")[:16],
            "is_origin": idx == 0,
        }

    def _append_group(
        *,
        gid: str,
        score: int,
        reasons: str,
        members: list,
        decision: str = "",
        note: str = "",
        parent_id: str = "",
        sub_idx: int = -1,
    ):
        groups.append(
            {
                "id": gid,
                "parent_id": parent_id,
                "sub_idx": sub_idx,
                "score": score,
                "tier": _tier(score),
                "tier_ar": _tier_ar(score),
                "reasons": reasons,
                "size": len(members),
                "decision": decision,
                "note": note,
                "members": [_member_view(m, idx) for idx, m in enumerate(members)],
            }
        )

    for g in scan.get("groups", []):
        members = g.get("members", [])
        gid = "_".join(sorted(m["sr"] for m in members))
        score = g.get("score", 0)
        reasons = g.get("reasons", "")
        if isinstance(reasons, list):
            reasons = " · ".join(reasons)
        sorted_members = sorted(members, key=lambda m: (m.get("reported_dt") or datetime.max))
        dec = decisions.get(gid, {})
        decision = dec.get("decision", "")
        if decision in ("duplicate", "different"):
            _append_group(
                gid=gid,
                score=score,
                reasons=reasons,
                members=sorted_members,
                decision=decision,
                note=dec.get("note", ""),
            )
            continue

        used_srs = set()
        for idx, sub in enumerate(dec.get("sub_decisions", []) or []):
            sub_dec = sub.get("decision", "")
            if sub_dec not in ("duplicate", "different"):
                continue
            sub_srs = set(str(x) for x in sub.get("srs", []))
            sub_members = [m for m in sorted_members if str(m.get("sr", "")) in sub_srs]
            if not sub_members:
                continue
            used_srs.update(str(m.get("sr", "")) for m in sub_members)
            _append_group(
                gid=f"{gid}__sub_{idx}",
                parent_id=gid,
                sub_idx=idx,
                score=score,
                reasons=reasons,
                members=sub_members,
                decision=sub_dec,
                note=sub.get("note", ""),
            )

        remaining = [m for m in sorted_members if str(m.get("sr", "")) not in used_srs]
        if remaining:
            _append_group(gid=gid, score=score, reasons=reasons, members=remaining)
    groups.sort(key=lambda g: (bool(g["decision"]), -g["score"]))
    age = time.time() - CFG.scan_pkl.stat().st_mtime if CFG.scan_pkl.exists() else None
    # ── Aggregations for KPIs / map ────────────────────────────────
    from collections import Counter

    total_srs = scan.get("sr_count", 0)
    sr_in_groups = set()
    contractor_c = Counter()
    fault_c = Counter()
    n_confirmed_dec = 0  # groups decided as "duplicate" (للإغلاق)
    n_different_dec = 0
    geo_dup_points: list = []
    geo_normal_points: list = []
    dup_locs = set()

    for g in groups:
        for m in g["members"]:
            sr = m.get("sr", "")
            if sr:
                sr_in_groups.add(sr)
            og = m.get("ownergroup") or m.get("assigned_ownergroup") or ""
            if og:
                contractor_c[og] += 1
            f = m.get("fault") or ""
            if f:
                fault_c[f] += 1
            dup_locs.add(m.get("loc", ""))
            try:
                lat = float(m.get("lat") or 0)
                lon = float(m.get("lon") or 0)
                if lat and lon:
                    geo_dup_points.append(
                        {
                            "lat": lat,
                            "lon": lon,
                            "sr": sr,
                            "loc": m.get("loc", ""),
                            "fault": (m.get("fault") or "")[:60],
                        }
                    )
            except Exception:
                pass
        dec = g.get("decision", "")
        if dec == "duplicate":
            n_confirmed_dec += 1
        elif dec == "different":
            n_different_dec += 1

    # Pull normal-location points from the raw scan (all_rows)
    for r in scan.get("all_rows") or []:
        sr = r.get("sr", "")
        if sr in sr_in_groups:
            continue
        try:
            lat = float(r.get("lat") or 0)
            lon = float(r.get("lon") or 0)
            if lat and lon:
                geo_normal_points.append(
                    {
                        "lat": lat,
                        "lon": lon,
                        "sr": sr,
                        "loc": r.get("loc", ""),
                        "fault": (r.get("fault_orig") or r.get("fault") or "")[:60],
                    }
                )
        except Exception:
            pass

    top_contractor = contractor_c.most_common(1)[0] if contractor_c else ("", 0)
    contractor_total = sum(contractor_c.values())
    top_faults = fault_c.most_common(3)
    n_groups = len(groups)
    n_dup_srs = len(sr_in_groups)
    n_decided = sum(1 for g in groups if g["decision"])
    progress = round(n_decided * 100 / n_groups) if n_groups else 0

    return JSONResponse(
        {
            "ready": True,
            "scanned_at": scan.get("scanned_at", ""),
            "sr_count": total_srs,
            "age_seconds": age,
            "groups": groups,
            "n_confirmed": sum(1 for g in groups if g["tier"] == "confirmed"),
            "n_possible": sum(1 for g in groups if g["tier"] == "possible"),
            "n_weak": sum(1 for g in groups if g["tier"] == "weak"),
            "n_decided": n_decided,
            # New KPI data
            "n_groups": n_groups,
            "n_dup_srs": n_dup_srs,
            "pct_dup": round(n_dup_srs * 100 / total_srs, 1) if total_srs else 0,
            "top_contractor": {
                "name": top_contractor[0],
                "count": top_contractor[1],
                "pct": round(top_contractor[1] * 100 / contractor_total, 1)
                if contractor_total
                else 0,
                "total": contractor_total,
            },
            "top_faults": [{"name": f, "count": c} for f, c in top_faults],
            "n_confirmed_dec": n_confirmed_dec,
            "n_different_dec": n_different_dec,
            "progress_pct": progress,
            "n_remaining": n_groups - n_decided,
            "geo_dup": geo_dup_points,
            "geo_normal": geo_normal_points[:2000],  # cap for browser perf
            "geo_dup_locs": len(dup_locs),
            "geo_normal_locs": max(0, total_srs - n_dup_srs),
            "open_alerts": db.alert_counts().get("open", 0),
            "health": health,
        },
        headers={"Cache-Control": "no-store"},
    )


@app.post("/api/decision")
async def api_decision(request: Request):
    body = await request.json()
    gid = body.get("gid", "").strip()
    decision = body.get("decision", "").strip()  # duplicate|different|""
    note = body.get("note", "").strip()
    srs = body.get("srs", [])

    if not gid or decision not in ("duplicate", "different", ""):
        raise HTTPException(400, "invalid")

    d = _load_dec()
    if decision == "":
        d.pop(gid, None)
    elif srs:
        entry = d.get(gid, {"decision": "partial", "sub_decisions": []})
        if "sub_decisions" not in entry:
            entry["sub_decisions"] = []
        entry["sub_decisions"].append(
            {
                "srs": srs,
                "decision": decision,
                "note": note,
                "ts": datetime.now().isoformat(timespec="seconds"),
            }
        )
        entry["decision"] = "partial"
        d[gid] = entry
    else:
        d[gid] = {
            "decision": decision,
            "note": note,
            "ts": datetime.now().isoformat(timespec="seconds"),
        }
    _save_dec(d)
    return JSONResponse({"ok": True})


@app.delete("/api/decision/{gid}")
def del_decision(gid: str, idx: int = -1):
    d = _load_dec()
    if gid not in d:
        raise HTTPException(404)
    if idx >= 0:
        subs = d[gid].get("sub_decisions", [])
        if 0 <= idx < len(subs):
            subs.pop(idx)
        if not subs:
            d[gid]["decision"] = ""
    else:
        del d[gid]
    _save_dec(d)
    return JSONResponse({"ok": True})


# ── Live alerts feed ──────────────────────────────────────────────────────────
@app.get("/api/alerts")
def api_alerts_feed(since: str = ""):
    """Return recent DB alerts newer than `since` (ISO timestamp)."""
    alerts = db.list_alerts(limit=100)
    if since:
        alerts = [a for a in alerts if a["detected_at"] > since]
    return JSONResponse({"alerts": alerts[:30]})


@app.get("/api/notifications")
def api_notifications():
    """Small topbar payload: current unresolved scan groups + live DB alerts."""
    scan = _load_scan()
    decisions = _load_dec()
    pending_groups = 0
    if scan:
        for g in scan.get("groups", []):
            members = g.get("members", [])
            gid = "_".join(sorted(str(m.get("sr", "")) for m in members if m.get("sr")))
            if gid and not decisions.get(gid, {}).get("decision"):
                pending_groups += 1
    open_alerts = db.list_alerts(state="open", limit=8)
    return JSONResponse(
        {
            "pending_groups": pending_groups,
            "open_alert_count": db.alert_counts().get("open", 0),
            "alerts": open_alerts,
        }
    )


# ── Analytics ─────────────────────────────────────────────────────────────────
@app.get("/api/analytics")
def api_analytics():
    scan = _load_scan()
    decisions = _load_dec()
    if not scan:
        return JSONResponse({"ready": False})
    groups = scan.get("groups", [])
    if not groups:
        running = _SCAN_STATE.get("running", False)
        return JSONResponse(
            {
                "ready": True,
                "total_groups": 0,
                "total_srs": scan.get("sr_count", 0),
                "scanning": running,
                "scan_phase": _SCAN_STATE.get("progress", {}).get("phase", ""),
                "tiers": {"confirmed": 0, "possible": 0, "weak": 0},
                "decisions": {"duplicate": 0, "different": 0, "pending": 0},
                "time_spread": {"same_day": 0, "same_week": 0, "farther": 0},
                "top_faults": [],
                "top_locs": [],
            }
        )

    from collections import Counter

    fault_c: Counter = Counter()
    loc_c: Counter = Counter()
    same_day = same_week = farther = 0
    tier_c: Counter = Counter()
    dec_c: Counter = Counter()

    for g in groups:
        members = g.get("members", [])
        score = g.get("score", 0)
        tier_c[_tier(score)] += 1
        gid = "_".join(sorted(m["sr"] for m in members))
        dec = decisions.get(gid, {}).get("decision", "")
        dec = dec if dec in ("duplicate", "different") else "pending"
        dec_c[dec] += 1

        for m in members:
            f = m.get("fault_orig") or m.get("fault") or ""
            f = f.split(",")[-1].strip()[:40] if "," in f else f[:40]
            if f:
                fault_c[f] += 1
            loc = (m.get("loc") or "")[:40]
            if loc:
                loc_c[loc] += 1

        # time spread within group
        dates = []
        for m in members:
            r = m.get("reported", "")
            if r and len(r) >= 10:
                try:
                    dates.append(datetime.strptime(r[:10], "%Y-%m-%d"))
                except:
                    pass
        if len(dates) >= 2:
            dates.sort()
            span = (dates[-1] - dates[0]).days
            if span == 0:
                same_day += 1
            elif span <= 7:
                same_week += 1
            else:
                farther += 1

    def top(c: Counter, n: int = 10):
        total = sum(c.values()) or 1
        return [
            {"label": k, "count": v, "pct": round(v * 100 / total)} for k, v in c.most_common(n)
        ]

    return JSONResponse(
        {
            "ready": True,
            "total_groups": len(groups),
            "total_srs": scan.get("sr_count", 0),
            "tiers": {k: tier_c.get(k, 0) for k in ("confirmed", "possible", "weak")},
            "decisions": {k: dec_c.get(k, 0) for k in ("duplicate", "different", "pending")},
            "time_spread": {"same_day": same_day, "same_week": same_week, "farther": farther},
            "top_faults": top(fault_c, 12),
            "top_locs": top(loc_c, 12),
        }
    )


# ── Export Excel ──────────────────────────────────────────────────────────────
@app.get("/api/export/excel")
def api_export_excel():
    try:
        return _do_export_excel()
    except Exception as _exc:
        import traceback as _tb

        return JSONResponse(
            status_code=500, content={"error": str(_exc), "trace": _tb.format_exc()}
        )


def _fmt_diff(delta) -> str:
    """Format a timedelta as Arabic short string: +X ثانية / دقيقة / ساعة / يوم."""
    s = int(delta.total_seconds())
    if s < 0:
        s = -s
    if s < 60:
        return f"+{s} ثانية"
    if s < 3600:
        return f"+{s // 60} دقيقة"
    if s < 86400:
        return f"+{s // 3600} ساعة"
    return f"+{s // 86400} يوم"


def _fmt_span(delta) -> str:
    s = int(abs(delta.total_seconds()))
    d = s // 86400
    h = (s % 86400) // 3600
    m = (s % 3600) // 60
    if d:
        return f"{d} يوم {h} ساعة"
    if h:
        return f"{h} ساعة {m} دقيقة"
    return f"{m} دقيقة"


def _parse_date(v):
    from datetime import datetime as _dt

    if not v:
        return None
    s = str(v).strip()
    # Try fromisoformat first — handles +03:00 timezone offsets (Python 3.7+)
    try:
        dt = _dt.fromisoformat(s)
        return dt.replace(tzinfo=None)  # strip tz → naive for comparison
    except Exception:
        pass
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M"):
        try:
            return _dt.strptime(s, fmt)
        except:
            pass
    return None


def _do_export_excel():
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(503, "openpyxl غير مثبّت — شغّل: pip install openpyxl")

    scan = _load_scan()
    decisions = _load_dec()
    groups = scan.get("groups", []) if scan else []
    sr_count = scan.get("sr_count", 0) if scan else 0

    def _grp_decision(members):
        """Resolve the user's review decision for a group.

        Returns (whole_decision, whole_note, per_sr_decision, per_sr_note).
        whole_decision is set when the ENTIRE group was judged at once;
        otherwise per_sr_* carry partial (sub-decision) judgements.
        Key matches api_scan: '_'.join(sorted SRs).
        """
        gid = "_".join(sorted(m.get("sr", "") for m in members))
        dec = decisions.get(gid, {}) or {}
        whole = dec.get("decision", "")
        if whole in ("duplicate", "different"):
            return whole, dec.get("note", ""), {}, {}
        sr_dec, sr_note = {}, {}
        for sub in dec.get("sub_decisions", []) or []:
            sd = sub.get("decision", "")
            if sd not in ("duplicate", "different"):
                continue
            for s in sub.get("srs", []):
                sr_dec[str(s)] = sd
                sr_note[str(s)] = sub.get("note", "")
        return "", dec.get("note", ""), sr_dec, sr_note

    # ── Styles (exact match to مكررات_2026-05-23.xlsx) ───────────────────
    NAVY = PatternFill("solid", fgColor="1C2E4A")
    BLUE_ALT = PatternFill("solid", fgColor="EFF6FF")  # odd groups
    GRN_ALT = PatternFill("solid", fgColor="F0FDF4")  # even groups
    GRN_DEC = PatternFill("solid", fgColor="DCFCE7")
    RED_DEC = PatternFill("solid", fgColor="FECACA")
    GRAY_F = PatternFill("solid", fgColor="F5F5F5")
    TITLE_FONT = Font(color="B9975B", bold=True, size=13)
    HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
    BOLD10 = Font(bold=True, size=10)
    NORM10 = Font(size=10)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=False)
    RIGHT_W = Alignment(horizontal="right", vertical="top", wrap_text=True)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    DEC_LBL = {"duplicate": "مكرر مؤكد", "different": "مختلف"}
    DEC_FILL = {"duplicate": GRN_DEC, "different": RED_DEC}

    def _s(ws_, r, c, val, fill=None, font=NORM10, align=RIGHT):
        cell = ws_.cell(row=r, column=c, value=val)
        cell.border = border
        cell.fill = fill or PatternFill()
        cell.font = font
        cell.alignment = align
        return cell

    # ── Sheet 1: مجموعات المكررات ─────────────────────────────────────────
    n_groups = len(groups)
    n_members = sum(len(g.get("members", [])) for g in groups)
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "مجموعات المكررات"
    ws1.sheet_view.rightToLeft = True

    HEADERS = [
        "رقم المجموعة",
        "ترتيب البلاغ",
        "الفارق عن السابق",
        "أسباب الكشف",
        "Service Request",
        "Site",
        "History",
        "Source",
        "Status",
        "Status Date",
        "Status Description",
        "Summary",
        "Work Zone",
        "LOCATION",
        "وصف المعلم",
        "المربع",
        "المنطقة",
        "Asset",
        "وصف الأصل",
        "REQUESTOR NO.",
        "Contract",
        "Contractor",
        "REPORTED NAME",
        "Reported By",
        "زمن الاستجابه",
        "Response Esclation",
        "Resolution Time",
        "Resolution Escalation",
        "الجهة",
        "Details",
        "Latitude(Y)",
        "Longitude(X)",
        "Internal Priority",
        "Ticket in Other Party",
        "تاريخ فتح البلاغ",
        "تاريخ المباشره",
        "تاريخ المعالجة",
        "القرار",
        "ملاحظة المراجعة",
    ]
    NCOLS = len(HEADERS)  # 39

    # Row 1: merged title
    title = f"مجموعات المكررات المحتملة — {n_groups} مجموعة | {n_members} بلاغ متأثر"
    ws1.cell(row=1, column=1, value=title)
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOLS)
    tc = ws1.cell(row=1, column=1)
    tc.fill = NAVY
    tc.font = TITLE_FONT
    tc.alignment = CENTER
    ws1.row_dimensions[1].height = 26

    # Row 2: headers
    for ci, h in enumerate(HEADERS, 1):
        cell = ws1.cell(row=2, column=ci, value=h)
        cell.fill = NAVY
        cell.font = HDR_FONT
        cell.alignment = CENTER
        cell.border = border
    ws1.row_dimensions[2].height = 22
    ws1.freeze_panes = "A3"

    # Column widths
    CW = [
        14,
        14,
        14,
        45,
        14,
        10,
        8,
        10,
        10,
        20,
        16,
        45,
        12,
        16,
        30,
        10,
        10,
        14,
        25,
        16,
        10,
        30,
        20,
        14,
        20,
        20,
        20,
        20,
        14,
        45,
        14,
        14,
        10,
        16,
        18,
        20,
        20,
        16,
        40,
    ]
    for ci, w in enumerate(CW, 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    cur_data_row = 3
    for gi, g in enumerate(groups, 1):
        members = g.get("members", [])
        reasons = g.get("reasons", "")
        if isinstance(reasons, list):
            reasons = " | ".join(reasons)
        row_fill = BLUE_ALT if gi % 2 == 1 else GRN_ALT
        whole_dec, whole_note, sr_dec, sr_note = _grp_decision(members)

        # Sort members by reported date ascending
        def _sort_key(m):
            d = _parse_date(m.get("reported", ""))
            from datetime import datetime as _dt2

            return d if d else _dt2.min

        members_sorted = sorted(members, key=_sort_key)

        prev_dt = None
        for order, m in enumerate(members_sorted, 1):
            r = cur_data_row
            dt = _parse_date(m.get("reported", ""))
            if order == 1:
                order_lbl = "1 (الأصل)"
                diff_lbl = "—"
            else:
                order_lbl = str(order)
                diff_lbl = _fmt_diff(dt - prev_dt) if dt and prev_dt else "—"
            prev_dt = dt

            vals = [
                f"#{gi}",  # A رقم المجموعة
                order_lbl,  # B ترتيب البلاغ
                diff_lbl,  # C الفارق عن السابق
                reasons if order == 1 else "",  # D أسباب الكشف
                m.get("sr", ""),  # E Service Request
                m.get("siteid", ""),  # F Site
                m.get("history", ""),  # G History
                m.get("source", ""),  # H Source
                m.get("status", ""),  # I Status
                m.get("statusdate") or m.get("reported", ""),  # J Status Date
                m.get("status_desc", ""),  # K Status Description
                (m.get("fault_full") or m.get("fault_orig") or m.get("summary", "")),  # L Summary
                m.get("workzone", ""),  # M Work Zone
                m.get("loc") or m.get("location", ""),  # N LOCATION
                m.get("loc_ar", ""),  # O وصف المعلم
                m.get("block", ""),  # P المربع
                m.get("region", ""),  # Q المنطقة
                m.get("asset", ""),  # R Asset
                m.get("asset_ar", ""),  # S وصف الأصل
                m.get("requestor_no") or m.get("caller_phone", ""),  # T REQUESTOR NO.
                m.get("contract", ""),  # U Contract
                m.get("contractor") or m.get("party") or m.get("ownergroup", ""),  # V Contractor
                m.get("reported_name") or m.get("reporter", ""),  # W REPORTED NAME
                m.get("reporter", ""),  # X Reported By
                m.get("resp_time", ""),  # Y زمن الاستجابه
                m.get("resp_esc", ""),  # Z Response Esclation
                m.get("resol_time", ""),  # AA Resolution Time
                m.get("resol_esc", ""),  # AB Resolution Escalation
                m.get("party") or m.get("caller_party", ""),  # AC الجهة
                m.get("detail", ""),  # AD Details
                m.get("lat", ""),  # AE Latitude(Y)
                m.get("lon", ""),  # AF Longitude(X)
                m.get("priority", ""),  # AG Internal Priority
                m.get("ticket_other", ""),  # AH Ticket in Other Party
                m.get("reported", ""),  # AI تاريخ فتح البلاغ
                m.get("actstart") or m.get("targetstart", ""),  # AJ تاريخ المباشره
                m.get("actfinish", ""),  # AK تاريخ المعالجة
            ]
            # ── Decision + note for THIS SR (whole-group or partial) ──
            _sr = str(m.get("sr", ""))
            if whole_dec:
                eff_dec, eff_note = whole_dec, whole_note
            else:
                eff_dec = sr_dec.get(_sr, "")
                eff_note = sr_note.get(_sr, "")
            vals.append(DEC_LBL.get(eff_dec, "لم يُراجع"))  # AL القرار
            vals.append(eff_note or "")  # AM ملاحظة المراجعة
            eff_fill = DEC_FILL.get(eff_dec, row_fill)  # decided → green/red
            for ci, val in enumerate(vals, 1):
                font = BOLD10 if ci in (1, 2, 3, 5, 38) else NORM10
                al = RIGHT_W if ci in (4, 12, 30, 39) else RIGHT
                _s(ws1, r, ci, val, fill=eff_fill, font=font, align=al)
            ws1.row_dimensions[r].height = 16
            cur_data_row += 1

        # Empty separator row between groups
        cur_data_row += 1

    # ── Sheet 2: ملخص المجموعات ───────────────────────────────────────────
    ws2 = wb.create_sheet("ملخص المجموعات")
    ws2.sheet_view.rightToLeft = True

    ws2.cell(row=1, column=1, value=f"ملخص المجموعات ({n_groups})")
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    tc2 = ws2.cell(row=1, column=1)
    tc2.fill = NAVY
    tc2.font = TITLE_FONT
    tc2.alignment = CENTER
    ws2.row_dimensions[1].height = 26

    H2 = [
        "رقم المجموعة",
        "حجم المجموعة",
        "الثقة",
        "النقاط",
        "أعضاء المجموعة",
        "تاريخ أول بلاغ",
        "تاريخ آخر بلاغ",
        "المدى الزمني",
        "أسباب الكشف",
        "القرار",
        "ملاحظة المراجعة",
    ]
    for ci, h in enumerate(H2, 1):
        cell = ws2.cell(row=2, column=ci, value=h)
        cell.fill = NAVY
        cell.font = HDR_FONT
        cell.alignment = CENTER
        cell.border = border
    ws2.row_dimensions[2].height = 22
    ws2.freeze_panes = "A3"
    for ci, w in enumerate([14, 12, 12, 8, 40, 18, 18, 16, 55, 16, 40], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    TIER_AR = {"confirmed": "مؤكد", "possible": "محتمل", "weak": "ضعيف"}
    for gi, g in enumerate(groups, 1):
        members = g.get("members", [])
        score = g.get("score", 0)
        tier = _tier(score)
        reasons = g.get("reasons", "")
        if isinstance(reasons, list):
            reasons = " | ".join(reasons)
        srs_str = " ، ".join(m.get("sr", "") for m in members)
        dates = sorted(filter(None, (_parse_date(m.get("reported", "")) for m in members)))
        d_first = str(dates[0])[:16] if dates else ""
        d_last = str(dates[-1])[:16] if dates else ""
        span = _fmt_span(dates[-1] - dates[0]) if len(dates) >= 2 else "—"
        # ── group-level decision label ──
        w_dec, w_note, p_dec, p_note = _grp_decision(members)
        if w_dec:
            dec_lbl = DEC_LBL.get(w_dec, "")
            dec_note = w_note
            grp_fill = DEC_FILL.get(w_dec, BLUE_ALT if gi % 2 == 1 else GRN_ALT)
        elif p_dec:
            dec_lbl = "جزئي (راجع التفاصيل)"
            dec_note = w_note
            grp_fill = BLUE_ALT if gi % 2 == 1 else GRN_ALT
        else:
            dec_lbl = "لم يُراجع"
            dec_note = w_note
            grp_fill = BLUE_ALT if gi % 2 == 1 else GRN_ALT
        row_fill = grp_fill
        r = gi + 2
        for ci, val in enumerate(
            [
                f"#{gi}",
                len(members),
                TIER_AR.get(tier, ""),
                score,
                srs_str,
                d_first,
                d_last,
                span,
                reasons,
                dec_lbl,
                dec_note,
            ],
            1,
        ):
            _s(
                ws2,
                r,
                ci,
                val,
                fill=row_fill,
                font=BOLD10 if ci in (1, 2, 3, 4, 10) else NORM10,
                align=RIGHT_W if ci in (5, 9, 11) else RIGHT,
            )
        ws2.row_dimensions[r].height = 16

    # ── Sheet 3: إحصاءات ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("إحصاءات")
    ws3.sheet_view.rightToLeft = True
    ws3.cell(row=1, column=1, value="إحصاءات المجموعات")
    ws3.merge_cells("A1:B1")
    tc3 = ws3.cell(row=1, column=1)
    tc3.fill = NAVY
    tc3.font = TITLE_FONT
    tc3.alignment = CENTER
    ws3.row_dimensions[1].height = 26
    ws3.column_dimensions["A"].width = 40
    ws3.column_dimensions["B"].width = 14

    same_day = sum(
        1
        for g in groups
        if len(
            set(
                (
                    _parse_date(m.get("reported", "")) or type("", (), {"date": lambda s: None})()
                ).date
                if hasattr(_parse_date(m.get("reported", "")), "date")
                else None
                for m in g.get("members", [])
            )
        )
        == 1
    )
    from collections import Counter

    source_counts = Counter(
        m.get("source", "") for g in groups for m in g.get("members", []) if m.get("source", "")
    )
    fault_counts = Counter(
        (m.get("fault_full") or m.get("fault_orig") or m.get("summary", ""))
        for g in groups
        for m in g.get("members", [])
    )

    # ── decision tallies across all groups ──
    n_conf = n_diff = n_partial = n_unrev = 0
    for _g in groups:
        _wd, _, _pd, _ = _grp_decision(_g.get("members", []))
        if _wd == "duplicate":
            n_conf += 1
        elif _wd == "different":
            n_diff += 1
        elif _pd:
            n_partial += 1
        else:
            n_unrev += 1

    stats = [
        ("إجمالي البلاغات في الملف", sr_count),
        ("إجمالي مجموعات المكررات المحتملة", n_groups),
        ("البلاغات المتأثرة (داخل مجموعات)", n_members),
        None,
        ("— قرارات المراجعة —", None),
        ("مؤكدة كتكرار", n_conf),
        ("مختلفة (ليست تكراراً)", n_diff),
        ("قرارات جزئية", n_partial),
        ("لم تُراجَع بعد", n_unrev),
        None,
        ("مجموعات تحوي 3 بلاغات أو أكثر", sum(1 for g in groups if len(g.get("members", [])) >= 3)),
        (
            "مجموعات بلاغاتها في نفس اليوم",
            sum(
                1
                for g in groups
                if len(
                    set(
                        str(_parse_date(m.get("reported", ""))).split()[0]
                        if _parse_date(m.get("reported", ""))
                        else ""
                        for m in g.get("members", [])
                    )
                )
                == 1
            ),
        ),
        None,
        ("— أكثر 10 أعطال تكراراً —", None),
    ]
    for fault, cnt in fault_counts.most_common(10):
        stats.append((fault[:60] if fault else "—", cnt))
    stats.append(None)
    stats.append(("— توزيع حسب المصدر —", None))
    for src, cnt in source_counts.most_common():
        stats.append((src or "—", cnt))

    for r, item in enumerate(stats, 2):
        if item is None:
            continue
        k, v = item
        ws3.cell(row=r, column=1, value=k).font = BOLD10 if v is None else NORM10
        if v is not None:
            ws3.cell(row=r, column=2, value=v).font = BOLD10

    # ── Save ──────────────────────────────────────────────────────────────
    fname = f"مكررات_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    out_path = _ROOT / fname
    wb.save(str(out_path))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from urllib.parse import quote as _quote

    fname_encoded = _quote(fname.encode("utf-8"))
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{fname_encoded}"},
    )


# ── Background Maximo full scan ───────────────────────────────────────────────

import threading as _threading

_SCAN_STATE: dict = {"running": False, "progress": {}, "error": ""}
_QUICK_SCAN_STATE: dict = {"running": False, "last": {}, "error": ""}

# ── Email notification helper ─────────────────────────────────────────────────


def _send_email_notification(new_groups: list[dict]) -> None:
    """Send an HTML email listing newly-detected duplicate groups.

    Reads SMTP config from CFG (sourced from .env).  Silently logs on error.
    """
    import smtplib, logging as _log
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText

    _elog = _log.getLogger("live_monitor.notify")

    if not CFG.notify_email or not CFG.smtp_host:
        return  # notifications not configured

    recipients = [r.strip() for r in CFG.notify_email.split(",") if r.strip()]
    if not recipients:
        return

    tier_label = {"confirmed": "🔴 مؤكد", "possible": "🟡 محتمل", "weak": "🔵 ضعيف"}
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")

    rows_html = ""
    for g in new_groups[:20]:  # cap at 20 to keep email concise
        tier = tier_label.get(g.get("tier", ""), g.get("tier", ""))
        srs = " · ".join(g.get("srs", []))
        rows_html += (
            f"<tr>"
            f"<td style='padding:6px 10px;border-bottom:1px solid #eee'>{tier}</td>"
            f"<td style='padding:6px 10px;border-bottom:1px solid #eee'>{g.get('score', 0)}</td>"
            f"<td style='padding:6px 10px;border-bottom:1px solid #eee'>{g.get('size', 0)}</td>"
            f"<td style='padding:6px 10px;border-bottom:1px solid #eee;direction:rtl'>{g.get('fault', '')}</td>"
            f"<td style='padding:6px 10px;border-bottom:1px solid #eee'>{g.get('loc', '')}</td>"
            f"<td style='padding:6px 10px;border-bottom:1px solid #eee'>{srs}</td>"
            f"</tr>"
        )

    html = f"""
<html><body dir="rtl" style="font-family:Arial,sans-serif;font-size:13px;color:#222">
<div style="background:#1a1a2e;padding:16px 20px;border-radius:8px 8px 0 0">
  <span style="color:#c8a044;font-weight:800;font-size:16px">منظومة جودة البلاغات — كدانة مالك</span>
  <span style="color:#aaa;font-size:11px;float:left">{now_str}</span>
</div>
<div style="border:1px solid #ddd;border-top:none;padding:16px;border-radius:0 0 8px 8px">
  <h3 style="margin:0 0 12px;color:#c0392b">
    ⚠️ تم رصد {len(new_groups)} مجموعة تكرار جديدة
  </h3>
  <table width="100%" cellspacing="0" style="border-collapse:collapse;font-size:12px">
    <thead>
      <tr style="background:#f5f5f5">
        <th style="padding:8px 10px;text-align:right;border-bottom:2px solid #ddd">الثقة</th>
        <th style="padding:8px 10px;text-align:right;border-bottom:2px solid #ddd">النقاط</th>
        <th style="padding:8px 10px;text-align:right;border-bottom:2px solid #ddd">الحجم</th>
        <th style="padding:8px 10px;text-align:right;border-bottom:2px solid #ddd">العطل</th>
        <th style="padding:8px 10px;text-align:right;border-bottom:2px solid #ddd">الموقع</th>
        <th style="padding:8px 10px;text-align:right;border-bottom:2px solid #ddd">البلاغات</th>
      </tr>
    </thead>
    <tbody>
      {rows_html}
    </tbody>
  </table>
  <p style="margin-top:16px;font-size:11px;color:#666">
    افتح لوحة المراقبة للمراجعة والحكم على المجموعات الجديدة.
  </p>
</div>
</body></html>"""

    subject = f"[كدانة] {len(new_groups)} مجموعة تكرار جديدة — {now_str}"

    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"] = CFG.smtp_from or CFG.smtp_user
        msg["To"] = ", ".join(recipients)
        msg.attach(MIMEText(html, "html", "utf-8"))

        smtp_cls = smtplib.SMTP
        with smtp_cls(CFG.smtp_host, CFG.smtp_port, timeout=15) as server:
            if CFG.smtp_tls:
                server.starttls()
            if CFG.smtp_user and CFG.smtp_pass:
                server.login(CFG.smtp_user, CFG.smtp_pass)
            server.sendmail(msg["From"], recipients, msg.as_string())
        _elog.info("Email notification sent → %s (%d groups)", recipients, len(new_groups))
    except Exception as exc:
        _elog.warning("Email notification failed: %s", exc)


def _bg_quick_scan_maximo():
    global _QUICK_SCAN_STATE
    if _QUICK_SCAN_STATE.get("running"):
        return
    _QUICK_SCAN_STATE.update({"running": True, "error": ""})
    try:
        from duplicate_monitor.maximo_source import MaximoSource
        from duplicate_monitor import scanner as _sc

        src = MaximoSource()
        if not src.configured():
            _QUICK_SCAN_STATE["error"] = "بيانات اعتماد Maximo غير مكوّنة في ملف .env"
            return
        result = _sc.run_quick_scan(src)
        _QUICK_SCAN_STATE["last"] = result
        if result.get("error"):
            _QUICK_SCAN_STATE["error"] = result["error"]
        # Send email for truly-new groups (background thread so scan isn't blocked)
        new_grps = result.get("new_groups") or []
        if new_grps:
            _threading.Thread(
                target=_send_email_notification,
                args=(new_grps,),
                daemon=True,
                name="email_notify",
            ).start()
    except Exception as e:
        _QUICK_SCAN_STATE["error"] = f"{type(e).__name__}: {e}"
    finally:
        _QUICK_SCAN_STATE["running"] = False


def _bg_scan_maximo(force: bool = False, max_days: Optional[int] = None):
    global _SCAN_STATE
    _SCAN_STATE.update(
        {
            "running": True,
            "error": "",
            "progress": {"phase": "جارٍ الاتصال بـ Maximo…", "fetched": 0},
        }
    )
    try:
        from duplicate_monitor.maximo_source import MaximoSource, MaximoSourceError
        from duplicate_monitor import scanner as _sc

        src = MaximoSource()
        if not src.configured():
            _SCAN_STATE["error"] = "بيانات اعتماد Maximo غير مكوّنة في ملف .env"
            return

        # Wrap fetch_all (no status filter) with live progress reporting.
        # _fetch_paginated accepts an on_page callback; we monkey-patch
        # fetch_all to inject it so the UI sees the running count.
        def _on_page(total):
            _SCAN_STATE["progress"] = {
                "phase": "جارٍ جلب البلاغات من Maximo…",
                "fetched": total,
            }

        _orig_fetch_paginated = src._fetch_paginated

        def _tracked_paginated(where, *, label="fetch", on_page=None):
            return _orig_fetch_paginated(
                where,
                label=label,
                on_page=on_page if on_page is not None else _on_page,
            )

        src._fetch_paginated = _tracked_paginated
        _SCAN_STATE["progress"] = {
            "phase": "جارٍ جلب جميع البلاغات من Maximo… (قد يستغرق بضع دقائق)",
            "fetched": 0,
        }
        result = _sc.run_scan(src, force=force, max_days=max_days)
        if result.get("error"):
            _SCAN_STATE["error"] = result["error"]
        else:
            _SCAN_STATE["progress"] = {
                "phase": "اكتمل",
                "fetched": result["sr_count"],
                "groups": result["group_count"],
                "pairs": result["pair_count"],
            }
            # Email for new groups (non-blocking)
            new_grps = result.get("new_groups") or []
            if new_grps:
                _threading.Thread(
                    target=_send_email_notification,
                    args=(new_grps,),
                    daemon=True,
                    name="email_notify_full",
                ).start()
    except Exception as e:
        _SCAN_STATE["error"] = f"{type(e).__name__}: {e}"
    finally:
        _SCAN_STATE["running"] = False


@app.get("/api/new-groups")
def api_new_groups():
    """Return groups detected in the last quick/full scan that are truly new.
    The frontend polls this endpoint to decide whether to show a browser notification.
    """
    last = _QUICK_SCAN_STATE.get("last", {})
    new_grps = last.get("new_groups") or []
    return JSONResponse(
        {
            "new_groups": new_grps,
            "count": len(new_grps),
            "scanned_at": last.get("scanned_at", ""),
        }
    )


@app.get("/api/notify-config")
def api_notify_config():
    """Return whether email notifications are configured (no secrets exposed)."""
    return JSONResponse(
        {
            "email_enabled": bool(CFG.notify_email and CFG.smtp_host),
            "email_recipient": CFG.notify_email if CFG.notify_email else "",
        }
    )


@app.post("/api/scan-maximo")
async def api_scan_maximo(request: Request):
    if _SCAN_STATE.get("running"):
        return JSONResponse({"ok": False, "message": "الفحص جارٍ بالفعل — انتظر اكتماله"})
    if not CFG.has_maximo_credentials:
        return JSONResponse(
            {
                "ok": False,
                "message": "بيانات اعتماد Maximo غير مكوّنة — أضف MAXIMO_BASE_URL و MAXIMO_USER و MAXIMO_PASS في .env",
            }
        )
    # Optional user-chosen comparison window (days). Clamp to 1–30.
    max_days = None
    try:
        body = await request.json()
        if body and body.get("max_days") is not None:
            max_days = max(1, min(30, int(body["max_days"])))
    except Exception:
        max_days = None
    t = _threading.Thread(
        target=lambda: _bg_scan_maximo(force=True, max_days=max_days),
        daemon=True,
        name="manual_maximo_scan",
    )
    t.start()
    msg = "بدأ الفحص في الخلفية"
    if max_days:
        msg += f" (نافذة المقارنة: {max_days} يوم)"
    return JSONResponse({"ok": True, "message": msg})


@app.post("/api/quick-scan-maximo")
def api_quick_scan_maximo():
    if not CFG.has_maximo_credentials:
        return JSONResponse(
            {
                "ok": False,
                "message": "بيانات اعتماد Maximo غير مكوّنة — أضف MAXIMO_BASE_URL و MAXIMO_USER و MAXIMO_PASS في .env",
            }
        )
    try:
        from duplicate_monitor.maximo_source import MaximoSource
        from duplicate_monitor import scanner as _sc

        result = _sc.run_quick_scan(MaximoSource())
        _QUICK_SCAN_STATE["last"] = result
        _QUICK_SCAN_STATE["error"] = result.get("error", "")
        return JSONResponse(
            {
                "ok": not bool(result.get("error")),
                "message": result.get("error") or "تم تحديث أحدث البلاغات من Maximo",
                "result": result,
            }
        )
    except Exception as e:
        return JSONResponse(
            {
                "ok": False,
                "message": f"{type(e).__name__}: {e}",
            }
        )


@app.get("/api/scan-status")
def api_scan_status():
    scan = _load_scan()
    return JSONResponse(
        {
            "running": _SCAN_STATE.get("running", False),
            "quick_running": _QUICK_SCAN_STATE.get("running", False),
            "quick_last": _QUICK_SCAN_STATE.get("last", {}),
            "quick_error": _QUICK_SCAN_STATE.get("error", ""),
            "progress": _SCAN_STATE.get("progress", {}),
            "error": _SCAN_STATE.get("error", ""),
            "has_credentials": CFG.has_maximo_credentials,
            "maximo_url": CFG.maximo_base_url or "",
            "last_scan": {
                "scanned_at": scan.get("scanned_at", "") if scan else "",
                "sr_count": scan.get("sr_count", 0) if scan else 0,
                "group_count": len(scan.get("groups", [])) if scan else 0,
                "source": scan.get("source", "") if scan else "",
                "file_name": scan.get("file_name", "") if scan else "",
            },
        }
    )


# ── File scan (upload path) ───────────────────────────────────────────────────
_UPLOAD_PATH = _HERE / "uploaded.xlsx"


def _run_file_scan(path: Path) -> dict:
    """Read an Excel file, run detect, save the scan pkl. Returns a summary dict."""
    from duplicate_monitor.matching.legacy import detect as _detect, read_file as _read_file

    # `read_file` handles HTML-based .xls exports from Maximo, double-header
    # rows where headers repeat in row 0, and the openpyxl/xlrd engine choice.
    try:
        df = _read_file(str(path))
    except SystemExit as e:
        return {"error": str(e), "ok": False}
    except Exception as e:
        return {"error": f"خطأ في قراءة الملف: {e}", "ok": False}
    try:
        result = _detect(df, min_score=CFG.min_score, max_days=CFG.max_days)
    except SystemExit as e:
        return {"error": str(e), "ok": False}
    except Exception as e:
        return {"error": f"خطأ في كشف المكررات: {e}", "ok": False}
    result["scanned_at"] = datetime.now().isoformat(timespec="seconds")
    result["sr_count"] = len(df)
    result["source"] = "file"
    result["file_name"] = path.name
    CFG.scan_pkl.write_bytes(pickle.dumps(result))
    groups = result.get("groups", [])
    return {
        "ok": True,
        "sr_count": len(df),
        "group_count": len(groups),
        "pair_count": len(result.get("pairs", [])),
        "scanned_at": result["scanned_at"],
        "file_name": path.name,
        "error": "",
    }


@app.post("/api/upload")
async def api_upload(file: UploadFile = File(...)):
    name = file.filename or ""
    if not name.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "يجب رفع ملف Excel (.xlsx أو .xls)")
    content = await file.read()
    _UPLOAD_PATH.write_bytes(content)
    summary = _run_file_scan(_UPLOAD_PATH)
    if not summary.get("ok"):
        raise HTTPException(500, summary.get("error", "خطأ غير محدد"))
    return JSONResponse(summary)


@app.get("/api/reports")
def api_reports():
    """Return all individual SRs from the last scan, with their group membership."""
    scan = _load_scan()
    if not scan:
        return JSONResponse({"ready": False, "srs": [], "total": 0})
    # all_rows contains every SR from the DataFrame (set by detect())
    all_rows = scan.get("all_rows") or []
    # Build sr  group mapping
    sr_to_group: dict = {}
    for i, g in enumerate(scan.get("groups", []), 1):
        members = g.get("members", [])
        score = g.get("score", 0)
        tier = _tier(score)
        for m in members:
            sr = m.get("sr", "")
            if sr:
                sr_to_group[sr] = {"group_num": i, "tier": tier, "score": score}
    # If all_rows is empty (old pkl), fall back to collecting from group members
    if not all_rows:
        seen: dict = {}
        for g in scan.get("groups", []):
            for m in g.get("members", []):
                sr = m.get("sr", "")
                if sr and sr not in seen:
                    seen[sr] = m
        all_rows = list(seen.values())

    def _reported_key(row: dict) -> str:
        return str(row.get("reported", "") or row.get("reported_dt", "") or "")

    all_rows = sorted(all_rows, key=_reported_key, reverse=True)
    result = []
    for row in all_rows:
        sr = str(row.get("sr", ""))
        ginfo = sr_to_group.get(sr, {})
        result.append(
            {
                "sr": sr,
                "loc": str(row.get("loc", "") or ""),
                "asset": str(row.get("asset", "") or ""),
                "fault": str(row.get("fault_orig") or row.get("fault") or "")[:70],
                "status": str(row.get("status", "") or ""),
                "reported": str(row.get("reported", "") or "")[:16],
                "detail": str(row.get("detail", "") or "")[:150],
                "group_num": ginfo.get("group_num"),
                "tier": ginfo.get("tier"),
                "score": ginfo.get("score"),
            }
        )
    return JSONResponse(
        {
            "ready": True,
            "total": len(result),
            "in_groups": len(sr_to_group),
            "srs": result,
        },
        headers={"Cache-Control": "no-store"},
    )


# ── HTML ──────────────────────────────────────────────────────────────────────
@app.get("/", response_class=HTMLResponse)
def index():
    return HTMLResponse(_HTML, headers={"Cache-Control": "no-store"})


_HTML = r"""<!DOCTYPE html>
<html lang="ar" dir="rtl">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title> منظومة جودة البلاغات – مركز مالك </title>
<link href="https://fonts.googleapis.com/css2?family=Cairo:wght@400;500;600;700;800&family=Inter:wght@400;500;600&display=swap" rel="stylesheet">
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css">
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<script src="https://unpkg.com/leaflet.markercluster@1.5.3/dist/leaflet.markercluster.js"></script>
<link rel="stylesheet" href="https://unpkg.com/leaflet.markercluster@1.5.3/dist/MarkerCluster.css">
<link rel="stylesheet" href="https://unpkg.com/leaflet.markercluster@1.5.3/dist/MarkerCluster.Default.css">
<style>
:root{
  --gold:#a9824b;--gold-dk:#6f542f;--gold-lt:#c9a977;
  --gold-bg:#fbf5e9;--gold-bd:#dcc49d;
  --brown:#f5ead8;          /* KIDANA cream - sidebar bg */
  --brown-deep:#e6d3b6;     /* secondary cream */
  --brown-ink:#4f3e25;      /* text on cream */
  --sand:#f7f2ea;--sand2:#efe3d1;
  --border:#dfcfb9;--card:#fffdf9;
  --red:#dc2626;--red-bg:#fef2f2;--red-bd:#fecaca;
  --amber:#d97706;--amb-bg:#fffbeb;--amb-bd:#fcd34d;
  --green:#16a34a;--grn-bg:#f0fdf4;--grn-bd:#86efac;
  --blue:#2563eb;--blu-bg:#eff6ff;--blu-bd:#bfdbfe;
  --txt1:#24211c;--txt2:#51483c;--txt3:#74695a;--txt4:#9b9286;
  --r:8px;--rl:12px;
  --sh:0 1px 4px rgba(0,0,0,.06),0  1px 2px rgba(0,0,0,.04);
  --sh2:0 4px 20px rgba(0,0,0,.10),0 2px 6px rgba(0,0,0,.06);
}
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
html,body{font-family:"Cairo","Segoe UI",sans-serif;background:var(--sand);color:var(--txt1);font-size:13px;height:100%;overflow:hidden}
.shell{display:flex;height:100vh}
/* Sidebar */
.sidebar{width:252px;flex-shrink:0;background:linear-gradient(180deg,#fbf4e8 0%,#eadcc8 100%);display:flex;flex-direction:column;border-left:1px solid var(--border);transition:width .25s ease,opacity .2s ease;overflow:hidden}
.shell.sb-closed .sidebar{width:0;border-left:none;opacity:0;pointer-events:none}
.brand{padding:18px 14px 14px;border-bottom:1px solid var(--border);display:flex;align-items:center;gap:11px}
.brand-icon{
  width:40px;
  height:40px;
  border-radius:9px;
  background:#fff;
  display:flex;
  align-items:center;
  justify-content:center;
  overflow:hidden;
  padding:4px;
  border:1px solid var(--border);
}

.brand-icon img{
  width:100%;
  height:100%;
  object-fit:contain;
}
.brand-name{font-size:11.5px;font-weight:700;color:var(--gold-dk);line-height:1.3}
.brand-sub{font-size:9px;color:var(--txt3);margin-top:2px}
.nav{flex:1;padding:8px;overflow-y:auto}
.nav-item{display:flex;align-items:center;gap:10px;padding:9px 12px;border-radius:7px;cursor:pointer;color:var(--txt2);font-size:12.5px;font-weight:500;border:none;background:none;width:100%;text-align:right;font-family:inherit;transition:all .15s;position:relative;margin-bottom:2px}
.nav-item:hover{background:rgba(181,150,102,.10);color:var(--brown-ink)}
.nav-item.active{background:rgba(181,150,102,.18);color:var(--gold-dk);font-weight:700}
.nav-item.active::after{content:"";position:absolute;right:0;top:25%;height:50%;width:3px;background:var(--gold);border-radius:3px 0 0 3px}
.nav-icon{font-size:14px;width:18px;text-align:center;flex-shrink:0}
.nav-badge{margin-right:auto;background:var(--red);color:#fff;font-size:9px;font-weight:800;padding:1px 5px;border-radius:8px;min-width:16px;text-align:center}
.sidebar-foot{padding:10px 14px;border-top:1px solid var(--border);font-size:10px;color:var(--txt4);text-align:center}
.sb-toggle{width:34px;height:34px;border-radius:7px;border:1px solid var(--border);background:var(--sand);display:flex;align-items:center;justify-content:center;cursor:pointer;font-size:18px;color:var(--gold-dk);transition:all .15s;font-family:inherit;font-weight:700}
.sb-toggle:hover{background:var(--gold-bg);border-color:var(--gold-bd)}
/* Main */
.main-area{flex:1;display:flex;flex-direction:column;min-width:0;overflow:hidden}
/* Topbar */
.topbar{background:var(--card);border-bottom:1px solid var(--border);padding:0 18px;height:64px;flex-shrink:0;display:flex;align-items:center;gap:12px;overflow:visible;position:relative;z-index:10}
.tb-brand{min-width:300px;max-width:420px;display:flex;flex-direction:row;align-items:center;gap:10px;line-height:1.25;flex-shrink:0}
.tb-brand-copy{min-width:0;display:flex;flex-direction:column;gap:2px}
.tb-brand-sup{font-size:10px;font-weight:800;color:var(--gold-dk);white-space:nowrap}
.tb-brand-title{font-size:16px;font-weight:800;color:var(--brown-ink);white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.tb-logo{
  width:88px;
  height:50px;
  object-fit:contain;
  display:block;
  flex-shrink:0;
}
.tb-brand-sub{font-size:10.5px;color:var(--txt3);white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.tb-search{flex:1;max-width:380px;display:flex;align-items:center;gap:8px;background:var(--sand);border:1px solid var(--border);border-radius:8px;padding:6px 12px}
.tb-search input{border:none;background:none;font-family:inherit;font-size:11.5px;color:var(--txt2);width:100%;outline:none}
.tb-search input::placeholder{color:var(--txt4)}
.tb-date{font-size:10.5px;color:var(--txt3)}
.tb-gap{flex:1}
.tb-icon{width:34px;height:34px;border-radius:7px;border:1px solid var(--border);background:var(--sand);display:flex;align-items:center;justify-content:center;cursor:pointer;position:relative;transition:all .15s;font-size:14px;color:var(--txt2)}
.tb-icon:hover{background:var(--sand2);border-color:var(--gold-bd)}
.notif-dot{position:absolute;top:-3px;right:-3px;width:14px;height:14px;border-radius:50%;background:var(--red);color:#fff;font-size:8px;font-weight:800;display:flex;align-items:center;justify-content:center;border:2px solid var(--card)}
.tb-avatar{width:34px;height:34px;border-radius:50%;background:var(--gold);color:#fff;font-size:13px;font-weight:700;display:flex;align-items:center;justify-content:center;cursor:pointer}
/* Pages */
.pages{flex:1;overflow:hidden;position:relative}
.page{display:none;height:100%;overflow-y:auto;flex-direction:column}
.page.active{display:flex}
/* Hero */
.hero{display:none}
.hero-text .hero-sup{color:var(--gold-dk) !important}
.hero-text .hero-title{color:var(--brown-ink) !important}
.hero-text .hero-desc{color:var(--txt2) !important}
.hero::after{content:"";position:absolute;inset:0;pointer-events:none;background:radial-gradient(ellipse at 80% 50%,rgba(185,151,91,.08) 0%,transparent 60%)}
.hero-inner{position:relative;display:flex;align-items:center;gap:20px}
.hero-text{flex:1}
.hero-sup{font-size:10px;font-weight:700;color:var(--gold);letter-spacing:.12em;text-transform:uppercase;margin-bottom:8px}
.hero-title{font-size:22px;font-weight:800;color:#fff;line-height:1.2;margin-bottom:8px}
.hero-desc{font-size:11px;color:rgba(255,255,255,.5);line-height:1.7;max-width:480px}
.hero-btn{padding:10px 20px;border-radius:8px;background:var(--gold);border:none;color:#fff;font-family:inherit;font-size:12px;font-weight:700;cursor:pointer;white-space:nowrap;transition:all .15s;display:inline-flex;align-items:center;gap:6px;flex-shrink:0}
.hero-btn:hover{background:var(--gold-dk)}
.hero-btn:disabled{opacity:.55;cursor:not-allowed}
/* KPI */
.kpi-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;padding:16px 20px 0;flex-shrink:0}
.kpi-grid-6{grid-template-columns:repeat(6,1fr);gap:10px}
.kpi-grid-6 .kpi-card{padding:14px 12px;flex-direction:column;align-items:stretch;gap:6px}
.kpi-v-sm{font-size:13px !important;line-height:1.3 !important;font-weight:700}
.kpi-v-red{color:var(--red)}
.kpi-faults{margin-top:6px;display:flex;flex-direction:column;gap:3px}
.kpi-fault-row{display:flex;align-items:center;gap:6px;font-size:10.5px;color:var(--txt2)}
.kpi-fault-bar{flex:1;height:6px;background:var(--sand2);border-radius:3px;overflow:hidden;min-width:30px}
.kpi-fault-fill{height:100%;background:var(--gold)}
.kpi-fault-n{font-weight:700;color:var(--gold-dk);min-width:18px;text-align:center}
.kpi-fault-name{flex:1;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;font-size:10px}
/* ── Dashboard: full-viewport, no scroll ── */
#pg-dash{overflow:hidden !important}
/* MAP fills all space between topbar and KPI row */
.map-section{flex:1;min-height:0;display:flex;flex-direction:column;overflow:hidden;border-bottom:1px solid var(--border)}
.map-head{padding:8px 14px;border-bottom:1px solid var(--border);display:flex;justify-content:space-between;align-items:center;gap:12px;background:var(--sand);flex-shrink:0}
.map-title{font-size:12.5px;font-weight:800;color:var(--brown-ink)}
.map-sub{font-size:10px;color:var(--txt3)}
.map-actions{display:flex;align-items:center;gap:6px;flex-wrap:wrap;justify-content:flex-end}
.map-zoom-btn{border:1px solid var(--border);background:var(--card);color:var(--txt2);border-radius:6px;padding:4px 9px;font-family:inherit;font-size:10.5px;font-weight:700;cursor:pointer;transition:all .12s}
.map-zoom-btn:hover{border-color:var(--gold);background:var(--gold-bg);color:var(--gold-dk)}
#livemap{flex:1;width:100%;min-height:0}
.map-legend{padding:5px 14px;display:flex;gap:18px;background:var(--sand);border-top:1px solid var(--border);font-size:10.5px;color:var(--txt2);flex-shrink:0}
/* Dashboard KPI cards row (below map, fixed height) */
.dash-kpi-row{display:grid;grid-template-columns:repeat(4,1fr) 1.4fr 1.4fr;border-top:2px solid var(--gold-bd);background:var(--card);flex-shrink:0}
.dkpi-card{padding:10px 16px 8px;border-left:1px solid var(--border)}
.dkpi-card:first-child{border-left:none}
.dkpi-btn{cursor:pointer;transition:background .12s}
.dkpi-btn:hover{background:var(--sand2)}
.dkpi-n{font-size:30px;font-weight:800;line-height:1;margin-bottom:2px}
.dkpi-lbl{font-size:11px;font-weight:700;color:var(--txt2);margin-bottom:2px}
.dkpi-sub{font-size:10px;color:var(--txt4)}
.dkpi-list{display:flex;flex-direction:column;justify-content:center}
.dkpi-toplist{display:flex;flex-direction:column;gap:2px}
.dkpi-topitem{display:flex;align-items:center;gap:6px;font-size:10.5px;cursor:pointer;padding:1px 0;transition:color .1s}
.dkpi-topitem:hover{color:var(--gold-dk)}
.dkpi-topitem-n{font-weight:800;color:var(--red);min-width:18px;text-align:center;font-size:12px}
.dkpi-topitem-lbl{color:var(--txt2);white-space:nowrap;overflow:hidden;text-overflow:ellipsis;max-width:140px}
.map-popup{min-width:190px;font-family:"Cairo","Segoe UI",sans-serif;direction:rtl}
.map-popup .mp-sr{font-size:13px;font-weight:800;color:var(--brown-ink);margin-bottom:4px}
.map-popup .mp-loc{font-size:11px;color:var(--txt2);line-height:1.5;margin-bottom:6px}
.map-popup .mp-fault{font-size:10.5px;color:var(--txt3);line-height:1.5;margin-bottom:8px}
.map-popup .mp-tag{display:inline-flex;align-items:center;padding:2px 7px;border-radius:10px;font-size:10px;font-weight:800;background:var(--red-bg);color:var(--red);border:1px solid var(--red-bd)}
.map-popup .mp-link{display:inline-flex;margin-top:8px;padding:4px 10px;border-radius:6px;background:var(--gold);color:#fff;text-decoration:none;font-size:10.5px;font-weight:700}
.map-popup .mp-dist{font-size:11.5px;font-weight:800;color:#0369a1;margin:6px 0}
.map-popup .mp-goto{display:inline-flex;align-items:center;gap:4px;margin-top:8px;padding:5px 12px;border-radius:6px;background:#0284c7;color:#fff;border:none;font-family:inherit;font-size:10.5px;font-weight:700;cursor:pointer;width:100%}
.map-popup .mp-goto:hover{background:#0369a1}
/* permanent distance label on the line */
.leaflet-tooltip.map-line-dist{background:rgba(255,255,255,.82);border:none!important;box-shadow:0 1px 2px rgba(0,0,0,.18);font-size:9px;font-weight:700;color:#374151;padding:1px 4px;border-radius:4px;pointer-events:none;white-space:nowrap}
.leaflet-tooltip.map-line-dist::before{display:none!important}
.leg-i{display:flex;align-items:center;gap:6px}
.leg-dot{width:10px;height:10px;border-radius:50%;display:inline-block}
.leg-red{background:#dc2626}
.leg-gold{background:var(--gold)}
.kpi-card{background:var(--card);border:1px solid var(--border);border-radius:var(--rl);padding:16px;display:flex;align-items:center;gap:14px;box-shadow:var(--sh);transition:box-shadow .15s}
.kpi-card:hover{box-shadow:var(--sh2)}
.kpi-ico{width:46px;height:46px;border-radius:10px;display:flex;align-items:center;justify-content:center;font-size:20px;flex-shrink:0}
.kpi-ico.blue{background:var(--blu-bg)}.kpi-ico.gold{background:var(--gold-bg)}.kpi-ico.amber{background:var(--amb-bg)}.kpi-ico.red{background:var(--red-bg)}
.kpi-v{font-size:24px;font-weight:800;color:var(--txt1);line-height:1}
.kpi-l{font-size:10.5px;color:var(--txt3);margin-top:4px}
.kpi-s{font-size:9.5px;color:var(--txt4);margin-top:2px}
/* Dash grid */
.dash-grid{display:grid;grid-template-columns:1fr 340px;gap:16px;padding:16px 20px;flex:1}
.dash-col-title{font-size:12px;font-weight:700;color:var(--txt2);margin-bottom:10px;display:flex;align-items:center;gap:6px}
.recent-list{background:var(--card);border:1px solid var(--border);border-radius:var(--rl);padding:0;overflow:hidden}
.ri{display:flex;align-items:center;gap:10px;padding:10px 14px;border-bottom:1px solid var(--border);cursor:pointer;transition:background .1s}
.ri:last-child{border-bottom:none}
.ri:hover{background:var(--sand)}
.rdot{width:7px;height:7px;border-radius:50%;flex-shrink:0}
.rdot.open{background:var(--green)}.rdot.closed{background:var(--txt4)}.rdot.wip{background:var(--amber)}.rdot.other{background:var(--blue)}
.ri-sr{font-size:12px;font-weight:700;color:var(--gold-dk)}
.ri-loc{font-size:10px;color:var(--txt3);margin-top:2px}
.ri-time{font-size:10px;color:var(--txt4);white-space:nowrap;flex-shrink:0}
/* Page header */
.pg-hd{padding:18px 20px 0;display:flex;align-items:flex-start;justify-content:space-between;flex-shrink:0}
.pg-title{font-size:18px;font-weight:800;color:var(--txt1)}
.pg-sub{font-size:11px;color:var(--txt3);margin-top:4px}
/* Stats bar */
.stats-bar{display:flex;gap:10px;padding:14px 20px 0;flex-shrink:0}
.sbox{flex:1;background:var(--card);border:1px solid var(--border);border-radius:var(--r);padding:12px 14px;display:flex;align-items:center;gap:10px}
.sbox-n{font-size:22px;font-weight:800;color:var(--txt1)}
.sbox-l{font-size:11px;color:var(--txt3)}
/* Filter bar */
.fbar{display:flex;align-items:center;gap:8px;padding:12px 20px 0;flex-shrink:0;flex-wrap:wrap}
.fbar select,.fbar input[type=text]{height:32px;border:1px solid var(--border);border-radius:6px;background:var(--card);font-family:inherit;font-size:11.5px;color:var(--txt2);padding:0 10px;outline:none;min-width:120px}
.fbar select:focus,.fbar input:focus{border-color:var(--gold)}
.ftier{height:32px;border:1px solid var(--border);border-radius:6px;background:var(--card);padding:0 12px;font-family:inherit;font-size:11.5px;font-weight:600;cursor:pointer;transition:all .12s}
.ftier.on.confirmed{background:var(--red-bg);border-color:var(--red);color:var(--red)}
.ftier.on.possible{background:var(--amb-bg);border-color:var(--amber);color:var(--amber)}
.ftier.on.weak{background:var(--blu-bg);border-color:var(--blue);color:var(--blue)}
.fbar-gap{flex:1}
.decision-tabs{display:grid;grid-template-columns:repeat(3,minmax(120px,1fr));gap:6px;padding:8px 16px 0;background:var(--sand)}
.decision-tab{border:1px solid var(--border);background:var(--card);border-radius:7px;padding:7px 8px;font-family:inherit;cursor:pointer;text-align:right;color:var(--txt2);display:flex;align-items:center;justify-content:space-between;gap:8px}
.decision-tab:hover{border-color:var(--gold-bd);background:var(--gold-bg)}
.decision-tab.on{border-color:var(--gold);background:var(--gold-bg);color:var(--gold-dk)}
.decision-tab b{font-size:14px;color:var(--txt1)}
.decision-tab span{font-size:10.5px;font-weight:700;white-space:nowrap}
.sf-age{font-size:10px;color:var(--txt4)}
/* Alerts body */
.alerts-body{flex:1;overflow-y:auto;padding:14px 20px}
#pg-alerts .pg-hd{display:none}
#pg-alerts .pg-title{font-size:16px;line-height:1.3}
#pg-alerts .pg-sub{font-size:10px;margin-top:1px}
#pg-alerts .stats-bar{display:none}
#pg-alerts .sbox{padding:7px 10px;border-radius:7px;min-height:40px}
#pg-alerts .sbox-n{font-size:17px;line-height:1}
#pg-alerts .sbox-l{font-size:10.5px}
#pg-alerts .fbar{position:sticky;top:0;z-index:5;background:var(--sand);border-bottom:1px solid var(--border);padding:8px 16px;gap:6px}
#pg-alerts .fbar select,#pg-alerts .fbar input[type=text],#pg-alerts .ftier{height:30px;font-size:11px}
#pg-alerts .alerts-body{padding:0}
/* Group card */
.gcard{background:var(--card);border:none;border-bottom:2px solid var(--border);border-radius:0;margin-bottom:0;overflow:hidden;transition:box-shadow .15s}
.gcard:hover{background:var(--sand)}
.gcard.decided{opacity:.75}
.gcard-head{display:flex;align-items:flex-start;gap:8px;padding:7px 10px;background:linear-gradient(180deg,var(--card) 0%,var(--sand) 100%);border-bottom:1px solid var(--border)}
.gcard-info{flex:1;min-width:0}
.gcard-title{font-size:11px;font-weight:800;color:var(--txt1);letter-spacing:.02em;line-height:1.3;word-break:break-word}
.gcard-fault-full{font-size:10.5px;color:var(--txt2);margin-top:3px;line-height:1.4;word-break:break-word}
.gcard-meta{font-size:9.5px;color:var(--txt3);margin-top:2px;line-height:1.4;display:flex;flex-wrap:wrap;align-items:center;gap:4px}
.gcard-right{display:flex;align-items:center;gap:5px;flex-shrink:0}
.gcard-right .gc-st{order:1}
.gcard-right .sel-toggle{order:2;font-size:10.5px;padding:4px 10px;border-radius:6px;border:1px solid var(--border);background:var(--card);cursor:pointer;color:var(--txt2);font-family:inherit;transition:all .15s}
.gcard-right .sel-toggle:hover{border-color:var(--gold);color:var(--gold-dk)}
.gcard-right .score-ring{order:3}
.gc-pdf-btn{order:2;font-size:9.5px;padding:3px 8px;border-radius:6px;border:1px solid var(--border);background:var(--card);cursor:pointer;color:var(--txt2);font-family:inherit;transition:all .15s}
.gc-pdf-btn:hover{border-color:var(--gold);color:var(--gold-dk);background:var(--sand)}
@media print{
  body > *{display:none!important}
  #print-area{display:block!important;direction:rtl;font-family:'Cairo',sans-serif;font-size:11px}
  .gc-pdf-btn,.dec-dock,.gcard-right button,.sel-toggle,.sr-chk{display:none!important}
  .gcard{border:1px solid #ccc;border-radius:8px;margin:0;page-break-inside:avoid}
}
.gc-st{font-size:10px;font-weight:700;padding:3px 8px;border-radius:12px;white-space:nowrap}
.gc-st.pending{background:var(--amb-bg);color:var(--amber);border:1px solid var(--amb-bd)}
.gc-st.confirmed{background:var(--grn-bg);color:var(--green);border:1px solid var(--grn-bd)}
.gc-st.rejected{background:var(--red-bg);color:var(--red);border:1px solid var(--red-bd)}
.score-ring{width:30px;height:30px;border-radius:50%;border:2px solid var(--border);display:flex;align-items:center;justify-content:center;font-size:9.5px;font-weight:800;color:var(--txt3);flex-shrink:0}
.score-ring.confirmed{color:var(--red);border-color:var(--red);background:var(--red-bg)}
.score-ring.possible{color:var(--amber);border-color:var(--amber);background:var(--amb-bg)}
.score-ring.weak{color:var(--blue);border-color:var(--blue);background:var(--blu-bg)}
/* Toggle selection mode button */
.sel-toggle{width:26px;height:26px;border-radius:6px;border:1px solid var(--border);background:var(--sand);display:flex;align-items:center;justify-content:center;cursor:pointer;font-size:12px;color:var(--txt3);transition:all .12s;flex-shrink:0}
.sel-toggle:hover{background:var(--sand2);color:var(--gold)}
.gcard.sel-mode .sel-toggle{background:var(--gold-bg);border-color:var(--gold);color:var(--gold-dk)}
/* SR pair */
.sr-pair{display:grid;grid-template-columns:1fr 1fr;border-top:1px solid var(--border)}
.sr-col{padding:12px 14px;position:relative}
.sr-col:first-child{border-left:1px solid var(--border)}
.sr-col-lbl{font-size:9.5px;font-weight:700;color:var(--txt4);text-transform:uppercase;letter-spacing:.06em;margin-bottom:6px;display:flex;align-items:center;gap:6px}
/* Checkbox — always visible for selecting members */
.sr-chk{display:inline-block;width:16px;height:16px;accent-color:var(--gold);cursor:pointer;flex-shrink:0}
.sr-num{font-size:18px;font-weight:800;color:var(--gold-dk);cursor:pointer;display:block;margin-bottom:6px}
.sr-num:hover{color:var(--gold)}
.sr-row{display:flex;align-items:flex-start;gap:6px;margin-bottom:4px}
.sr-ri{font-size:11px;width:16px;flex-shrink:0;text-align:center;margin-top:1px}
.sr-rv{font-size:11px;color:var(--txt2);line-height:1.4}
.sr-mx{display:inline-flex;align-items:center;gap:4px;margin-top:8px;font-size:10.5px;color:var(--gold-dk);text-decoration:none;border:1px solid var(--gold-bd);border-radius:5px;padding:3px 8px;background:var(--gold-bg);transition:all .12s}
.sr-mx:hover{background:var(--gold-bd)}
.sr-extra{position:absolute;bottom:8px;right:50%;transform:translateX(50%);background:var(--sand2);border:1px solid var(--border);border-radius:12px;font-size:10px;font-weight:700;color:var(--txt3);padding:2px 10px}
/* ── Comparison TABLE: rows=fields, cols=SRs, horizontal scroll if wide ── */
.cmp-wrap{padding:0;background:var(--sand);max-height:none;overflow-x:auto;overflow-y:hidden;border-top:1px solid var(--border)}
.cmp-tbl{border-collapse:separate;border-spacing:0;width:100%;min-width:500px;font-size:10px;background:var(--card);direction:rtl}
.cmp-tbl thead{position:sticky;top:0;background:var(--card);z-index:2}
.cmp-th-row th{background:linear-gradient(180deg,#715633 0%,#4f3e25 100%);color:#fffaf2;padding:5px 7px;text-align:right;font-weight:700;border-bottom:2px solid var(--gold);white-space:nowrap;position:sticky;top:0}
.cmp-th-lbl{width:90px;min-width:90px;background:#6a5232 !important;color:#f6e6cb !important;font-size:9.5px;letter-spacing:0}
.cmp-col-head{min-width:155px;border-right:1px solid rgba(255,255,255,.08)}
.cmp-col-head.origin{background:linear-gradient(180deg,#7b5a31 0%,#5a4228 100%) !important}
.cmp-col-head-top{display:flex;align-items:center;gap:4px;margin-bottom:2px}
.cmp-sr{color:var(--gold-lt);font-size:11px;font-weight:800;cursor:pointer;letter-spacing:.02em}
.cmp-sr:hover{text-decoration:underline;color:#fff}
.cmp-origin-mini{background:var(--gold);color:#fff;font-size:8.5px;font-weight:700;padding:1px 5px;border-radius:8px;letter-spacing:.04em}
.cmp-mx{display:inline-block;font-size:9px;color:var(--gold-lt);text-decoration:none;padding:1px 6px;border:1px solid rgba(255,255,255,.2);border-radius:4px;background:rgba(0,0,0,.2);font-weight:500}
.cmp-mx:hover{background:var(--gold);color:#fff;border-color:var(--gold)}
.cmp-sec td{background:linear-gradient(90deg,var(--gold-bg) 0%,var(--sand) 100%);color:var(--gold-dk);font-size:9px;font-weight:800;padding:2px 7px;letter-spacing:0;border-top:1px solid var(--gold-bd);border-bottom:1px solid var(--gold-bd);text-transform:uppercase}
.cmp-lbl{background:var(--sand);color:var(--txt2);font-size:9.5px;font-weight:700;padding:4px 7px;border-bottom:1px solid var(--border);border-left:2px solid var(--gold-bd);white-space:nowrap;position:sticky;right:0;z-index:1}
.cmp-cell{padding:4px 7px;border-bottom:1px solid var(--border);border-left:1px solid var(--border);vertical-align:top;color:var(--txt1);font-size:10px;line-height:1.4;word-break:break-word;min-width:155px}
.cmp-cell.origin{background:var(--gold-bg)}
.cmp-cell:last-child{border-left:none}
.cmp-empty{color:var(--txt4);font-style:italic}
.cmp-dist-near{color:#15803d;font-weight:600;font-size:11px}
.cmp-dist-mid {color:#b45309;font-weight:600;font-size:11px}
.cmp-dist-far {color:#dc2626;font-weight:600;font-size:11px}
.cmp-map-btn{border:none;background:none;cursor:pointer;font-size:13px;padding:0 2px;vertical-align:middle;opacity:.65;transition:opacity .12s;line-height:1}
.cmp-map-btn:hover{opacity:1;transform:scale(1.2)}
.cmp-stack{display:flex;flex-direction:column;gap:2px}
.cmp-code{font-weight:700;color:var(--txt1)}
.cmp-sub{font-size:10px;color:var(--txt3);line-height:1.35}
.cmp-origin-pill{background:var(--gold);color:#fff;font-size:9.5px;font-weight:700;padding:1px 8px;border-radius:10px;display:inline-block}
.cmp-gap{background:var(--blu-bg);color:var(--blue);border:1px solid var(--blu-bd);font-size:9.8px;font-weight:700;padding:1px 7px;border-radius:10px;display:inline-block}
.cmp-fault{font-size:10.8px;line-height:1.45;color:var(--txt1);white-space:normal}
.cmp-detail{background:var(--sand);padding:6px 8px;border-right:3px solid var(--gold-lt);border-radius:4px;font-size:10.5px;line-height:1.5;color:var(--txt2);white-space:pre-wrap}
.cmp-mini{display:flex;flex-direction:column;gap:2px}
.cmp-line{display:flex;gap:5px;align-items:baseline;line-height:1.35}
.cmp-line b{color:var(--txt3);font-size:9.8px;min-width:46px}
.cmp-line span{color:var(--txt1)}
.gap-inline{margin-right:6px;display:inline-flex;vertical-align:middle}
.cmp-reporter{margin-top:5px;padding-top:4px;border-top:1px dashed var(--border)}
/* keep srd-list rule for back-compat (now unused) */
.srd-list{display:grid;grid-template-columns:repeat(auto-fit,minmax(300px,1fr));gap:12px;padding:14px 16px;background:var(--sand);align-items:start}
.srd-card{padding:14px 16px;background:var(--card);border:1px solid var(--border);border-radius:10px;box-shadow:var(--sh);transition:box-shadow .15s;display:flex;flex-direction:column;min-width:0}
.srd-card:hover{box-shadow:var(--sh2)}
.srd-card.origin{border-color:var(--gold);background:linear-gradient(180deg,var(--gold-bg) 0%,var(--card) 40%);position:relative}
.srd-card.origin::before{content:"";position:absolute;top:0;right:0;width:4px;height:100%;background:var(--gold);border-radius:0 10px 10px 0}
.srd-head{display:flex;align-items:center;gap:10px;margin-bottom:10px;flex-wrap:wrap;padding-bottom:10px;border-bottom:1px solid var(--border)}
.sr-origin-tag{background:var(--gold);color:#fff;font-size:9.5px;font-weight:700;padding:3px 10px;border-radius:10px;letter-spacing:.05em}
.sr-seq{background:var(--sand2);color:var(--txt2);font-size:10px;font-weight:700;padding:2px 8px;border-radius:10px;border:1px solid var(--border)}
.srd-sr{font-size:14px;font-weight:800;color:var(--gold-dk);cursor:pointer;letter-spacing:.02em}
.srd-sr:hover{text-decoration:underline}
.sr-gap{background:var(--blu-bg);color:var(--blue);border:1px solid var(--blu-bd);font-size:10px;font-weight:700;padding:3px 10px;border-radius:10px}
.srd-mx{margin-right:auto;font-size:10.5px;color:var(--gold-dk);text-decoration:none;padding:4px 10px;border:1px solid var(--gold-bd);border-radius:6px;background:var(--gold-bg);font-weight:600;transition:all .15s}
.srd-mx:hover{background:var(--gold);color:#fff}
/* vertical label:value lines — single column inside each card */
.srd-lines{display:flex;flex-direction:column;gap:0}
.srd-line{display:grid;grid-template-columns:95px 1fr;gap:10px;align-items:start;padding:6px 0;border-bottom:1px dashed transparent;transition:border-color .15s}
.srd-line:not(:last-child){border-bottom:1px dashed var(--border)}
.srd-line-full{}
.srd-l{font-size:10.5px;font-weight:700;color:var(--txt3);letter-spacing:.02em;padding-top:2px}
.srd-v{font-size:12px;color:var(--txt1);font-weight:500;word-break:break-word;line-height:1.6}
.srd-empty{color:var(--txt4);font-weight:400}
.srd-detail{margin-top:12px;padding-top:10px;border-top:1px dashed var(--border)}
.srd-detail .srd-l{display:block;margin-bottom:6px}
.srd-detail-txt{font-size:11.5px;color:var(--txt2);line-height:1.8;word-break:break-word;background:var(--sand);padding:10px 12px;border-radius:6px;border-right:3px solid var(--gold-lt)}
.tier-pill{display:inline-block;font-size:10px;font-weight:700;padding:3px 10px;border-radius:10px}
.tier-pill.confirmed{background:var(--red-bg);color:var(--red);border:1px solid var(--red-bd)}
.tier-pill.possible{background:var(--amb-bg);color:var(--amber);border:1px solid var(--amb-bd)}
.tier-pill.weak{background:var(--blu-bg);color:var(--blue);border:1px solid var(--blu-bd)}
/* Sub-decisions list */
.sub-dec-list{border-top:1px solid var(--border);padding:8px 14px;display:flex;flex-direction:column;gap:4px;background:var(--sand)}
.sub-dec-item{display:flex;align-items:center;gap:6px;font-size:10.5px;padding:4px 8px;border-radius:5px;border:1px solid var(--border)}
.sub-dec-item.confirmed{background:var(--grn-bg);border-color:var(--grn-bd);color:var(--green)}
.sub-dec-item.rejected{background:var(--red-bg);border-color:var(--red-bd);color:var(--red)}
.sub-dec-txt{flex:1}
.sub-dec-del{border:none;background:none;cursor:pointer;font-size:11px;color:inherit;opacity:.6;padding:0 2px;line-height:1}
.sub-dec-del:hover{opacity:1}
/* Decision banner */
.dec-banner{padding:8px 14px;font-size:11.5px;font-weight:600;border-top:1px solid var(--border)}
.dec-banner.confirmed{background:var(--grn-bg);color:var(--green)}
.dec-banner.rejected{background:var(--red-bg);color:var(--red)}
/* Decision dock */
.dec-dock{padding:6px 10px;border-top:1px solid var(--border);background:var(--sand);display:grid;grid-template-columns:minmax(160px,1fr) auto;gap:6px;align-items:center}
.note-inp{width:100%;border:1px solid var(--border);border-radius:6px;background:var(--card);font-family:inherit;font-size:10.5px;color:var(--txt2);padding:5px 10px;outline:none;transition:border .12s}
.note-inp:focus{border-color:var(--gold)}
.dec-hint{display:none}
.dec-hint-i{width:16px;height:16px;border-radius:50%;background:var(--gold-bg);color:var(--gold-dk);font-size:10px;font-weight:800;display:inline-flex;align-items:center;justify-content:center;font-style:italic;font-family:Georgia,serif}
.dec-actions{display:flex;gap:5px;flex-wrap:wrap;justify-content:flex-end}
.dec-grp{display:flex;align-items:center;gap:5px;padding:4px 6px;background:var(--card);border:1px solid var(--border);border-radius:6px;flex-wrap:wrap}
.dec-grp-sel{border-color:var(--gold-bd);background:var(--gold-bg)}
.dec-grp-lbl{font-size:9.5px;font-weight:700;color:var(--txt2);margin-left:auto}
.dec-grp-sel.disabled{opacity:.45;pointer-events:none}
.btn-confirm{padding:5px 12px;border:none;border-radius:5px;font-family:inherit;font-size:10.5px;font-weight:700;cursor:pointer;background:var(--green);color:#fff;transition:all .12s}
.btn-confirm:hover{filter:brightness(1.1)}
.btn-reject{padding:5px 12px;border:none;border-radius:5px;font-family:inherit;font-size:10.5px;font-weight:700;cursor:pointer;background:var(--red-bg);color:var(--red);border:1px solid var(--red-bd);transition:all .12s}
.btn-reject:hover{background:var(--red);color:#fff}
.btn-sel-confirm{padding:6px 14px;border:none;border-radius:6px;font-family:inherit;font-size:11.5px;font-weight:700;cursor:pointer;background:var(--green);color:#fff}
.btn-sel-reject{padding:6px 14px;border:none;border-radius:6px;font-family:inherit;font-size:11.5px;font-weight:700;cursor:pointer;background:var(--red-bg);color:var(--red);border:1px solid var(--red-bd)}
.btn-sel-cancel{padding:6px 10px;border:1px solid var(--border);border-radius:6px;font-family:inherit;font-size:11px;color:var(--txt3);cursor:pointer;background:var(--card)}
.btn-undo{padding:5px 10px;border:1px solid var(--gold-bd);border-radius:5px;font-family:inherit;font-size:10px;font-weight:700;cursor:pointer;background:var(--gold-bg);color:var(--gold-dk)}
.btn-undo:hover{background:var(--gold);color:#fff}
/* Status badges */
.st-open{background:var(--grn-bg);color:var(--green);border:1px solid var(--grn-bd);padding:1px 6px;border-radius:8px;font-size:9.5px;font-weight:700}
.st-closed{background:var(--sand2);color:var(--txt3);border:1px solid var(--border);padding:1px 6px;border-radius:8px;font-size:9.5px;font-weight:700}
.st-wip{background:var(--amb-bg);color:var(--amber);border:1px solid var(--amb-bd);padding:1px 6px;border-radius:8px;font-size:9.5px;font-weight:700}
.st-other{background:var(--blu-bg);color:var(--blue);border:1px solid var(--blu-bd);padding:1px 6px;border-radius:8px;font-size:9.5px;font-weight:700}
/* Analytics */
.ana-wrap{padding:16px 20px;flex:1;overflow-y:auto}
.ana-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:14px}
.ana-card{background:var(--card);border:1px solid var(--border);border-radius:var(--rl);padding:16px}
.ana-t{font-size:12px;font-weight:700;color:var(--txt2);margin-bottom:12px}
.spills{display:flex;gap:8px;flex-wrap:wrap}
.spill{flex:1;min-width:60px;background:var(--sand);border:1px solid var(--border);border-radius:var(--r);padding:10px 8px;text-align:center}
.spill.red{background:var(--red-bg);border-color:var(--red-bd)}.spill.amber{background:var(--amb-bg);border-color:var(--amb-bd)}.spill.blue{background:var(--blu-bg);border-color:var(--blu-bd)}.spill.green{background:var(--grn-bg);border-color:var(--grn-bd)}
.spill-n{display:block;font-size:22px;font-weight:800;color:var(--txt1)}
.spill-l{font-size:10px;color:var(--txt3);margin-top:4px}
.bar-row{display:flex;align-items:center;gap:8px;margin-bottom:6px}
.bar-lbl{width:120px;font-size:10.5px;color:var(--txt2);text-overflow:ellipsis;overflow:hidden;white-space:nowrap;flex-shrink:0;direction:rtl}
.bar-out{flex:1;height:8px;background:var(--sand2);border-radius:4px;overflow:hidden}
.bar-in{height:100%;background:var(--gold);border-radius:4px;transition:width .4s}
.bar-n{width:24px;text-align:left;font-size:10px;color:var(--txt3);flex-shrink:0}
/* Reports page */
.rpt-bar{display:flex;align-items:center;gap:8px;padding:12px 20px 0;flex-shrink:0;flex-wrap:wrap}
.rpt-bar input,.rpt-bar select{height:32px;border:1px solid var(--border);border-radius:6px;background:var(--card);font-family:inherit;font-size:11.5px;color:var(--txt2);padding:0 10px;outline:none}
.rpt-bar input{min-width:220px}
.rpt-bar select{min-width:120px}
.rpt-bar input:focus,.rpt-bar select:focus{border-color:var(--gold)}
.rpt-wrap{flex:1;overflow-y:auto;padding:12px 20px 20px;margin-top:0}
.rpt-tbl{width:100%;border-collapse:collapse;background:var(--card);border:1px solid var(--border);border-radius:var(--rl);overflow:hidden;font-size:11.5px}
.rpt-tbl thead tr{background:#5a4228;color:#f6e6cb}
.rpt-tbl th{padding:10px 12px;font-weight:700;font-size:11px;text-align:right;white-space:nowrap}
.rpt-tbl td{padding:8px 12px;border-bottom:1px solid var(--border);vertical-align:middle}
.rpt-tbl tbody tr:hover{background:var(--sand)}
.rpt-tbl tbody tr:last-child td{border-bottom:none}
.rpt-sr-num{font-weight:700;color:var(--gold-dk);cursor:pointer}
.rpt-sr-num:hover{color:var(--gold);text-decoration:underline}
.rpt-fault{max-width:200px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.rpt-grp-badge{display:inline-flex;align-items:center;padding:2px 7px;border-radius:10px;font-size:9.5px;font-weight:700}
.rpt-grp-badge.confirmed{background:var(--red-bg);color:var(--red);border:1px solid var(--red-bd)}
.rpt-grp-badge.possible{background:var(--amb-bg);color:var(--amber);border:1px solid var(--amb-bd)}
.rpt-grp-badge.weak{background:var(--blu-bg);color:var(--blue);border:1px solid var(--blu-bd)}
.rpt-more-btn{display:flex;justify-content:center;padding:12px}
/* Upload / Data source page */
.upd-src-card{background:var(--card);border:2px solid var(--gold-bd);border-radius:var(--rl);padding:18px;margin-bottom:0}
.upd-body{padding:20px;max-width:680px;margin:0 auto;flex:1;overflow-y:auto}
.drop-zone{border:2px dashed var(--gold-bd);border-radius:var(--rl);padding:48px 24px;text-align:center;cursor:pointer;background:var(--gold-bg);transition:all .15s;margin-bottom:16px}
.drop-zone:hover,.drop-zone.drag-over{background:var(--amb-bg);border-color:var(--amber)}
.dz-icon{font-size:36px;margin-bottom:12px}
.dz-text{font-size:14px;font-weight:700;color:var(--txt1);margin-bottom:6px}
.dz-hint{font-size:11px;color:var(--txt3)}
.upd-prog{background:var(--card);border:1px solid var(--border);border-radius:var(--r);padding:16px;margin-bottom:16px}
.prog-bar{height:8px;background:var(--sand2);border-radius:4px;overflow:hidden;margin-bottom:8px}
.prog-fill{height:100%;background:linear-gradient(90deg,var(--gold),var(--amber));border-radius:4px;width:0;transition:width .3s}
#prog-txt{font-size:11.5px;color:var(--txt3);text-align:center}
.upd-result{background:var(--grn-bg);border:1px solid var(--grn-bd);border-radius:var(--r);padding:16px;margin-bottom:16px;text-align:center}
.upd-ok{font-size:16px;font-weight:800;color:var(--green);margin-bottom:8px}
.upd-stats{font-size:12px;color:var(--txt2)}
.upd-last{background:var(--card);border:1px solid var(--border);border-radius:var(--r);padding:16px}
.upd-last-t{font-size:12px;font-weight:700;color:var(--txt2);margin-bottom:10px}
.upd-last-row{display:flex;justify-content:space-between;font-size:11.5px;color:var(--txt3);padding:4px 0;border-bottom:1px solid var(--border)}
.upd-last-row:last-child{border-bottom:none}
.upd-last-row b{color:var(--txt1)}
/* SR Modal */
.modal-bg{display:none;position:fixed;inset:0;background:rgba(0,0,0,.35);z-index:100;align-items:center;justify-content:center}
.modal-bg.open{display:flex}
.modal{background:var(--card);border-radius:var(--rl);width:560px;max-width:95vw;max-height:85vh;display:flex;flex-direction:column;box-shadow:0 20px 60px rgba(0,0,0,.25)}
.modal-hd{padding:18px 20px;border-bottom:1px solid var(--border);display:flex;align-items:center;justify-content:space-between}
.modal-title{font-size:14px;font-weight:800;color:var(--txt1)}
.modal-close{width:28px;height:28px;border:none;background:var(--sand);border-radius:6px;cursor:pointer;font-size:16px;display:flex;align-items:center;justify-content:center;color:var(--txt3)}
.modal-close:hover{background:var(--sand2)}
.modal-body{padding:18px 20px;overflow-y:auto;flex:1}
.mfg{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:14px}
.mf{background:var(--sand);border-radius:var(--r);padding:10px 12px}
.mf.wide{grid-column:span 2}
.mf-l{font-size:9.5px;font-weight:700;color:var(--txt4);text-transform:uppercase;letter-spacing:.04em;margin-bottom:4px}
.mf-v{font-size:12.5px;color:var(--txt1)}
.mx-btn{display:inline-flex;align-items:center;gap:6px;padding:8px 16px;border-radius:7px;background:var(--gold);color:#fff;font-family:inherit;font-size:12px;font-weight:700;text-decoration:none;margin-top:8px;transition:all .12s}
.mx-btn:hover{background:var(--gold-dk)}
/* Toast */
#toasts{position:fixed;top:72px;left:88px;z-index:200;display:flex;flex-direction:column;gap:8px}
.toast{padding:10px 16px;border-radius:8px;font-size:12px;font-weight:600;color:#fff;background:#1a1a1a;display:flex;align-items:center;gap:8px;box-shadow:var(--sh2);animation:tslide .25s ease-out;max-width:300px}
.toast.ok{background:var(--green)}.toast.err{background:var(--red)}.toast.alert{background:var(--amber);color:#1a1a1a}
.ti{font-size:14px}
@keyframes tslide{from{transform:translateY(-10px);opacity:0}to{transform:translateY(0);opacity:1}}
.notif-wrap{position:static}
.notif-panel{display:none;position:fixed;top:68px;width:340px;max-width:calc(100vw - 24px);background:var(--card);border:1px solid var(--border);border-radius:10px;box-shadow:0 8px 32px rgba(0,0,0,.22);z-index:99999;overflow:hidden}
.notif-panel.open{display:block;animation:tslide .15s ease}
.notif-head{padding:10px 12px;border-bottom:1px solid var(--border);display:flex;align-items:center;justify-content:space-between;gap:10px;background:var(--sand)}
.notif-title{font-size:12px;font-weight:800;color:var(--txt1)}
.notif-total{font-size:10px;font-weight:800;color:#fff;background:var(--red);border-radius:10px;padding:1px 7px}
.notif-list{max-height:300px;overflow:auto}
.notif-item{padding:10px 12px;border-bottom:1px solid var(--border);cursor:pointer;background:var(--card)}
.notif-item:hover{background:var(--sand)}
.notif-item:last-child{border-bottom:none}
.notif-item-sr{font-size:12px;font-weight:800;color:var(--gold-dk);margin-bottom:3px}
.notif-item-meta{font-size:10.5px;color:var(--txt3);line-height:1.5}
.notif-empty{padding:18px 12px;text-align:center;color:var(--txt3);font-size:11.5px}
@media (max-width:1100px){
  .tb-brand{min-width:180px;max-width:260px}
  .tb-brand-sub,.tb-date{display:none}
  .tb-search{max-width:260px}
}
@media (max-width:760px){
  .topbar{height:auto;min-height:64px;flex-wrap:wrap;padding:8px 12px}
  .tb-brand{order:2;min-width:calc(100% - 46px);max-width:none}
  .tb-search{order:5;flex-basis:100%;max-width:none}
  #livemap{height:calc(100vh - 260px);min-height:300px}
  #toasts{top:82px;left:12px;right:12px}
}
/* Empty */
.empty{display:flex;flex-direction:column;align-items:center;justify-content:center;padding:48px 24px;text-align:center;gap:12px;flex:1}
.ei{font-size:40px}
.empty h3{font-size:15px;font-weight:700;color:var(--txt2)}
.empty p{font-size:12px;color:var(--txt4);max-width:320px;line-height:1.6}
</style>
</head>
<body>
<div class="shell">
  <!-- SIDEBAR (RTL: first child = right side) -->
  <nav class="sidebar">
    <div class="brand">
      <div class="brand-icon">
        <img src="/static/kidana-logo.png" alt="Kidana Logo">
      </div>
      <div>
        <div class="brand-name">كدانة مالك</div>
        <div class="brand-sub">مركز الجودة</div>
      </div>
    </div>
    <div class="nav">
      <button class="nav-item active" data-page="dash"     onclick="goto('dash')">    <span class="nav-icon"></span>لوحة التحكم</button>
      <button class="nav-item"        data-page="alerts"   onclick="goto('alerts')">  <span class="nav-icon"></span>مراجعة المكررات<span class="nav-badge" id="nb" style="display:none">0</span></button>
      <button class="nav-item"        data-page="reports"  onclick="goto('reports');loadReports()"><span class="nav-icon"></span>البلاغات</button>
      <button class="nav-item"        data-page="upload"   onclick="goto('upload');loadUploadInfo()"><span class="nav-icon"></span>رفع البلاغات</button>
    </div>
    <div class="sidebar-foot">v2.0 · كدانه</div>
  </nav>

  <!-- MAIN -->
  <div class="main-area">
    <!-- Topbar -->
    <div class="topbar">
      <button class="sb-toggle" onclick="toggleSidebar()" title="إخفاء/إظهار القائمة">☰</button>
      <div class="tb-brand">
        <img class="tb-logo" src="/static/kidana-logo.png" alt="Kidana Logo">
        <div class="tb-brand-copy">
          <div class="tb-brand-sup">منظومة جودة البلاغات</div>
          <div class="tb-brand-title">مركز مالك — كشف البلاغات المكررة</div>
          <div class="tb-brand-sub">رصد بلاغات ماكسيمو وكشف التكرارات لحظياً</div>
        </div>
      </div>
      <span id="tb-scan-age" style="font-size:10px;color:var(--txt4)"></span>
      <button class="hero-btn" id="btn-topbar-refresh" onclick="topbarRefresh()" style="font-size:11px;padding:7px 14px">تحديث الآن</button>
      <div class="tb-search">
        <span></span>
        <input id="srch" placeholder="بحث برقم SR أو الموقع..." oninput="onSearch()">
      </div>
      <span id="tb-date" class="tb-date"></span>
      <div class="tb-gap"></div>
      <div class="notif-wrap">
        <button class="tb-icon" onclick="toggleNotifications(event)" title="التنبيهات" type="button">
          <span>🔔</span>
          <span class="notif-dot" id="nb2" style="display:none">0</span>
        </button>
      </div>
      <button class="hero-btn" id="btn-exp" onclick="exportExcel()" style="font-size:11px;padding:6px 14px"> تصدير Excel</button>
      <div class="tb-avatar" title="المستخدم">م</div>
    </div>

    <!-- hidden stubs — ks-* IDs kept for JS compatibility, not displayed -->
    <div style="display:none" aria-hidden="true">
      <span id="ks-groups"></span><span id="ks-decided"></span>
      <span id="ks-confirmed"></span><span id="ks-progress"></span>
      <span id="ks-total"></span><span id="ks-possible"></span>
    </div>

    <!-- Pages -->
    <div class="pages">

      <!-- ① DASHBOARD — fills 100% viewport, no scroll -->
      <div id="pg-dash" class="page active">
        <!-- MAP — flex:1 fills all remaining height -->
        <div class="map-section">
          <div class="map-head">
            <div>
              <div class="map-title">المشهد الجغرافي للبلاغات</div>
              <div class="map-sub" id="map-sub">جارٍ التحميل…</div>
            </div>
            <div class="map-actions">
              <button class="map-zoom-btn" type="button" onclick="zoomMashair('all')" title="عرض نطاق المشاعر">كل المشاعر</button>
              <button class="map-zoom-btn" type="button" onclick="zoomMashair('mina')" title="تكبير على مشعر منى">منى</button>
              <button class="map-zoom-btn" type="button" onclick="zoomMashair('muzdalifah')" title="تكبير على مشعر مزدلفة">مزدلفة</button>
              <button class="map-zoom-btn" type="button" onclick="zoomMashair('arafat')" title="تكبير على مشعر عرفات">عرفات</button>
              <span class="leg-i"><span class="leg-dot leg-red"></span><span style="font-size:10px;color:var(--txt3)">مواقع فيها مكررات</span></span>
              <span class="leg-i"><span class="leg-dot leg-gold"></span><span style="font-size:10px;color:var(--txt3)">مواقع عادية</span></span>
              <span class="leg-i"><span style="display:inline-block;width:22px;height:0;border-top:2px dashed #dc2626;margin-left:3px;vertical-align:middle;opacity:.75"></span><span style="font-size:10px;color:var(--txt3)">شبكة مكررات (انقر للتركيز)</span></span>
            </div>
          </div>
          <div id="livemap"></div>
        </div>

        <!-- KPI CARDS — interactive strip at bottom -->
        <div class="dash-kpi-row">
          <div class="dkpi-card dkpi-btn" onclick="goto('alerts')" title="كل المجموعات">
            <div class="dkpi-n" id="k-groups" style="color:var(--txt1)">—</div>
            <div class="dkpi-lbl">مجموعة مكررة</div>
            <div class="dkpi-sub" id="k-groups-s">— بلاغ متكرر</div>
          </div>
          <div class="dkpi-card dkpi-btn" onclick="goto('alerts');document.getElementById('fl-status').value='pending';renderAlerts()" title="المجموعات غير المراجَعة">
            <div class="dkpi-n" id="k-progress" style="color:var(--amber)">—</div>
            <div class="dkpi-lbl">معلقة</div>
            <div class="dkpi-sub" id="k-progress-s">لم تُراجَع بعد</div>
          </div>
          <div class="dkpi-card dkpi-btn" onclick="goto('alerts');document.getElementById('fl-status').value='duplicate';renderAlerts()" title="تكرارات محققة">
            <div class="dkpi-n" id="k-confirmed" style="color:var(--green)">—</div>
            <div class="dkpi-lbl">تكرار محقق</div>
            <div class="dkpi-sub" id="k-confirmed-s">جاهزة للإغلاق</div>
          </div>
          <div class="dkpi-card dkpi-btn" onclick="goto('alerts');document.getElementById('fl-status').value='different';renderAlerts()" title="بلاغات مختلفة">
            <div class="dkpi-n" id="k-different" style="color:var(--txt2)">—</div>
            <div class="dkpi-lbl">مختلفة</div>
            <div class="dkpi-sub">بلاغات غير مكررة</div>
          </div>
          <div class="dkpi-card dkpi-list">
            <div class="dkpi-lbl" style="margin-bottom:5px">أكثر الأعطال تكراراً</div>
            <div id="k-top-faults-dash" class="dkpi-toplist"></div>
          </div>
          <div class="dkpi-card dkpi-list" style="border-left:none">
            <div class="dkpi-lbl" style="margin-bottom:5px">أكثر المواقع تكراراً</div>
            <div id="k-top-locs-dash" class="dkpi-toplist"></div>
          </div>
        </div>

        <!-- hidden stubs -->
        <div style="display:none" aria-hidden="true">
          <span id="k-top-contractor"></span><span id="k-top-contractor-s"></span>
          <span id="k-top-faults"></span><span id="k-different-stub"></span>
          <span id="k-confirmed-stub"></span><span id="k-progress-stub"></span>
          <span id="k-groups-stub"></span>
          <div id="dash-grps" style="display:none"></div>
          <div id="recent-list" style="display:none"></div>
        </div>
      </div>

      <!-- ② LIVE ALERTS -->
      <div id="pg-alerts" class="page">
        <div class="pg-hd">
          <div>
            <div class="pg-title"> مراجعة المكررات <span style="background:var(--green);color:#fff;font-size:10px;font-weight:800;padding:2px 8px;border-radius:10px;vertical-align:middle">مباشر</span></div>
            <div class="pg-sub" id="sf-age">—</div>
          </div>
        </div>
        <div class="stats-bar">
          <div class="sbox"><span class="sbox-n" id="sb-pend">—</span><div><div class="sbox-l">معلقة</div></div></div>
          <div class="sbox"><span class="sbox-n" id="sb-conf">—</span><div><div class="sbox-l"> مؤكدة</div></div></div>
          <div class="sbox"><span class="sbox-n" id="sb-rej">—</span><div><div class="sbox-l"> مرفوضة</div></div></div>
        </div>
        <!-- progress bar -->
        <div id="review-prog-bar" style="padding:8px 16px 0;background:var(--sand);border-bottom:1px solid var(--border);flex-shrink:0">
          <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:5px">
            <span style="font-size:11px;font-weight:700;color:var(--txt2)">تقدم المراجعة</span>
            <span id="review-prog-lbl" style="font-size:11px;font-weight:800;color:var(--gold-dk)">0 / 0</span>
          </div>
          <div style="height:8px;background:var(--sand2);border-radius:6px;overflow:hidden;margin-bottom:6px">
            <div id="review-prog-fill" style="height:100%;background:linear-gradient(90deg,var(--gold),var(--gold-dk));border-radius:6px;width:0%;transition:width .4s ease"></div>
          </div>
          <div style="display:flex;gap:10px;font-size:10px;color:var(--txt3);margin-bottom:6px">
            <span> <span id="rpb-confirmed" style="font-weight:800;color:var(--green)">0</span> تكرار محقق</span>
            <span> <span id="rpb-different" style="font-weight:800;color:var(--txt2)">0</span> مختلفة</span>
            <span> <span id="rpb-pending"   style="font-weight:800;color:var(--amber)">0</span> معلقة</span>
          </div>
        </div>
        <div class="decision-tabs" id="decision-tabs">
          <button class="decision-tab on" data-status="pending" onclick="setStatusFilter('pending')"><span>الأساسية</span><b id="dt-pending">0</b></button>
          <button class="decision-tab" data-status="duplicate" onclick="setStatusFilter('duplicate')"><span>مؤكدة</span><b id="dt-duplicate">0</b></button>
          <button class="decision-tab" data-status="different" onclick="setStatusFilter('different')"><span>مستبعدة</span><b id="dt-different">0</b></button>
        </div>
        <div class="fbar">
          <select id="fl-status" onchange="renderAlerts()" style="display:none">
            <option value="">كل المجموعات</option>
            <option value="pending" selected>الأساسية</option>
            <option value="duplicate">مؤكدة كتكرار</option>
            <option value="different">مستبعدة</option>
          </select>
          <select id="fl-loc"   onchange="renderAlerts()"><option value="">كل المواقع</option></select>
          <select id="fl-fault" onchange="renderAlerts()"><option value="">كل الأعطال</option></select>
          <select id="fl-time"  onchange="renderAlerts()" title="الفترة الزمنية">
            <option value="" selected>كل الفترات</option>
            <option value="24">آخر ٢٤ ساعة</option>
            <option value="48">آخر ٤٨ ساعة</option>
            <option value="168">آخر أسبوع</option>
          </select>
          <select id="fl-sort"  onchange="renderAlerts()" title="ترتيب">
            <option value="newest">الأحدث أولاً</option>
            <option value="oldest">الأقدم أولاً</option>
            <option value="score">الأعلى نقاطاً</option>
          </select>
          <div class="fbar-gap"></div>
          <span class="sf-age" id="sf-age2"></span>
        </div>
        <div class="alerts-body" id="alerts-list"></div>
      </div>

      <!-- ③ REPORTS -->
      <div id="pg-reports" class="page">
        <div class="pg-hd">
          <div>
            <div class="pg-title"> البلاغات</div>
            <div class="pg-sub" id="rpt-count">—</div>
          </div>
          <button class="hero-btn" style="font-size:11px;padding:6px 14px" onclick="exportExcel()"> تصدير Excel</button>
        </div>
        <div class="rpt-bar">
          <input type="text" id="rpt-q" placeholder=" بحث برقم SR أو الموقع أو الأصل..." oninput="filterReports()">
          <select id="rpt-status" onchange="filterReports()"><option value="">كل الحالات</option></select>
          <select id="rpt-loc"    onchange="filterReports()"><option value="">كل المواقع</option></select>
          <select id="rpt-grp"    onchange="filterReports()">
            <option value="">كل السجلات</option>
            <option value="in_group">في مجموعة مكررة</option>
            <option value="no_group">بدون مجموعة</option>
          </select>
        </div>
        <div class="rpt-wrap">
          <table class="rpt-tbl">
            <thead><tr>
              <th>رقم SR</th><th>الحالة</th><th>الموقع</th><th>الأصل</th><th>نوع العطل</th><th>تاريخ الفتح</th><th>المجموعة</th>
            </tr></thead>
            <tbody id="rpt-tbody"></tbody>
          </table>
          <div class="rpt-more-btn" id="rpt-more" style="display:none">
            <button class="hero-btn" onclick="rptLoadMore()"> تحميل المزيد</button>
          </div>
        </div>
      </div>

      <!-- ⑤ DATA SOURCE -->
      <div id="pg-upload" class="page">
        <div class="pg-hd">
          <div>
            <div class="pg-title"> مصدر البيانات</div>
            <div class="pg-sub">سحب البلاغات من Maximo أو رفع ملف Excel</div>
          </div>
        </div>
        <div class="upd-body">
          <!-- ── Maximo (primary) ── -->
          <div class="upd-src-card" id="mx-card">
            <div style="display:flex;align-items:center;gap:12px;margin-bottom:14px">
              <div style="width:44px;height:44px;border-radius:10px;background:var(--gold-bg);border:1px solid var(--gold-bd);display:flex;align-items:center;justify-content:center;font-size:22px;flex-shrink:0"></div>
              <div>
                <div style="font-size:14px;font-weight:800;color:var(--txt1)">سحب من Maximo</div>
                <div style="font-size:11px;color:var(--txt3);margin-top:2px">جلب كل البلاغات المفتوحة مباشرة ثم كشف التكرارات</div>
              </div>
              <span id="mx-status-badge" style="margin-right:auto;font-size:10px;font-weight:700;padding:3px 10px;border-radius:10px;background:var(--sand2);color:var(--txt3)">جارٍ الفحص…</span>
            </div>
            <div id="mx-url-row" style="font-size:11px;color:var(--txt3);margin-bottom:12px;display:none">
               <b id="mx-url-txt"></b>
            </div>
            <div id="mx-scan-prog" style="display:none;margin-bottom:12px">
              <div class="prog-bar" style="margin-bottom:6px"><div class="prog-fill" id="mx-prog-fill" style="width:0%"></div></div>
              <div id="mx-prog-txt" style="font-size:11.5px;color:var(--txt3);text-align:center">جارٍ الفحص…</div>
            </div>
            <div id="mx-error" style="display:none;background:var(--red-bg);border:1px solid var(--red-bd);border-radius:6px;padding:10px 12px;font-size:11.5px;color:var(--red);margin-bottom:12px"></div>
            <div id="mx-result" style="display:none;background:var(--grn-bg);border:1px solid var(--grn-bd);border-radius:6px;padding:12px;font-size:12px;color:var(--green);margin-bottom:12px;text-align:center"></div>
            <div style="display:flex;align-items:center;gap:10px;margin-bottom:12px;background:var(--sand);border:1px solid var(--border);border-radius:8px;padding:10px 12px">
              <label for="mx-maxdays" style="font-size:11.5px;font-weight:700;color:var(--txt2);white-space:nowrap">نافذة المقارنة الزمنية:</label>
              <select id="mx-maxdays" style="height:32px;border:1px solid var(--border);border-radius:6px;background:var(--card);font-family:inherit;font-size:12px;color:var(--txt1);padding:0 8px;outline:none">
                <option value="1">يوم واحد</option>
                <option value="2" selected>يومان</option>
                <option value="3">٣ أيام</option>
                <option value="5">٥ أيام</option>
                <option value="7">أسبوع</option>
                <option value="14">أسبوعان</option>
                <option value="30">شهر</option>
              </select>
              <span style="font-size:10px;color:var(--txt4);line-height:1.4">أقصى فارق بين بلاغين ليُعدّا تكراراً محتملاً</span>
            </div>
            <button class="hero-btn" id="btn-mx-scan" onclick="startMaximoScan()" style="width:100%;justify-content:center;font-size:13px;padding:12px">
               سحب من Maximo الآن
            </button>
          </div>

          <!-- ── divider ── -->
          <div style="display:flex;align-items:center;gap:12px;margin:16px 0">
            <div style="flex:1;height:1px;background:var(--border)"></div>
            <span style="font-size:11px;color:var(--txt4);white-space:nowrap">أو (للاستخدام بدون اتصال)</span>
            <div style="flex:1;height:1px;background:var(--border)"></div>
          </div>

          <!-- ── Excel fallback ── -->
          <div class="drop-zone" id="drop-zone" style="margin-bottom:16px"
               onclick="document.getElementById('file-inp').click()"
               ondragover="event.preventDefault();this.classList.add('drag-over')"
               ondragleave="this.classList.remove('drag-over')"
               ondrop="handleDrop(event)">
            <input type="file" id="file-inp" accept=".xlsx,.xls" hidden onchange="handleFile(this.files[0])">
            <div class="dz-icon"></div>
            <div class="dz-text">رفع ملف Excel</div>
            <div class="dz-hint">.xlsx · .xls · الملف الكامل من Maximo</div>
          </div>
          <div id="upd-prog" class="upd-prog" style="display:none">
            <div class="prog-bar"><div class="prog-fill" id="prog-fill"></div></div>
            <div id="prog-txt">جارٍ الرفع والمعالجة…</div>
          </div>
          <div id="upd-result" style="display:none" class="upd-result"></div>

          <!-- ── last scan info ── -->
          <div id="upd-last" class="upd-last">
            <div class="upd-last-t"> آخر فحص</div>
            <div class="upd-last-row"><span>الحالة</span><span><b id="ul-ready">—</b></span></div>
            <div class="upd-last-row"><span>المصدر</span><span><b id="ul-source">—</b></span></div>
            <div class="upd-last-row"><span>تاريخ الفحص</span><span><b id="ul-date">—</b></span></div>
            <div class="upd-last-row"><span>عدد البلاغات</span><span><b id="ul-srs">—</b></span></div>
            <div class="upd-last-row"><span>المجموعات المكتشفة</span><span><b id="ul-grps">—</b></span></div>
          </div>
        </div>
      </div>

    </div><!-- /pages -->
  </div><!-- /main-area -->
</div><!-- /shell -->

<!-- SR Modal -->
<div class="modal-bg" id="sr-modal" onclick="if(event.target===this)this.classList.remove('open')">
  <div class="modal">
    <div class="modal-hd">
      <span class="modal-title" id="sr-modal-t">تفاصيل البلاغ</span>
      <button class="modal-close" onclick="document.getElementById('sr-modal').classList.remove('open')"></button>
    </div>
    <div class="modal-body" id="sr-modal-b"></div>
  </div>
</div>

<!-- Toasts -->
<div id="toasts"></div>

<!-- Print area — hidden normally, visible only during print -->
<div id="print-area" style="display:none;padding:16px"></div>

<script>
var G=[], D={}, _timer=null, REFRESH=15000;
var _active=new Set(["confirmed","possible","weak"]);
var _scanRunning=false, _scanPhase="", _statusTimer=null;
/* ── Notification tracking ── */
var _knownGroupIds=null;   // null = first load, skip auto-notify
var _lastNewGroupsAt="";   // scanned_at of last new-groups payload we notified
var _notifPermission="default";
/* ── Request browser notification permission (shown once) ── */
function _requestNotifPermission(){
  if(!("Notification" in window))return;
  if(Notification.permission==="granted"){_notifPermission="granted";return;}
  if(Notification.permission!=="denied"){
    Notification.requestPermission().then(function(p){_notifPermission=p;});
  } else {
    _notifPermission="denied";
  }
}
/* ── Show a single browser notification ── */
function _showBrowserNotif(title, body, tag){
  if(_notifPermission!=="granted")return;
  try{
    var n=new Notification(title,{body:body,tag:tag||"lm-dup",icon:"/static/logo.png",dir:"rtl",lang:"ar"});
    n.onclick=function(){window.focus();goto("alerts");n.close();};
    setTimeout(function(){try{n.close();}catch(e){}},12000);
  }catch(e){}
}
/* ── Notify about newly-detected duplicate groups ── */
function _notifyNewGroups(newGrps){
  if(!newGrps||!newGrps.length)return;
  var tierLabel={"confirmed":"🔴 مؤكد","possible":"🟡 محتمل","weak":"🔵 ضعيف"};
  newGrps.slice(0,5).forEach(function(g){
    var tier=tierLabel[g.tier]||g.tier||"";
    var srs=(g.srs||[]).slice(0,3).join(" · ");
    var body=(g.fault||"")+(g.loc?" — "+g.loc:"")+(srs?"\n"+srs:"");
    _showBrowserNotif("تكرار جديد "+tier, body, "lm-"+g.gid.substring(0,20));
  });
  // Also show an in-page toast
  var count=newGrps.length;
  var el=document.createElement("div");
  el.className="toast alert";
  el.innerHTML='<span class="ti">🔔</span><span>'+(count===1?"مجموعة تكرار جديدة واحدة":count+" مجموعات تكرار جديدة")
    +' — <b><u style="cursor:pointer" onclick="goto(\'alerts\');">مراجعة الآن</u></b></span>';
  document.getElementById("toasts").appendChild(el);
  setTimeout(function(){el.remove();},10000);
}
/* ── Poll /api/new-groups every 15 s to surface fresh notifications ── */
var _newGroupsPollTimer=null;
async function _pollNewGroups(){
  try{
    var d=await(await fetch("/api/new-groups?t="+Date.now(),{cache:"no-store"})).json();
    var at=d.scanned_at||"";
    if(at && at!==_lastNewGroupsAt && d.count>0){
      _lastNewGroupsAt=at;
      _notifyNewGroups(d.new_groups||[]);
      loadNotifications();
      setAlertBadges(G.filter(function(g){return!g.decision;}).length);
    }
  }catch(e){}
  _newGroupsPollTimer=setTimeout(_pollNewGroups, REFRESH);
}

// Poll scan-status every 3s — drives the progress banner across all pages
function _pollScanStatus(){
  fetch("/api/scan-status").then(function(r){return r.json();}).then(function(d){
    _scanRunning=d.running||false;
    _scanPhase=(d.progress||{}).phase||"";
    // Update the nav badge / topbar if needed
    if(!_scanRunning&&_statusTimer){
      clearInterval(_statusTimer);_statusTimer=null;
      loadData(true);  // refresh everything once done
    }
  }).catch(function(){});
}
function _ensureStatusPolling(){
  if(!_statusTimer)_statusTimer=setInterval(_pollScanStatus,3000);
}

/* ── utils ── */
function esc(s){
  if(!s)return"";
  return String(s).replace(/&/g,"&amp;").replace(/</g,"&lt;").replace(/>/g,"&gt;").replace(/"/g,"&quot;");
}
function maximoUrl(sr){
  var ticket = String(sr||"").trim();
  var qbe = "ticketid=" + ticket;
  return "https://maximo.kidana.com.sa/maximo/ui/?event=loadapp&value=sr"
    + "&additionalevent=useqbe&additionaleventvalue=" + encodeURIComponent(qbe);
}
function setAlertBadges(n){
  ["nb","nb2"].forEach(function(id){
    var el=document.getElementById(id);if(!el)return;
    if(n>0){el.textContent=n>99?"99+":n;el.style.display="flex";}else el.style.display="none";
  });
  /* notif-total in panel header — single source of truth */
  var total=document.getElementById("notif-total");
  if(total){total.textContent=n>99?"99+":n;total.style.display=n>0?"flex":"none";}
}
function toggleNotifications(e){
  if(e)e.stopPropagation();
  var p=document.getElementById("notif-panel");
  if(!p)return;
  var willOpen=!p.classList.contains("open");
  if(willOpen){
    /* position panel below bell button — keep inside viewport (panel width = 340) */
    var btn=e?e.currentTarget:document.querySelector(".notif-wrap button");
    if(btn){
      var r=btn.getBoundingClientRect();
      var lx=Math.max(8, Math.min(r.left, window.innerWidth-348));
      p.style.left=lx+"px";
    }
    p.classList.add("open");
    loadNotifications();
  } else {
    p.classList.remove("open");
  }
}
document.addEventListener("click",function(e){
  var p=document.getElementById("notif-panel");
  if(p && p.classList.contains("open") && !e.target.closest(".notif-wrap") && !e.target.closest("#notif-panel"))
    p.classList.remove("open");
});
async function loadNotifications(){
  var list=document.getElementById("notif-list");
  var panelEl=document.getElementById("notif-panel");
  var panelOpen=panelEl&&panelEl.classList.contains("open");
  /* Show loading spinner only when panel is actually open */
  if(panelOpen && list) list.innerHTML='<div class="notif-empty">جارٍ التحميل…</div>';
  try{
    var d=await(await fetch("/api/notifications")).json();
    var pending=(d.pending_groups||0);
    var items=d.alerts||[];
    /* notif-total is owned by setAlertBadges — don't overwrite it here */
    /* Show email status row */
    fetch("/api/notify-config").then(function(r){return r.json();}).then(function(cfg){
      var row=document.getElementById("notif-email-row");
      if(row)row.style.display=cfg.email_enabled?"block":"none";
    }).catch(function(){});
    if(!list)return;
    if(!pending && !items.length){
      list.innerHTML='<div class="notif-empty">لا توجد مجموعات معلقة</div>';
      return;
    }
    var pendingHtml = pending
      ? '<div class="notif-item" onclick="goto(\'alerts\');document.getElementById(\'fl-status\').value=\'pending\';renderAlerts();document.getElementById(\'notif-panel\').classList.remove(\'open\')">'
        + '<div style="display:flex;align-items:center;gap:8px">'
        + '<span style="font-size:22px;font-weight:800;color:var(--amber)">'+pending+'</span>'
        + '<div><div class="notif-item-sr">مجموعة معلقة تحتاج مراجعة</div>'
        + '<div class="notif-item-meta">اضغط للانتقال لصفحة التنبيهات</div></div></div></div>'
      : "";
    var liveSummary = (d.open_alert_count||0)
      ? '<div class="notif-item" onclick="goto(\'alerts\');document.getElementById(\'notif-panel\').classList.remove(\'open\')">'
        + '<div style="display:flex;align-items:center;gap:8px">'
        + '<span style="font-size:22px;font-weight:800;color:var(--red)">'+(d.open_alert_count||0)+'</span>'
        + '<div><div class="notif-item-sr">تنبيه لحظي مفتوح</div>'
        + '<div class="notif-item-meta">تطابقات جديدة رصدها النظام</div></div></div></div>'
      : "";
    var alertHtml = items.map(function(a){
      return '<div class="notif-item" onclick="goto(\'alerts\');document.getElementById(\'notif-panel\').classList.remove(\'open\')">'
        +'<div class="notif-item-sr">SR '+esc(a.new_sr||"")+' ↔ SR '+esc(a.match_sr||"")+'</div>'
        +'<div class="notif-item-meta">تطابق '+scorePct(a.score)+'% · '+esc((a.new_location||"").substring(0,30))+'</div>'
        +'</div>';
    }).join("");
    list.innerHTML=pendingHtml+liveSummary+alertHtml;
  }catch(err){
    if(list)list.innerHTML='<div class="notif-empty">تعذر تحميل التنبيهات</div>';
  }
}
function scorePct(s){return Math.min(100,Math.round((s||0)/15*100));}
function relTime(r){
  if(!r)return"—";
  var parts=String(r).split(/[T ]/);
  var dt=new Date(parts[0]+(parts[1]?" "+parts[1]:""));
  if(isNaN(dt))return r;
  var diff=Math.round((Date.now()-dt)/1000);
  if(diff<60)return"الآن";
  if(diff<3600)return"منذ "+Math.round(diff/60)+" د";
  if(diff<86400)return"منذ "+Math.round(diff/3600)+" س";
  return"منذ "+Math.round(diff/86400)+" يوم";
}
function timeGap(members){
  if(!members||members.length<2)return"";
  var dates=members.map(function(m){return new Date(String(m.reported||"").replace(" ","T"));}).filter(function(d){return!isNaN(d);});
  if(dates.length<2)return"";
  var span=Math.abs(dates[dates.length-1]-dates[0])/1000;
  if(span<120)return Math.round(span)+" ثانية";
  if(span<7200)return Math.round(span/60)+" دقيقة";
  if(span<172800)return Math.round(span/3600)+" ساعة";
  return Math.round(span/86400)+" يوم";
}
function stBadge(s){
  var v=(s||"").toUpperCase();
  if(v==="OPEN"||v==="NEW"||v==="QUEUED")return'<span class="st-open">'+esc(s)+'</span>';
  if(v==="CLOSED"||v==="COMP"||v==="RESOLVED")return'<span class="st-closed">'+esc(s)+'</span>';
  if(v==="INPRG"||v==="WPCOND"||v==="WAPPR"||v==="WMATL")return'<span class="st-wip">'+esc(s)+'</span>';
  if(s)return'<span class="st-other">'+esc(s)+'</span>';
  return"";
}
function stDot(s){
  var v=(s||"").toUpperCase();
  if(v==="OPEN"||v==="NEW"||v==="QUEUED")return"open";
  if(v==="CLOSED"||v==="COMP"||v==="RESOLVED")return"closed";
  if(v==="INPRG"||v==="WPCOND"||v==="WAPPR"||v==="WMATL")return"wip";
  return"other";
}
function toast(msg,type,icon){
  var el=document.createElement("div");
  el.className="toast "+(type||"");
  el.innerHTML='<span class="ti">'+(icon||"ℹ")+'</span><span>'+esc(msg)+'</span>';
  document.getElementById("toasts").appendChild(el);
  setTimeout(function(){el.remove();},4000);
}

/* ── navigation ── */
function goto(id){
  document.querySelectorAll(".page").forEach(function(p){p.classList.remove("active");});
  document.getElementById("pg-"+id).classList.add("active");
  document.querySelectorAll(".nav-item").forEach(function(n){n.classList.toggle("active",n.dataset.page===id);});
  if(id==="dash" && _mapInst){ setTimeout(function(){ _mapInst.invalidateSize(); }, 60); }
}
function activePage(){
  var p=document.querySelector(".page.active");
  return p ? p.id.replace("pg-","") : "";
}
/* navigate from SR list to the matching group in the alerts page */
function gotoGroup(sr){
  goto('alerts');
  /* reset all filters so the target group is definitely visible */
  var fs=document.getElementById('fl-status'); if(fs) fs.value='';
  var ft=document.getElementById('fl-tier');   if(ft) ft.value='';
  var fq=document.getElementById('srch');      if(fq) fq.value='';
  document.querySelectorAll('.ftier').forEach(function(b){b.classList.remove('active');});
  renderAlerts();
  setTimeout(function(){
    var cards=document.querySelectorAll('#alerts-list .gcard');
    for(var i=0;i<cards.length;i++){
      if(cards[i].querySelector('.sr-chk[value="'+sr+'"]')){
        cards[i].scrollIntoView({behavior:'smooth',block:'start'});
        cards[i].style.outline='3px solid var(--gold)';
        setTimeout(function(c){c.style.outline='';},2500,cards[i]);
        return;
      }
    }
  },400);
}
/* ── PDF print for one group card ── */
function printGroup(btn){
  var card = btn.closest(".gcard");
  if(!card){ alert("لم يتم العثور على الكارت"); return; }
  /* Collect SR numbers from hidden checkboxes inside the card */
  var srNums = [];
  card.querySelectorAll('.sr-chk').forEach(function(chk){
    if(chk.value) srNums.push(chk.value);
  });
  /* Build filename: "مكرر - SR47379 SR47380 SR47381" */
  var srPart = srNums.length ? srNums.join(' · ') : "مجموعة";
  var pdfTitle = 'مكرر — ' + srPart;

  var area = document.getElementById("print-area");
  var now  = new Date().toLocaleString("ar-SA");
  var clone = card.cloneNode(true);
  /* أخفِ أزرار الإجراءات من النسخة */
  clone.querySelectorAll(".dec-dock,.btn-undo,.sel-toggle,.sr-chk,.gc-pdf-btn,.gcard-right button").forEach(function(el){el.style.display="none";});
  area.innerHTML = '<div style="margin-bottom:10px;font-weight:800;font-size:13px;border-bottom:2px solid #c8a044;padding-bottom:6px;font-family:Cairo,sans-serif">منظومة جودة البلاغات — كدانة مالك &nbsp;|&nbsp; '+now+'</div>';
  area.appendChild(clone);
  /* Set document title → browser uses it as the default PDF filename */
  var _prevTitle = document.title;
  document.title = pdfTitle;
  window.print();
  /* Restore title after a short delay (print dialog is async) */
  setTimeout(function(){ document.title = _prevTitle; }, 1500);
}
function toggleSidebar(){
  document.querySelector(".shell").classList.toggle("sb-closed");
  if(_mapInst) setTimeout(function(){ _mapInst.invalidateSize(); }, 280);
}

/* ── topbar refresh: fetch newest Maximo page immediately then reload ── */
async function topbarRefresh(){
  var btn=document.getElementById("btn-topbar-refresh");
  if(btn){btn.textContent="جارٍ التحديث…";btn.disabled=true;}
  try{
    var r=await fetch("/api/quick-scan-maximo?t="+Date.now(),{method:"POST",cache:"no-store"});
    var d=await r.json();
    if(d.ok){
      var res=d.result||{};
      var added=res.new_count||0;
      var changed=res.changed_count||0;
      toast(added ? ("تم جلب "+added+" بلاغ جديد") : (changed ? ("تم تحديث "+changed+" بلاغ") : "لا توجد بلاغات جديدة"), "ok", "");
      await loadData(true);
      if(activePage()==="reports") await loadReports();
    } else {
      toast(d.message||"خطأ","err","");
    }
  }catch(e){
    toast("تعذّر الاتصال بـ Maximo","err","");
  }finally{
    if(btn){btn.textContent="تحديث الآن";btn.disabled=false;}
  }
}

/* ── data load ── */
async function loadData(force){
  try{
    var data=await(await fetch("/api/scan?t="+Date.now(),{cache:"no-store"})).json();
    applyData(data);
    if(activePage()==="reports") await loadReports();
    if(force)toast("تم التحديث","ok","");
  }catch(e){toast("خطأ في الاتصال","err","");}
}
function applyData(d){
  if(!d.ready){
    var msg='<div class="empty"><div class="ei"></div><h3>لا توجد بيانات بعد</h3>'
      +'<p>اضغط الزر لسحب كل البلاغات من Maximo وبدء كشف المكررات</p>'
      +'<button class="hero-btn" onclick="goto(\'upload\');loadUploadInfo();setTimeout(startMaximoScan,400)" style="margin-top:12px;font-size:13px;padding:12px 24px"> سحب من Maximo الآن</button></div>';
    document.getElementById("alerts-list").innerHTML=msg;
    document.getElementById("dash-grps").innerHTML=msg;
    setAlertBadges(0);
    clearTimeout(_timer);_timer=setTimeout(function(){loadData();},REFRESH);
    return;
  }
  // If scan is running, show progress hint
  if(_scanRunning){
    var ph=_scanPhase||"جارٍ الفحص…";
    document.getElementById("dash-grps").innerHTML='<div class="empty"><div class="ei"></div><h3>'+esc(ph)+'</h3><p>سيظهر المحتوى بعد اكتمال الفحص</p></div>';
  }
  G=d.groups||[];
  D=d;
  /* ── Detect new groups vs previous load ───────────────────────── */
  var currentIds=new Set(G.map(function(g){return g.id;}));
  if(_knownGroupIds!==null){
    var newPendingGroups=G.filter(function(g){
      return !g.decision && !_knownGroupIds.has(g.id);
    });
    if(newPendingGroups.length){
      // Build compact notification payload from group data we already have
      var notifPayload=newPendingGroups.slice(0,5).map(function(g){
        var m=(g.members||[])[0]||{};
        return {
          gid:   g.id,
          score: g.score,
          tier:  g.tier,
          size:  g.size,
          fault: m.fault||"",
          loc:   m.loc||"",
          srs:   (g.members||[]).slice(0,4).map(function(x){return x.sr;}),
        };
      });
      _notifyNewGroups(notifPayload);
    }
  }
  _knownGroupIds=currentIds;
  var pend=G.filter(function(g){return!g.decision;}).length;
  /* legacy KPI elements removed — guard each lookup */
  function _set(id, v){var el=document.getElementById(id); if(el) el.textContent=v;}
  var nConf=G.filter(function(g){return g.decision==="duplicate";}).length;
  var nDiff=G.filter(function(g){return g.decision==="different";}).length;
  var nDecided=G.filter(function(g){return!!g.decision;}).length;
  var nTotal=G.length;
  _set("sb-pend", pend);
  _set("sb-conf", nConf);
  _set("sb-rej",  nDiff);
  /* progress bar */
  var pct=nTotal>0?Math.round(nDecided*100/nTotal):0;
  var pfill=document.getElementById("review-prog-fill");
  var plbl=document.getElementById("review-prog-lbl");
  if(pfill) pfill.style.width=pct+"%";
  if(plbl)  plbl.textContent=nDecided+" / "+nTotal;
  _set("rpb-confirmed", nConf);
  _set("rpb-different", nDiff);
  _set("rpb-pending",   pend);
  setAlertBadges(pend);
  loadNotifications();
  var age=d.age_seconds,as="—";
  if(age!=null){if(age<60)as="تحديث منذ "+Math.round(age)+" ث";else if(age<3600)as="تحديث منذ "+Math.round(age/60)+" د";else as="تحديث منذ "+Math.round(age/3600)+" س";}
  document.getElementById("sf-age").textContent=as;
  document.getElementById("sf-age2").textContent=as;
  var tbAge=document.getElementById("tb-scan-age");
  if(tbAge) tbAge.textContent=as;
  var locs=new Set(),faults=new Set();
  G.forEach(function(g){g.members.forEach(function(m){if(m.loc)locs.add(m.loc);if(m.fault)faults.add(m.fault.substring(0,45));});});
  ["fl-loc","fl-fault"].forEach(function(id){
    var s=document.getElementById(id),prev=s.value;
    var items=id==="fl-loc"?[...locs].sort():[...faults].sort();
    s.innerHTML='<option value="">'+(id==="fl-loc"?"كل المواقع":"كل الأعطال")+'</option>';
    items.forEach(function(v){var o=document.createElement("option");o.value=v;o.textContent=v;if(v===prev)o.selected=true;s.appendChild(o);});
  });
  renderAlerts();renderDash();
  clearTimeout(_timer);_timer=setTimeout(function(){loadData();},REFRESH);
}

/* ── filters ── */
function toggleTier(btn){
  var t=btn.dataset.tier;
  if(_active.has(t)){if(_active.size===1)return;_active.delete(t);btn.classList.remove("on");}
  else{_active.add(t);btn.classList.add("on");}
  renderAlerts();
}
function filtered(){
  var sf=document.getElementById("fl-status").value;
  var lf=document.getElementById("fl-loc").value;
  var ff=document.getElementById("fl-fault").value;
  var tf=document.getElementById("fl-time"); var timeHours=tf?parseInt(tf.value||"0",10):0;
  var sortEl=document.getElementById("fl-sort");
  var sortMode = sortEl ? sortEl.value : "newest";
  var cutoff=timeHours>0?new Date(Date.now()-timeHours*3600000):null;
  var out = G.filter(function(g){
    if(sf==="pending"&&g.decision)return false;
    if(sf==="duplicate"&&g.decision!=="duplicate")return false;
    if(sf==="different"&&g.decision!=="different")return false;
    if(lf&&!g.members.some(function(m){return m.loc===lf;}))return false;
    if(ff&&!g.members.some(function(m){return(m.fault||"").startsWith(ff);}))return false;
    if(cutoff){
      /* keep group only if at least one SR was reported within the time window */
      var hasRecent=g.members.some(function(m){
        if(!m.reported)return false;
        try{return new Date(m.reported)>=cutoff;}catch(e){return false;}
      });
      if(!hasRecent)return false;
    }
    return true;
  });
  /* sort */
  function newestSr(g){
    var dates = g.members.map(function(m){return m.reported||"";}).filter(Boolean).sort();
    return dates.length ? dates[dates.length-1] : "";
  }
  if(sortMode === "newest"){
    out.sort(function(a,b){return newestSr(b).localeCompare(newestSr(a));});
  } else if(sortMode === "oldest"){
    out.sort(function(a,b){return newestSr(a).localeCompare(newestSr(b));});
  } else if(sortMode === "score"){
    out.sort(function(a,b){return (b.score||0)-(a.score||0);});
  }
  return out;
}
function setStatusFilter(status){
  var sel=document.getElementById("fl-status");
  if(sel)sel.value=status;
  renderAlerts();
}
function updateDecisionTabs(){
  var counts={pending:0,duplicate:0,different:0};
  G.forEach(function(g){
    var d=(g.decision==="duplicate"||g.decision==="different")?g.decision:"pending";
    counts[d]++;
  });
  Object.keys(counts).forEach(function(k){
    var el=document.getElementById("dt-"+k);
    if(el)el.textContent=counts[k]||0;
  });
  var sf=document.getElementById("fl-status");
  var active=sf?sf.value:"";
  document.querySelectorAll(".decision-tab").forEach(function(b){
    b.classList.toggle("on", (b.dataset.status||"")===active);
  });
}
function onSearch(){
  var q=document.getElementById("srch").value.toLowerCase().trim();
  if(!q){renderAlerts();return;}
  renderList(G.filter(function(g){return g.members.some(function(m){
    return(m.sr||"").toLowerCase().includes(q)||(m.loc||"").toLowerCase().includes(q)||(m.fault||"").toLowerCase().includes(q)||(m.asset||"").toLowerCase().includes(q);
  });}),document.getElementById("alerts-list"));
}

/* ── helper: format a value with optional sub-label (e.g. "B1-N\\nName") ── */
function valWith(code, sub){
  if(!code && !sub) return '<span class="cmp-empty">—</span>';
  if(!sub) return esc(code);
  return '<div class="cmp-stack"><div class="cmp-code">'+esc(code)+'</div><div class="cmp-sub">'+esc(sub)+'</div></div>';
}
/* ── helper: gap from origin ── */
function gapFromOrigin(originDate, mDate, isOrigin){
  if(isOrigin) return '<span class="cmp-origin-pill">الأصل</span>';
  if(!originDate || !mDate) return '<span class="cmp-empty">—</span>';
  var d1=new Date(originDate), d2=new Date(mDate);
  if(isNaN(d1)||isNaN(d2)) return '<span class="cmp-empty">—</span>';
  var diffSec = Math.round((d2-d1)/1000);
  if(diffSec < 0) diffSec = -diffSec;
  if(diffSec < 60) return '<span class="cmp-gap">+'+diffSec+' ث</span>';
  if(diffSec < 3600) return '<span class="cmp-gap">+'+Math.round(diffSec/60)+' د</span>';
  if(diffSec < 86400) return '<span class="cmp-gap">+'+Math.round(diffSec/3600)+' س</span>';
  return '<span class="cmp-gap">+'+Math.round(diffSec/86400)+' يوم</span>';
}
function compactLine(lbl, value){
  if(!value)return"";
  return '<div class="cmp-line"><b>'+esc(lbl)+'</b><span>'+value+'</span></div>';
}
function dateAndGap(originDate, m, isOrigin){
  var reported = esc(m.reported || "—");
  var gap = gapFromOrigin(originDate, m.reported, isOrigin);

  return '<div class="cmp-mini">'
    + compactLine(
        'التاريخ',
        reported + ' <span class="gap-inline">' + gap + '</span>'
      )
    + '</div>';
}

function workAreaSummary(m){
  return '<div class="cmp-mini">'
    + compactLine('منطقة العمل', esc(m.workzone || "—"))
    + compactLine('المشعر', esc(m.site || "—"))
    + '</div>';
}
/* ── render comparison table for a group (rows=fields, cols=SRs) ── */
function renderComparison(g){
  var origin = g.members[0];
  /* column headers */
  var headers = '<tr class="cmp-th-row"><th class="cmp-th-lbl">الحقل</th>'
    + g.members.map(function(m, i){
        var mu=maximoUrl(m.sr);
        var mdata=JSON.stringify(m).replace(/</g,"\\u003c");
        var originBadge = i===0?'<span class="cmp-origin-mini">الأصل</span>':'';
        var escIcon = m.resp_esc ? '<span title="'+esc(m.resp_esc)+'" style="margin-right:3px;font-size:11px;cursor:default" aria-label="تصعيد">🔴</span>' : '';
        return '<th class="cmp-col-head'+(i===0?' origin':'')+'">'
          +'<div class="cmp-col-head-top">'
            +'<input type="checkbox" class="sr-chk" value="'+esc(m.sr)+'" onclick="event.stopPropagation()">'
            +escIcon
            +'<span class="cmp-sr" onclick="openSrModal('+mdata+')">SR '+esc(m.sr)+'</span>'
            +originBadge
          +'</div>'
          +'<div style="display:flex;align-items:center;justify-content:space-between;margin-top:4px">'
            +'<a class="cmp-mx" href="'+mu+'" target="_blank" rel="noopener">Maximo ↗</a>'
            +stBadge(m.status)
          +'</div>'
        +'</th>';
      }).join("")
    +'</tr>';

  /* fields organized into sections */
  function row(lbl, getter){
    var cells = g.members.map(function(m, i){
      var v = getter(m, i);
      if(v == null || v === "") v = '<span class="cmp-empty">—</span>';
      return '<td class="cmp-cell'+(i===0?' origin':'')+'">'+v+'</td>';
    }).join("");
    return '<tr><td class="cmp-lbl">'+lbl+'</td>'+cells+'</tr>';
  }
  function sec(title){
    return '<tr class="cmp-sec"><td colspan="'+(g.members.length+1)+'">'+title+'</td></tr>';
  }
  var body = ""
    + sec('معلومات البلاغ')
    + row('التاريخ والفرق',   function(m,i){return dateAndGap(origin.reported, m, i===0);})
    + row('المصدر',           function(m){ return valWith(m.source, m.source_desc);})
    + row('الأولوية',         function(m){ return m.priority_desc?esc(m.priority)+' — '+esc(m.priority_desc):esc(m.priority);})
    + row('العطل',            function(m){ return '<div class="cmp-fault" style="white-space:normal;word-break:break-word">'+esc(m.fault||"")+'</div>';})
    + sec('الموقع والأصل')
    + row('الموقع',           function(m){ return valWith(m.loc, m.loc_ar);})
    + row('المنطقة / المربع', function(m){ var z=[m.region,m.block].filter(Boolean).join(' · '); return z?esc(z):'<span class="cmp-empty">—</span>';})
    + row('الأصل',            function(m){ return valWith(m.asset, m.asset_ar);})
    + row('منطقة العمل / المشعر', function(m){ return workAreaSummary(m);})
    + row('المسافة عن الأصل', function(m,i){
        var mLat=parseFloat(m.lat||0), mLon=parseFloat(m.lon||0);
        var gidSafe=(g.id||'').replace(/'/g,"\\'");
        var mapBtn=(mLat&&mLon)
          ? ' <button class="cmp-map-btn" onclick="mapFocusGroup(\''+gidSafe+'\','+mLat+','+mLon+')" title="عرض في الخريطة">📍</button>'
          : '';
        if(i===0){ return mapBtn||'<span class="cmp-empty">—</span>'; }
        var oLat=parseFloat(origin.lat||0), oLon=parseFloat(origin.lon||0);
        if(!oLat||!oLon||!mLat||!mLon)
          return '<span class="cmp-empty" title="إحداثيات غير متاحة">—</span>'+mapBtn;
        var d=haversineM(oLat,oLon,mLat,mLon);
        var fmt=fmtDist(d);
        var note='', cls='', ttl='';
        if(d<50){  cls='cmp-dist-near'; ttl='إحداثيات متقاربة جداً';}
        else if(d<500){ cls='cmp-dist-mid'; ttl='مسافة معقولة — قد يكون خطأ إدخال طفيف'; note=' ⚠️';}
        else { cls='cmp-dist-far'; ttl='إحداثيات متباعدة — قد يكون خطأ في الإدخال'; note=' ⚠️';}
        return '<span class="'+cls+'" title="'+ttl+'">'+fmt+note+'</span>'+mapBtn;
      })
    + sec('الأطراف')
    + row('رقم المبلّغ',       function(m){ return esc(m.requestor_no||m.caller_phone||"");})
    + row('جهة المبلّغ',       function(m){ return esc(m.caller_party||m.party||"");})
    + row('اسم المبلّغ',       function(m){ return esc(m.reported_name||"");})
    + row('اسم مدخل البلاغ',   function(m){
          var display = m.reporter_display || "";
          var uname   = m.reporter || "";
          if(!display && !uname) return '<span class="cmp-empty">—</span>';
          if(display && uname && display !== uname)
            return display + '<div class="cmp-sub">'+esc(uname)+'</div>';
          return esc(display || uname);
      })
    + row('المقاول',  function(m){
          var name = m.contractor||m.party||"";
          var c = m.contract ? ('العقد: '+m.contract) : "";
          if(!name && !c) return '<span class="cmp-empty">—</span>';
          if(!name)       return '<div class="cmp-sub">'+esc(c)+'</div>';
          if(!c)          return esc(name);
          return '<div class="cmp-stack"><div class="cmp-sub">'+esc(c)+'</div><div>'+esc(name)+'</div></div>';
      })
    + sec('التفاصيل')
    + row('التفاصيل',         function(m){ return m.detail?'<div class="cmp-detail">'+esc(m.detail)+'</div>':'<span class="cmp-empty">—</span>';});

  return '<div class="cmp-wrap"><table class="cmp-tbl"><thead>'+headers+'</thead><tbody>'+body+'</tbody></table></div>';
}
/* legacy srDetail kept only for the small dashboard summary (not used in alerts) */
function srDetail(m, idx, gap){ return ''; }
/* legacy 2-col helper kept for the dashboard summary */
function srCol(m,label){
  var mu=maximoUrl(m.sr);
  var af=[m.asset,m.fault].filter(Boolean).join(" — ");
  var mdata=JSON.stringify(m).replace(/</g,"\\u003c");
  return'<div class="sr-col">'
    +'<div class="sr-col-lbl"><input type="checkbox" class="sr-chk" value="'+esc(m.sr)+'" onclick="event.stopPropagation()"> '+esc(label)+' '+stBadge(m.status)+'</div>'
    +'<span class="sr-num" onclick="openSrModal('+mdata+')">'+esc(m.sr)+'</span>'
    +(m.loc?'<div class="sr-row"><span class="sr-rv">'+esc(m.loc)+'</span></div>':"")
    +(af?'<div class="sr-row"><span class="sr-rv">'+esc(af)+'</span></div>':"")
    +(m.reported?'<div class="sr-row"><span class="sr-rv">'+esc(relTime(m.reported))+'</span></div>':"")
    +'<a class="sr-mx" href="'+mu+'" target="_blank" rel="noopener"> فتح في Maximo</a>'
    +'</div>';
}

/* ── sub-decisions list ── */
function subDecHtml(subs,gid){
  if(!subs||!subs.length)return"";
  var eid=CSS.escape(gid);
  return'<div class="sub-dec-list">'
    +subs.map(function(s,i){
      var cls=s.decision==="duplicate"?"confirmed":s.decision==="different"?"rejected":"";
      var icon=s.decision==="duplicate"?"":s.decision==="different"?"":"";
      return'<div class="sub-dec-item '+cls+'">'
        +'<span class="sub-dec-txt">'+icon+' <b>'+esc(s.srs.join(" · "))+'</b>'+(s.note?' — '+esc(s.note):"")+'</span>'
        +'<button class="sub-dec-del" title="حذف" onclick="delSubDec(\''+eid+'\',\''+gid+'\','+i+')"></button>'
        +'</div>';
    }).join("")
    +'</div>';
}

/* ── Haversine distance in metres between two lat/lon pairs ── */
function haversineM(lat1, lon1, lat2, lon2){
  if(!lat1||!lon1||!lat2||!lon2) return null;
  var R=6371000, toR=Math.PI/180;
  var dLat=(lat2-lat1)*toR, dLon=(lon2-lon1)*toR;
  var a=Math.sin(dLat/2)*Math.sin(dLat/2)
       +Math.cos(lat1*toR)*Math.cos(lat2*toR)
       *Math.sin(dLon/2)*Math.sin(dLon/2);
  return R*2*Math.atan2(Math.sqrt(a),Math.sqrt(1-a));
}
function fmtDist(m){
  if(m==null) return null;
  if(m<1000) return Math.round(m)+' م';
  return (m/1000).toFixed(2)+' كم';
}

/* ── time gap between two ISO dates in Arabic ── */
function gapBetween(a, b){
  if(!a||!b) return "";
  var d1=new Date(a), d2=new Date(b);
  if(isNaN(d1)||isNaN(d2)) return "";
  var diffMin = Math.round((d2-d1)/60000);
  if(diffMin < 0) diffMin = -diffMin;
  if(diffMin < 60) return diffMin + " دقيقة";
  if(diffMin < 1440) return Math.round(diffMin/60) + " ساعة";
  return Math.round(diffMin/1440) + " يوم";
}

/* ── render group card with ALL members expanded ── */
function renderCard(g){
  var pct=scorePct(g.score);
  var dec=g.decision||"";
  var decided=!!dec;
  var sbadge={
    "":"<span class='gc-st pending'>معلقة</span>",
    duplicate:"<span class='gc-st confirmed'> تكرار محقق</span>",
    different:"<span class='gc-st rejected'> مختلفة</span>",
  }[dec]||"";
  var bannerTxt={duplicate:"تم تأكيد هذه البلاغات كتكرار",different:"تم استبعاد هذه البلاغات — ليست تكراراً"}[dec]||"";
  var bannerCls=dec==="duplicate"?"confirmed":"rejected";
  var undoCall=g.parent_id
    ? "delSubDec('', '"+g.parent_id+"', "+g.sub_idx+")"
    : "undoDecision('"+g.id+"')";
  var banner=decided?'<div class="dec-banner '+bannerCls+'">'+bannerTxt+' <button class="btn-undo" onclick="'+undoCall+'">تراجع عن القرار</button></div>':"";
  var eid=CSS.escape(g.id);
  var hasSubDec=g.sub_decisions&&g.sub_decisions.length>0;
  /* render comparison table (rows=fields, cols=SRs) */
  var tableHtml = renderComparison(g);
  return'<div class="gcard'+(decided?" decided":"")+'" id="gc-'+eid+'">'
    +'<div class="gcard-head">'
      +'<div class="gcard-info">'
        +'<div class="gcard-title"><b>'+g.members.length+'</b> بلاغات مكررة · <b>'+g.score+'</b> نقطة</div>'
        +((g.members[0]&&g.members[0].fault)?'<div class="gcard-fault-full">'+esc(g.members[0].fault)+'</div>':"")
        +(g.reasons?'<div class="gcard-meta"><b>أسباب الكشف:</b> '+esc(g.reasons)+'</div>':"")
      +'</div>'
      +'<div class="gcard-right">'
        +sbadge
        +'<button class="gc-pdf-btn" onclick="printGroup(this)" title="تنزيل PDF">⬇ PDF</button>'
        +'<div class="score-ring '+g.tier+'">'+pct+'%</div>'
      +'</div>'
    +'</div>'
    +tableHtml
    +(hasSubDec?subDecHtml(g.sub_decisions,g.id):"")
    +banner
    +(decided?"":'<div class="dec-dock">'
      +'<div class="dec-hint"><span class="dec-hint-i">i</span><span>حدد البلاغات التي تريد نقلها للمؤكدة أو المستبعدة. غير المحدد يبقى في الأساسية.</span></div>'
      +'<input class="note-inp" id="ni-'+eid+'" type="text" placeholder="أضف ملاحظة (اختياري)…" value="'+esc(g.note||"")+'">'
      +'<div class="dec-actions">'
        +'<div class="dec-grp dec-grp-all">'
          +'<span class="dec-grp-lbl">قرار للمجموعة كاملة:</span>'
          +'<button class="btn-reject"  onclick="decide(\''+g.id+'\',\'different\')">استبعاد</button>'
          +'<button class="btn-confirm" onclick="decide(\''+g.id+'\',\'duplicate\')">تأكيد التكرار</button>'
        +'</div>'
        +'<div class="dec-grp dec-grp-sel" id="dgs-'+eid+'">'
          +'<span class="dec-grp-lbl">للمحدّدة (<span id="sel-cnt-'+eid+'">0</span>):</span>'
          +'<button class="btn-sel-reject"  onclick="decideSelected(\''+g.id+'\',\'different\')">رفض المحدّدة</button>'
          +'<button class="btn-sel-confirm" onclick="decideSelected(\''+g.id+'\',\'duplicate\')">تأكيد المحدّدة</button>'
        +'</div>'
      +'</div>'
    +'</div>')
  +'</div>';
}
function toggleSelMode(eid){
  var card=document.getElementById("gc-"+eid);
  if(card)card.classList.toggle("sel-mode");
}
function getChecked(eid){
  var card=document.getElementById("gc-"+eid);
  if(!card)return[];
  var checked=[];
  card.querySelectorAll(".sr-chk:checked").forEach(function(cb){checked.push(cb.value);});
  return checked;
}

/* ── render lists ── */
function renderList(groups,container){
  if(!groups.length){container.innerHTML='<div class="empty"><div class="ei"></div><h3>لا توجد مجموعات بهذا الفلتر</h3></div>';return;}
  container.innerHTML=groups.map(renderCard).join("");
  /* attach live count updater on checkboxes */
  container.querySelectorAll(".gcard").forEach(function(card){
    var eid = card.id.substring(3);   /* strip "gc-" */
    var dgs = card.querySelector(".dec-grp-sel");
    var cnt = card.querySelector("[id^='sel-cnt-']");
    function refresh(){
      var checked = card.querySelectorAll(".sr-chk:checked").length;
      if(cnt) cnt.textContent = checked;
      if(dgs){
        if(checked === 0) dgs.classList.add("disabled");
        else dgs.classList.remove("disabled");
      }
    }
    card.querySelectorAll(".sr-chk").forEach(function(cb){
      cb.addEventListener("change", refresh);
    });
    refresh();
  });
}
function renderAlerts(){
  updateDecisionTabs();
  var container = document.getElementById("alerts-list");
  var groups = filtered();
  /* update badge to reflect current time-filtered pending count */
  var filteredPend = groups.filter(function(g){return !g.decision;}).length;
  setAlertBadges(filteredPend);
  /* update stats-bar counters to reflect current filter */
  function _set(id,v){var el=document.getElementById(id);if(el)el.textContent=v;}
  _set("sb-pend", filteredPend);
  _set("sb-conf", groups.filter(function(g){return g.decision==="duplicate";}).length);
  _set("sb-rej",  groups.filter(function(g){return g.decision==="different";}).length);
  /* ── Skip DOM rewrite if data hasn't changed ──────────────────────────
     Build a fingerprint of the current filtered set (id + decision).
     If it matches the last render, leave the DOM untouched — the user's
     scroll position and any open state are naturally preserved. */
  var fingerprint = groups.map(function(g){return g.id+':'+( g.decision||'');}).join('|');
  if(container._renderKey === fingerprint && container.querySelector('.gcard')){
    return;   /* nothing changed — don't touch the DOM */
  }
  container._renderKey = fingerprint;
  renderList(groups, container);
  /* (أ) keep the dashboard map + numbers in sync with the active filters */
  try{ _drawFilteredMapAndKpis(); }catch(e){}
}
/* ── (أ) Dashboard filter — applies ONLY the visible loc/fault/time filters.
   The hidden status filter (pending/duplicate/different) belongs to the alert
   LIST, not the dashboard, so we ignore it here — otherwise the headline
   numbers would collapse to the "pending" subset. Returns the groups whose
   members fall inside the chosen location / fault / time window. ── */
function _dashFiltered(){
  var lfEl=document.getElementById("fl-loc");   var lf=lfEl?lfEl.value:"";
  var ffEl=document.getElementById("fl-fault"); var ff=ffEl?ffEl.value:"";
  var tfEl=document.getElementById("fl-time");  var th=tfEl?parseInt(tfEl.value||"0",10):0;
  var cutoff=th>0?new Date(Date.now()-th*3600000):null;
  return (G||[]).filter(function(g){
    if(lf&&!(g.members||[]).some(function(m){return m.loc===lf;}))return false;
    if(ff&&!(g.members||[]).some(function(m){return (m.fault||"").startsWith(ff);}))return false;
    if(cutoff&&!(g.members||[]).some(function(m){
      if(!m.reported)return false;
      try{return new Date(m.reported)>=cutoff;}catch(e){return false;}
    }))return false;
    return true;
  });
}
function _filterActive(){
  var lf=document.getElementById("fl-loc");   var ff=document.getElementById("fl-fault");
  var tf=document.getElementById("fl-time");
  return !!((lf&&lf.value)||(ff&&ff.value)||(tf&&tf.value));
}
/* Redraw the map + headline KPIs from the filtered group set so the dashboard
   always matches the chosen filters. Called by renderDash() (every refresh)
   AND by renderAlerts() (the moment a filter changes). */
function _drawFilteredMapAndKpis(){
  var groups=_dashFiltered();
  var _set=function(id,v){var el=document.getElementById(id);if(el)el.textContent=v;};
  /* build dup geo points from the filtered groups (skip excluded ones) */
  var dups=[], locSet={};
  groups.forEach(function(g){
    if(g.decision==='different')return;
    (g.members||[]).forEach(function(mm){
      var la=parseFloat(mm.lat||0), lo=parseFloat(mm.lon||0);
      if(la&&lo){ dups.push({sr:mm.sr,lat:la,lon:lo,loc:mm.loc||'',fault:mm.fault||''}); if(mm.loc)locSet[mm.loc]=1; }
    });
  });
  drawMap(dups, D.geo_normal||[], groups);
  _set("map-sub", Object.keys(locSet).length+" موقع فيه مكررات"+(_filterActive()?" (مفلتر)":""));
  /* headline counts from the filtered set */
  var nG=groups.length;
  var nConf=groups.filter(function(g){return g.decision==='duplicate';}).length;
  var nDiff=groups.filter(function(g){return g.decision==='different';}).length;
  var nRem =groups.filter(function(g){return !g.decision;}).length;
  _set("k-groups",nG); _set("k-confirmed",nConf); _set("k-different",nDiff); _set("k-progress",nRem);
  _set("ks-groups",nG);_set("ks-confirmed",nConf);_set("ks-possible",nDiff);_set("ks-decided",nRem);
}
function renderDash(){
  /* null-safe setter */
  function _set(id,v){var el=document.getElementById(id);if(el)el.textContent=v;}
  function _html(id,v){var el=document.getElementById(id);if(el)el.innerHTML=v;}

  /* ── Topbar KPI strip — values from user decisions, not algorithm tiers ── */
  _set("ks-groups",    D.n_groups         || 0);
  _set("ks-confirmed", D.n_confirmed_dec  || 0);  /* user confirmed as duplicate */
  _set("ks-possible",  D.n_different_dec  || 0);  /* user confirmed as different */
  _set("ks-decided",   D.n_remaining      || 0);  /* still awaiting decision */
  _set("ks-progress",  (D.progress_pct||0)+"%");
  _set("ks-total",     D.sr_count         || 0);

  /* ── Dashboard KPI big cards ── */
  _set("k-groups",      D.n_groups     || 0);
  _set("k-groups-s",    (D.n_dup_srs||0)+" بلاغ متكرر · "+(D.pct_dup||0)+"% من الإجمالي");
  _set("k-progress",    D.n_remaining  || 0);
  _set("k-progress-s",  (D.n_remaining||0)+" مجموعة لم تُراجَع بعد");
  _set("k-confirmed",   D.n_confirmed_dec || 0);
  _set("k-confirmed-s", D.n_confirmed_dec? D.n_confirmed_dec+" جاهزة للإغلاق في Maximo" : "لا يوجد بعد");
  _set("k-different",   D.n_different_dec || 0);
  /* ── Top faults (from server data) ── */
  var tf=D.top_faults||[];
  var tfEl=document.getElementById("k-top-faults-dash");
  if(tfEl) tfEl.innerHTML=tf.slice(0,3).map(function(f){
    return'<div class="dkpi-topitem" onclick="var s=document.getElementById(\'fl-fault\');if(s){s.value=\''+esc(f.name.substring(0,45))+'\';goto(\'alerts\');renderAlerts();}">'
      +'<span class="dkpi-topitem-n">'+f.count+'</span>'
      +'<span class="dkpi-topitem-lbl">'+esc(f.name.substring(0,32))+'</span>'
      +'</div>';
  }).join("")||'<span style="font-size:10px;color:var(--txt4)">لا توجد بيانات</span>';
  /* ── Top locations (computed from groups) ── */
  var locC={};
  G.forEach(function(g){g.members.forEach(function(m){if(m.loc)locC[m.loc]=(locC[m.loc]||0)+1;});});
  var topLocs=Object.entries(locC).sort(function(a,b){return b[1]-a[1];}).slice(0,3);
  var tlEl=document.getElementById("k-top-locs-dash");
  if(tlEl) tlEl.innerHTML=topLocs.map(function(e){
    return'<div class="dkpi-topitem" onclick="var s=document.getElementById(\'fl-loc\');if(s){s.value=\''+esc(e[0])+'\';goto(\'alerts\');renderAlerts();}">'
      +'<span class="dkpi-topitem-n">'+e[1]+'</span>'
      +'<span class="dkpi-topitem-lbl">'+esc(e[0].substring(0,32))+'</span>'
      +'</div>';
  }).join("")||'<span style="font-size:10px;color:var(--txt4)">لا توجد بيانات</span>';
  /* stubs */
  _set("k-top-contractor", "");
  _set("k-top-contractor-s", "");

  /* ── Map + headline KPIs follow the active filters (أ) ── */
  _drawFilteredMapAndKpis();

  /* ── Top priority list (undecided groups, highest score first) ── */
  var top=G.filter(function(g){return!g.decision;}).sort(function(a,b){return (b.score||0)-(a.score||0);}).slice(0,5);
  var dg=document.getElementById("dash-grps"); if(dg)renderList(top,dg);
  var seen={};
  G.forEach(function(g){g.members.forEach(function(m){if(!seen[m.sr])seen[m.sr]=m;});});
  var rec=Object.values(seen).sort(function(a,b){return(b.reported||"")>(a.reported||"")?1:-1;}).slice(0,12);
  var rl=document.getElementById("recent-list");
  if(rl)rl.innerHTML=rec.map(function(m){
    var mdata=JSON.stringify(m).replace(/</g,"\\u003c");
    return'<div class="ri" onclick="openSrModal('+mdata+')">'
      +'<span class="rdot '+stDot(m.status)+'"></span>'
      +'<div style="flex:1"><div class="ri-sr">'+esc(m.sr)+'</div>'
      +'<div class="ri-loc">'+esc([m.loc,m.fault].filter(Boolean).join(" · ").substring(0,55))+'</div></div>'
      +'<div class="ri-time">'+relTime(m.reported)+'</div>'
      +'</div>';
  }).join("");
}

/* ── Leaflet map ── */
var _mapInst = null, _mapLayers = null, _groupLayers = null;
var _mapUserFocused = false;
var _focusedGroupId = null, _lineClicked = false;

/* from comparison card → switch to dashboard, zoom map, focus group */
function mapFocusGroup(gid, lat, lon){
  goto('dash');
  _mapUserFocused = true;          /* lock auto-zoom */
  setTimeout(function(){
    if(_mapInst && lat && lon)
      _mapInst.flyTo([lat, lon], 17, {duration:0.45});
    if(gid) _applyMapFocus(gid);
  }, 200);
}

/* navigate to group card from map popup */
function mapGotoCard(gid){
  if(_mapInst) _mapInst.closePopup();
  goto('alerts');
  setTimeout(function(){
    function findCard(id){
      var c = document.getElementById('gc-'+CSS.escape(id));
      if(c) return c;
      /* try parent group if this was a sub-group id */
      var parent = id.replace(/__sub_\d+$/, '');
      return parent !== id ? document.getElementById('gc-'+CSS.escape(parent)) : null;
    }

    var card = findCard(gid);

    /* Card not in DOM: the status filter is hiding it (e.g. group was
       decided but filter shows only pending).  Clear only the status
       filter — keep location/fault filters as the user set them. */
    if(!card){
      var flStatus = document.getElementById('fl-status');
      if(flStatus && flStatus.value !== '') {
        flStatus.value = '';
        container_renderKey_reset();   /* force re-render on next renderAlerts */
        renderAlerts();
        card = findCard(gid);
      }
    }

    if(!card) return;
    card.scrollIntoView({behavior:'smooth', block:'center'});
    card.style.boxShadow = '0 0 0 3px #0284c7, 0 4px 20px rgba(2,132,199,.25)';
    setTimeout(function(){ card.style.boxShadow = ''; }, 2500);
  }, 200);
}
function container_renderKey_reset(){
  var c = document.getElementById('alerts-list');
  if(c) c._renderKey = null;
}

/* dim all group layers except focused one */
function _applyMapFocus(gid){
  _focusedGroupId = gid;
  if(!_groupLayers) return;
  Object.keys(_groupLayers).forEach(function(id){
    var focused = (id === gid);
    (_groupLayers[id]||[]).forEach(function(layer){
      if(!layer.setStyle) return;
      if(focused){
        layer.setStyle({
          opacity:     layer._baseOp||0.65,
          fillOpacity: layer._baseFillOp!=null ? layer._baseFillOp : 0.85,
          weight:      layer._baseW||1.5
        });
      } else {
        layer.setStyle({opacity:0.08, fillOpacity:0.08, weight:layer._baseW||1.5});
      }
    });
  });
}
function _resetMapFocus(){
  _focusedGroupId = null;
  if(!_groupLayers) return;
  Object.keys(_groupLayers).forEach(function(id){
    (_groupLayers[id]||[]).forEach(function(layer){
      if(layer.setStyle)
        layer.setStyle({
          opacity:     layer._baseOp||0.65,
          fillOpacity: layer._baseFillOp!=null ? layer._baseFillOp : 0.85,
          weight:      layer._baseW||1.5
        });
    });
  });
}

var _mashairViews = {
  all:        {bounds:[[21.335,39.855],[21.435,40.015]]},
  mina:       {center:[21.4135,39.8933], zoom:15},
  muzdalifah: {center:[21.3891,39.9138], zoom:15},
  arafat:     {center:[21.3548,39.9841], zoom:14}
};
function inMashairBounds(p){
  if(!p || !p.lat || !p.lon) return false;
  var b=_mashairViews.all.bounds;
  return p.lat>=b[0][0] && p.lat<=b[1][0] && p.lon>=b[0][1] && p.lon<=b[1][1];
}
function fitMashairOverview(m){
  m.fitBounds(_mashairViews.all.bounds, {padding:[24,24], maxZoom:13});
}
function ensureMap(){
  var el = document.getElementById("livemap");
  if(!el || typeof L === "undefined") return null;
  if(_mapInst) return _mapInst;
  _mapInst = L.map(el, {zoomControl:true, attributionControl:false}).setView([21.389, 39.93], 13);
  L.tileLayer('https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png', {maxZoom:18}).addTo(_mapInst);
  fitMashairOverview(_mapInst);
  /* ── Lock auto-zoom once user touches the map ── */
  _mapInst.on('dragstart', function(){ _mapUserFocused = true; });
  el.addEventListener('wheel', function(){ _mapUserFocused = true; }, {passive:true});
  _mapInst.on('dblclick', function(){ _mapUserFocused = true; });
  /* ── Click background → reset group focus ── */
  _mapInst.on('click', function(){
    if(_lineClicked){ _lineClicked=false; return; }
    if(_focusedGroupId) _resetMapFocus();
  });
  return _mapInst;
}
function zoomMashair(name){
  var m = ensureMap();
  if(!m) return;
  var v = _mashairViews[name] || _mashairViews.all;
  _mapUserFocused = name !== "all";
  if(v.bounds){
    fitMashairOverview(m);
  } else {
    m.flyTo(v.center, v.zoom, {duration:.45});
  }
  setTimeout(function(){ m.invalidateSize(); }, 80);
}
function drawMap(dups, normals, groups){
  var m = ensureMap();
  if(!m) return;
  if(_mapLayers){ _mapLayers.forEach(function(L_){ m.removeLayer(L_); }); }
  _mapLayers = [];
  _groupLayers = {};
  _focusedGroupId = null;
  var bounds = [];

  /* ── single colour for ALL lines — same red as dup markers ── */
  var LINE_CLR = '#dc2626';

  /* ── build SR → groupId lookup for dup markers ── */
  var activeGroups = (groups||[]).filter(function(g){ return g.decision !== 'different'; });
  var srToGid = {};
  activeGroups.forEach(function(g){
    (g.members||[]).forEach(function(mm){ if(mm.sr) srToGid[String(mm.sr)] = g.id; });
  });

  /* ── exclude "different" SR markers ── */
  var excludedSRs = {};
  (groups||[]).forEach(function(g){
    if(g.decision==='different')
      (g.members||[]).forEach(function(mm){ if(mm.sr) excludedSRs[String(mm.sr)]=1; });
  });
  var activeDups = (dups||[]).filter(function(p){ return !excludedSRs[String(p.sr||'')]; });

  /* ── reusable group popup builder ── */
  function groupPopup(g, distTxt){
    var gidSafe = (g.id||'').replace(/'/g,"\\'");
    var srs = (g.members||[]).map(function(mm){return mm.sr||'';}).filter(Boolean).join(' · ');
    var fault = esc(((g.members||[])[0]||{}).fault||'');
    var tierLabel = g.tier==='confirmed'?'مؤكد':g.tier==='possible'?'محتمل':'ضعيف';
    return '<div class="map-popup">'
      +'<div class="mp-sr">'+tierLabel+' · '+g.score+' نقطة</div>'
      +'<div class="mp-loc">'+esc(srs)+'</div>'
      +(fault?'<div class="mp-fault">'+fault+'</div>':'')
      +(distTxt?'<div class="mp-dist">📍 '+distTxt+'</div>':'')
      +'<button class="mp-goto" onclick="mapGotoCard(\''+gidSafe+'\')">↩ عرض كرت المجموعة</button>'
      +'</div>';
  }

  function addFocusClick(layer, gid){
    layer.on('click', function(e){
      L.DomEvent.stopPropagation(e);
      _lineClicked = true;
      _applyMapFocus(gid);
    });
  }

  /* ── 1. Network lines (below markers) ── */
  activeGroups.forEach(function(g){
    var gid = g.id || '';
    if(!_groupLayers[gid]) _groupLayers[gid] = [];

    var memberPts = (g.members||[]).map(function(mm){
      var la=parseFloat(mm.lat||0), lo=parseFloat(mm.lon||0);
      return (la&&lo)?[la,lo]:null;
    }).filter(Boolean);
    if(memberPts.length < 2) return;

    var origin = memberPts[0];
    for(var i=1;i<memberPts.length;i++){
      var dest = memberPts[i];
      var dist = haversineM(origin[0],origin[1],dest[0],dest[1]);
      var distTxt = dist!=null ? fmtDist(dist) : null;

      var line = L.polyline([origin, dest], {
        color: LINE_CLR, weight: 2, opacity: 0.7,
        dashArray: '6,4', lineCap:'round', lineJoin:'round'
      }).bindPopup(groupPopup(g, distTxt));

      /* permanent distance label — always visible on the line */
      if(distTxt){
        line.bindTooltip(distTxt, {
          permanent:true, direction:'center',
          className:'map-line-dist', interactive:false, opacity:0.9
        });
      }

      line._baseOp=0.6; line._baseFillOp=null; line._baseW=1.5;
      addFocusClick(line, gid);
      line.addTo(m); _mapLayers.push(line); _groupLayers[gid].push(line);
    }
  });

  /* ── 2. Normal markers ── */
  (normals||[]).forEach(function(p){
    var pop = '<div class="map-popup"><div class="mp-sr">SR '+esc(p.sr||'')+'</div>'
      +'<div class="mp-loc">'+esc(p.loc||'غير محدد')+'</div>'
      +(p.fault?'<div class="mp-fault">'+esc(p.fault)+'</div>':'')
      +'<br><a class="mp-link" href="'+maximoUrl(p.sr)+'" target="_blank" rel="noopener">فتح في Maximo</a>'
      +'</div>';
    var c = L.circleMarker([p.lat,p.lon], {
      radius:3, color:'#b9975b', fillColor:'#d4b896', fillOpacity:0.55, weight:1
    }).bindPopup(pop).bindTooltip('SR '+esc(p.sr||''), {direction:'top', opacity:.9});
    c.addTo(m); _mapLayers.push(c);
    bounds.push([p.lat,p.lon]);
  });

  /* ── 3. Dup markers — click focuses their group ── */
  activeDups.forEach(function(p){
    var gid = srToGid[String(p.sr||'')] || null;
    var g   = gid ? activeGroups.filter(function(ag){ return ag.id===gid; })[0] : null;

    /* popup: show group info + goto button if belongs to a group */
    var pop = g ? groupPopup(g, null)
      : ('<div class="map-popup"><div class="mp-sr">SR '+esc(p.sr||'')+'</div>'
        +'<div class="mp-loc">'+esc(p.loc||'غير محدد')+'</div>'
        +(p.fault?'<div class="mp-fault">'+esc(p.fault)+'</div>':'')
        +'<a class="mp-link" href="'+maximoUrl(p.sr)+'" target="_blank" rel="noopener">فتح في Maximo</a>'
        +'</div>');

    var c = L.circleMarker([p.lat,p.lon], {
      radius:6, color:'#991b1b', fillColor:'#dc2626', fillOpacity:0.85, weight:1.5
    }).bindPopup(pop)
      .bindTooltip('SR '+esc(p.sr||'')+' · مكرر', {direction:'top', opacity:.95});

    c._baseOp=0.85; c._baseFillOp=0.85; c._baseW=1.5;
    if(gid){ addFocusClick(c, gid); }
    c.addTo(m); _mapLayers.push(c);
    if(gid && _groupLayers[gid]) _groupLayers[gid].push(c);
    bounds.push([p.lat,p.lon]);
  });

  if(!_mapUserFocused){
    var localBounds = activeDups.concat(normals||[]).filter(inMashairBounds).map(function(p){return [p.lat,p.lon];});
    try{
      if(localBounds.length){
        m.fitBounds(localBounds, {padding:[24,24], maxZoom:14, animate:false});
      } else {
        fitMashairOverview(m);
      }
    }catch(e){ fitMashairOverview(m); }
  }
  setTimeout(function(){ m.invalidateSize(); }, 30);
}
/* initialise the map immediately on first load, even before scan data arrives */
window.addEventListener("load", function(){
  ensureMap();
  /* re-measure after fonts/layout settle — critical for flex-based map container */
  setTimeout(function(){ if(_mapInst) _mapInst.invalidateSize(); }, 200);
  setTimeout(function(){ if(_mapInst) _mapInst.invalidateSize(); }, 600);
});

/* ── decisions ── */
async function decide(gid,decision){
  var eid=CSS.escape(gid);
  var noteEl=document.getElementById("ni-"+eid);
  var note=noteEl?noteEl.value:"";
  try{
    var g=G.find(function(x){return x.id===gid;});
    var srs=g?(g.members||[]).map(function(m){return m.sr;}).filter(Boolean):[];
    var parent=g&&g.parent_id?g.parent_id:gid;
    var r=await fetch("/api/decision",{method:"POST",headers:{"Content-Type":"application/json"},body:JSON.stringify({gid:parent,decision:decision,note:note,srs:srs})});
    if(!(await r.json()).ok)throw 0;
    var labels={duplicate:" تم تأكيد التكرار",different:" تم الاستبعاد"};
    toast(labels[decision]||"تم","ok","");
    await loadData(true);
    var pend=G.filter(function(g){return!g.decision;}).length;
    function _set(id, v){var el=document.getElementById(id); if(el) el.textContent=v;}
    _set("sb-pend", pend);
    _set("sb-conf", G.filter(function(g){return g.decision==="duplicate";}).length);
    _set("sb-rej",  G.filter(function(g){return g.decision==="different";}).length);
    setAlertBadges(pend);
    loadNotifications();
  }catch(e){toast("خطأ في الحفظ","err","");}
}
async function decideSelected(gid,decision){
  var eid=CSS.escape(gid);
  var srs=getChecked(eid);
  if(!srs.length){toast("اختر بلاغاً على الأقل","err","");return;}
  var noteEl=document.getElementById("ni-"+eid);
  var note=noteEl?noteEl.value:"";
  try{
    var r=await fetch("/api/decision",{method:"POST",headers:{"Content-Type":"application/json"},body:JSON.stringify({gid:gid,decision:decision,note:note,srs:srs})});
    if(!(await r.json()).ok)throw 0;
    var lbl=decision==="duplicate"?" تم تأكيد المحددة":" تم رفض المحددة";
    toast(lbl,"ok","");
    toggleSelMode(eid);
    await loadData(true);
    loadNotifications();
  }catch(e){toast("خطأ في الحفظ","err","");}
}
async function delSubDec(eid,gid,idx){
  try{
    await fetch("/api/decision/"+encodeURIComponent(gid)+"?idx="+idx,{method:"DELETE"});
    await loadData(true);
    loadNotifications();
  }catch(e){toast("خطأ","err","");}
}
async function undoDecision(gid){
  try{
    await fetch("/api/decision/"+encodeURIComponent(gid),{method:"DELETE"});
    toast("تم التراجع عن القرار","ok","");
    await loadData(true);
    loadNotifications();
  }catch(e){toast("تعذر التراجع","err","");}
}

/* ── SR modal ── */
function openSrModal(m){
  var mu=maximoUrl(m.sr);
  document.getElementById("sr-modal-t").textContent="تفاصيل البلاغ — "+m.sr;
  document.getElementById("sr-modal-b").innerHTML=
    '<div class="mfg">'
    +'<div class="mf"><div class="mf-l">رقم البلاغ</div><div class="mf-v" style="font-size:18px;font-weight:800;color:var(--gold-dk)">'+esc(m.sr)+'</div></div>'
    +'<div class="mf"><div class="mf-l">الحالة</div><div class="mf-v">'+stBadge(m.status)+'</div></div>'
    +'<div class="mf"><div class="mf-l">تاريخ الفتح</div><div class="mf-v">'+esc(m.reported||"—")+'</div></div>'
    +'<div class="mf"><div class="mf-l">الموقع</div><div class="mf-v">'+esc(m.loc||"—")+'</div></div>'
    +'<div class="mf"><div class="mf-l">الأصل (Asset)</div><div class="mf-v">'+esc(m.asset||"—")+'</div></div>'
    +'<div class="mf wide"><div class="mf-l">نوع العطل</div><div class="mf-v">'+esc(m.fault||"—")+'</div></div>'
    +'<div class="mf wide"><div class="mf-l">التفاصيل</div><div class="mf-v" style="white-space:pre-wrap;font-size:11px">'+esc(m.detail||"—")+'</div></div>'
    +'</div>'
    +'<a class="mx-btn" href="'+mu+'" target="_blank" rel="noopener"> فتح في Maximo</a>';
  document.getElementById("sr-modal").classList.add("open");
}

/* ── analytics ── */
async function loadAnalytics(){
  document.getElementById("ana-content").innerHTML='<div class="empty"><div class="ei"></div><h3>جارٍ التحميل…</h3></div>';
  try{
    var d=await(await fetch("/api/analytics")).json();
    if(!d.ready){
      document.getElementById("ana-content").innerHTML='<div class="empty"><div class="ei"></div><h3>لا توجد بيانات</h3>'
        +'<p>اضغط "سحب من Maximo" لبدء الفحص</p>'
        +'<button class="hero-btn" onclick="goto(\'upload\');loadUploadInfo();startMaximoScan()" style="margin-top:8px"> سحب من Maximo الآن</button></div>';
      return;
    }
    if(d.scanning||d.total_groups===0){
      var msg=d.scanning?(' '+esc(d.scan_phase||"جارٍ الفحص…")):"لا توجد مجموعات مكررة بعد";
      document.getElementById("ana-content").innerHTML='<div class="empty"><div class="ei">'+(d.scanning?"":"")+'</div><h3>'+msg+'</h3>'
        +(d.scanning?'<p>سيتم تحديث الإحصاءات عند الانتهاء</p>':'<p>البلاغات المفتوحة: '+d.total_srs+' · لا توجد تكرارات مكتشفة</p>')
        +'</div>';
      if(d.scanning)setTimeout(loadAnalytics,5000);
      return;
    }
    var tf=d.tiers,dc=d.decisions,ts=d.time_spread;
    var mxF=d.top_faults.length?d.top_faults[0].count:1,mxL=d.top_locs.length?d.top_locs[0].count:1;
    function bars(items,mx){return items.map(function(x){return'<div class="bar-row"><span class="bar-lbl" title="'+esc(x.label)+'">'+esc(x.label)+'</span><div class="bar-out"><div class="bar-in" style="width:'+Math.round(x.count*100/mx)+'%"></div></div><span class="bar-n">'+x.count+'</span></div>';}).join("");}
    document.getElementById("ana-content").innerHTML=
      '<div class="ana-grid">'
      +'<div class="ana-card"><div class="ana-t">حالة المراجعة</div>'
        +'<div class="spills"><div class="spill green"><span class="spill-n">'+(dc.duplicate||0)+'</span><div class="spill-l"> مكرر</div></div><div class="spill red"><span class="spill-n">'+(dc.different||0)+'</span><div class="spill-l"> مختلف</div></div><div class="spill amber"><span class="spill-n">'+(dc.pending||0)+'</span><div class="spill-l"> معلق</div></div></div>'
        +'<div style="margin-top:10px;font-size:11px;color:var(--txt3)">إجمالي: <b>'+d.total_groups+'</b> مجموعة · <b>'+d.total_srs+'</b> بلاغ</div>'
      +'</div>'
      +'<div class="ana-card"><div class="ana-t">أكثر الأعطال تكراراً</div>'+bars(d.top_faults,mxF)+'</div>'
      +'<div class="ana-card"><div class="ana-t">أكثر المواقع تكراراً</div>'+bars(d.top_locs,mxL)+'</div>'
      +'</div>';
  }catch(e){document.getElementById("ana-content").innerHTML="<p style='padding:20px;color:var(--red)'>خطأ في التحميل</p>";}
}

/* ── Reports page ── */
var _allSRs=[], _filteredSRs=[], _rptOffset=0;
async function loadReports(){
  document.getElementById("rpt-count").textContent="جارٍ التحميل…";
  try{
    var d=await(await fetch("/api/reports?t="+Date.now(),{cache:"no-store"})).json();
    if(!d.ready){
      document.getElementById("rpt-tbody").innerHTML='<tr><td colspan="7" style="text-align:center;padding:24px"><button class="hero-btn" onclick="goto(\'upload\')"> ارفع ملف Excel أولاً</button></td></tr>';
      document.getElementById("rpt-count").textContent="لا توجد بيانات";
      return;
    }
    _allSRs=d.srs||[];
    // populate filter dropdowns
    var statuses=new Set(),locs=new Set();
    _allSRs.forEach(function(r){if(r.status)statuses.add(r.status);if(r.loc)locs.add(r.loc);});
    var rptStatus=document.getElementById("rpt-status");
    var rptLoc=document.getElementById("rpt-loc");
    rptStatus.innerHTML='<option value="">كل الحالات</option>';
    rptLoc.innerHTML='<option value="">كل المواقع</option>';
    [...statuses].sort().forEach(function(s){var o=document.createElement("option");o.value=s;o.textContent=s;rptStatus.appendChild(o);});
    [...locs].sort().forEach(function(s){var o=document.createElement("option");o.value=s;o.textContent=s;rptLoc.appendChild(o);});
    filterReports();
  }catch(e){document.getElementById("rpt-count").textContent="خطأ في التحميل";}
}
function filterReports(){
  var q=document.getElementById("rpt-q").value.toLowerCase().trim();
  var sf=document.getElementById("rpt-status").value;
  var lf=document.getElementById("rpt-loc").value;
  var gf=document.getElementById("rpt-grp").value;
  _filteredSRs=_allSRs.filter(function(r){
    if(q&&![r.sr,r.loc,r.asset,r.fault].join(" ").toLowerCase().includes(q))return false;
    if(sf&&r.status!==sf)return false;
    if(lf&&r.loc!==lf)return false;
    if(gf==="in_group"&&!r.group_num)return false;
    if(gf==="no_group"&&r.group_num)return false;
    return true;
  });
  _rptOffset=0;
  document.getElementById("rpt-tbody").innerHTML="";
  renderReportRows();
  var shown=Math.min(200,_filteredSRs.length);
  document.getElementById("rpt-count").textContent="عرض "+shown.toLocaleString("ar-SA")+" من "+_filteredSRs.length.toLocaleString("ar-SA")+" بلاغ";
}
function renderReportRows(){
  var tbody=document.getElementById("rpt-tbody");
  var chunk=_filteredSRs.slice(_rptOffset,_rptOffset+200);
  _rptOffset+=chunk.length;
  chunk.forEach(function(r){
    var tr=document.createElement("tr");
    var groupCell=r.group_num?'<span class="rpt-grp-badge '+(r.tier||"")+'" style="cursor:pointer;text-decoration:underline dotted" onclick="gotoGroup(\''+r.sr+'\')" title="انتقل إلى المجموعة">#'+r.group_num+'</span>':"—";
    var mdata=JSON.stringify(r).replace(/</g,"\\u003c");
    tr.innerHTML=
      '<td><span class="rpt-sr-num" onclick="openSrModal('+mdata+')">'+esc(r.sr)+'</span></td>'
      +'<td>'+stBadge(r.status)+'</td>'
      +'<td>'+esc(r.loc||"—")+'</td>'
      +'<td>'+esc(r.asset||"—")+'</td>'
      +'<td class="rpt-fault" title="'+esc(r.fault||"")+'">'+esc((r.fault||"").substring(0,50))+'</td>'
      +'<td>'+esc(r.reported||"—")+'</td>'
      +'<td>'+groupCell+'</td>';
    tbody.appendChild(tr);
  });
  document.getElementById("rpt-more").style.display=_filteredSRs.length>_rptOffset?"block":"none";
}
function rptLoadMore(){
  renderReportRows();
  var shown=Math.min(_rptOffset,_filteredSRs.length);
  document.getElementById("rpt-count").textContent="عرض "+shown.toLocaleString("ar-SA")+" من "+_filteredSRs.length.toLocaleString("ar-SA")+" بلاغ";
}

/* ── Data source page ── */
var _mxPollTimer=null;
function loadUploadInfo(){
  fetch("/api/scan-status").then(function(r){return r.json();}).then(function(d){
    // Maximo connection badge
    var badge=document.getElementById("mx-status-badge");
    var urlRow=document.getElementById("mx-url-row");
    if(d.has_credentials){
      badge.textContent=" مكوّن";badge.style.background="var(--grn-bg)";badge.style.color="var(--green)";badge.style.border="1px solid var(--grn-bd)";
      if(d.maximo_url){urlRow.style.display="block";document.getElementById("mx-url-txt").textContent=d.maximo_url;}
    }else{
      badge.textContent=" غير مكوّن";badge.style.background="var(--red-bg)";badge.style.color="var(--red)";badge.style.border="1px solid var(--red-bd)";
      document.getElementById("btn-mx-scan").disabled=true;
    }
    // Running state
    if(d.running){_startMxPoll();}
    // Last scan info
    var ls=d.last_scan||{};
    document.getElementById("ul-ready").textContent=ls.sr_count?" جاهز":" لا توجد بيانات";
    document.getElementById("ul-source").textContent=ls.source==="maximo"?" Maximo":ls.source==="file"?" ملف Excel":ls.source||"—";
    document.getElementById("ul-date").textContent=ls.scanned_at||"—";
    document.getElementById("ul-srs").textContent=ls.sr_count?ls.sr_count.toLocaleString("ar-SA"):"—";
    document.getElementById("ul-grps").textContent=ls.group_count?ls.group_count+" مجموعة":"—";
  }).catch(function(){});
}
function startMaximoScan(){
  document.getElementById("mx-error").style.display="none";
  document.getElementById("mx-result").style.display="none";
  document.getElementById("btn-mx-scan").disabled=true;
  document.getElementById("btn-mx-scan").textContent=" جارٍ الفحص…";
  var mdEl=document.getElementById("mx-maxdays");
  var maxDays=mdEl?parseInt(mdEl.value||"2",10):2;
  fetch("/api/scan-maximo",{method:"POST",headers:{"Content-Type":"application/json"},body:JSON.stringify({max_days:maxDays})}).then(function(r){return r.json();}).then(function(d){
    if(!d.ok){
      document.getElementById("mx-error").textContent=d.message||"خطأ";
      document.getElementById("mx-error").style.display="block";
      document.getElementById("btn-mx-scan").disabled=false;
      document.getElementById("btn-mx-scan").textContent=" سحب من Maximo الآن";
      return;
    }
    toast("بدأ الفحص من Maximo","ok","");
    _scanRunning=true;
    _ensureStatusPolling();
    _startMxPoll();
  }).catch(function(e){
    document.getElementById("mx-error").textContent="خطأ في الاتصال: "+e;
    document.getElementById("mx-error").style.display="block";
    document.getElementById("btn-mx-scan").disabled=false;
    document.getElementById("btn-mx-scan").textContent=" سحب من Maximo الآن";
  });
}
function _startMxPoll(){
  clearInterval(_mxPollTimer);
  document.getElementById("mx-scan-prog").style.display="block";
  var pct=5;
  var fill=document.getElementById("mx-prog-fill");
  var txt=document.getElementById("mx-prog-txt");
  _mxPollTimer=setInterval(function(){
    fetch("/api/scan-status").then(function(r){return r.json();}).then(function(d){
      var prog=d.progress||{};
      txt.textContent=prog.phase||"جارٍ…";
      if(d.running){pct=Math.min(pct+3,88);fill.style.width=pct+"%";}
      else{
        clearInterval(_mxPollTimer);
        fill.style.width="100%";
        document.getElementById("btn-mx-scan").disabled=false;
        document.getElementById("btn-mx-scan").textContent=" سحب من Maximo الآن";
        document.getElementById("mx-scan-prog").style.display="none";
        if(d.error){
          document.getElementById("mx-error").textContent=d.error;
          document.getElementById("mx-error").style.display="block";
        }else if(prog.fetched){
          document.getElementById("mx-result").innerHTML=
            " اكتمل! جُلب <b>"+prog.fetched.toLocaleString("ar-SA")+"</b> بلاغ"
            +" · <b>"+(prog.groups||0)+"</b> مجموعة مكررة"
            +"<br><button class='hero-btn' style='margin-top:10px' onclick=\"goto('dash');loadData(true)\"> الانتقال إلى الداشبورد</button>";
          document.getElementById("mx-result").style.display="block";
          loadUploadInfo();
          loadData();
        }
      }
    });
  },2500);
}
function handleFile(file){
  if(!file)return;
  if(!file.name.match(/\.(xlsx|xls)$/i)){toast("يجب رفع ملف Excel فقط","err","");return;}
  document.getElementById("upd-result").style.display="none";
  var prog=document.getElementById("upd-prog");
  prog.style.display="block";
  document.getElementById("prog-fill").style.width="5%";
  document.getElementById("prog-txt").textContent="جارٍ الرفع… (قد يستغرق 1–3 دقائق حسب حجم الملف)";
  var pct=5;
  var pint=setInterval(function(){pct=Math.min(pct+1.5,88);document.getElementById("prog-fill").style.width=pct+"%";},600);
  var fd=new FormData();
  fd.append("file",file);
  fetch("/api/upload",{method:"POST",body:fd})
    .then(function(r){return r.json();})
    .then(function(d){
      clearInterval(pint);
      document.getElementById("prog-fill").style.width="100%";
      setTimeout(function(){prog.style.display="none";},600);
      if(d.detail||d.error){
        document.getElementById("prog-txt").textContent="خطأ: "+(d.detail||d.error);
        prog.style.display="block";
        return;
      }
      var res=document.getElementById("upd-result");
      res.style.display="block";
      res.innerHTML='<div class="upd-ok"> اكتمل الفحص بنجاح!</div>'
        +'<div class="upd-stats">البلاغات: <b>'+d.sr_count.toLocaleString("ar-SA")+'</b>'
        +' · المجموعات: <b>'+d.group_count+'</b>'
        +' · الأزواج: <b>'+d.pair_count+'</b></div>'
        +'<button class="hero-btn" style="margin-top:14px" onclick="goto(\'dash\');loadData(true)"> الانتقال إلى الداشبورد</button>';
      loadData();
      loadUploadInfo();
    })
    .catch(function(e){
      clearInterval(pint);
      document.getElementById("prog-txt").textContent="خطأ في الاتصال: "+e;
    });
}
function handleDrop(e){
  e.preventDefault();
  document.getElementById("drop-zone").classList.remove("drag-over");
  var file=e.dataTransfer.files[0];
  if(file)handleFile(file);
}

/* ── export ── */
function exportExcel(){
  var b=document.getElementById("btn-exp");b.textContent=" جارٍ…";b.disabled=true;
  var a=document.createElement("a");a.href="/api/export/excel";a.download="";document.body.appendChild(a);a.click();document.body.removeChild(a);
  setTimeout(function(){b.textContent=" تصدير Excel";b.disabled=false;},3000);
}

/* ── live alerts polling ── */
var _lat="";
async function pollAlerts(){
  try{
    var url="/api/alerts"+(_lat?"?since="+encodeURIComponent(_lat):"");
    var d=await(await fetch(url)).json();
    var al=d.alerts||[];
    if(al.length){
      _lat=al[0].detected_at;
      loadNotifications();
      al.slice(0,3).forEach(function(a){
        var el=document.createElement("div");
        el.className="toast alert";
        el.innerHTML='<span class="ti"></span><span>تكرار محتمل مكتشف<br><b>'+a.new_sr+'</b> · تطابق '+scorePct(a.score)+'%</span>';
        document.getElementById("toasts").appendChild(el);
        setTimeout(function(){el.remove();},6000);
      });
    }
  }catch(e){}
  setTimeout(pollAlerts,15000);
}

/* ── init ── */
document.getElementById("tb-date").textContent=new Date().toLocaleDateString("ar-SA",{weekday:"long",year:"numeric",month:"long",day:"numeric"});
// Check if a background scan is already running (auto-started by server)
fetch("/api/scan-status").then(function(r){return r.json();}).then(function(d){
  _scanRunning=d.running||false;
  _scanPhase=(d.progress||{}).phase||"";
  if(d.running){
    _ensureStatusPolling();
    toast("جارٍ فحص Maximo تلقائياً…","alert","");
  }
  loadData();
}).catch(function(){loadData();});
setTimeout(pollAlerts,8000);
loadNotifications();
/* ── Request browser notification permission after 3 s (non-intrusive) ── */
setTimeout(_requestNotifPermission, 3000);
/* ── Start new-groups poller ── */
setTimeout(_pollNewGroups, 12000);
</script>
<!-- notification panel at body level — outside any stacking context -->
<div class="notif-panel" id="notif-panel">
  <div class="notif-head">
    <span class="notif-title">التنبيهات</span>
    <span class="notif-total" id="notif-total" style="display:none">0</span>
    <button onclick="_requestNotifPermission();toast('تم تفعيل الإشعارات','ok','')" title="تفعيل إشعارات المتصفح" style="margin-right:0;background:none;border:1px solid var(--border);border-radius:5px;cursor:pointer;font-size:10px;color:var(--txt2);padding:2px 7px;white-space:nowrap">🔔 تفعيل</button>
    <button onclick="document.getElementById('notif-panel').classList.remove('open')" style="margin-right:auto;background:none;border:none;cursor:pointer;font-size:14px;color:var(--txt3);padding:0 4px" title="إغلاق">✕</button>
  </div>
  <div id="notif-email-row" style="display:none;padding:6px 12px;background:var(--grn-bg);border-bottom:1px solid var(--grn-bd);font-size:10px;color:var(--green);font-weight:700">
    📧 إشعارات البريد الإلكتروني مفعّلة
  </div>
  <div class="notif-list" id="notif-list">
    <div class="notif-empty">لا توجد تنبيهات مفتوحة</div>
  </div>
</div>
</body>
</html>"""
