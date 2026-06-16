#!/usr/bin/env python3
"""
find_duplicates.py v3 -- كشف البلاغات المكررة في ماكسيمو (Dropdowns + التفاصيل)

نظام النقاط:
  نفس العطل (fault_category)              +3   (من Summary dropdown)
  نفس الموقع (LOCATION)                   +4   (dropdown موثوق)
  نفس الأصل (Asset)                       +4   (dropdown موثوق)
  مقارنة ذكية للتفاصيل (Details):
    identical — قالب+أرقام متطابقة        +5
    similar   — قالب متشابه >= 90%        +3
    template_only — قالب نفس/أرقام مختلفة +0  (يمنع الإيجابيات الزائفة)

ملاحظة: الزمن لا يدخل في النقاط — البلاغ المكرر قد يُفتح بعد ساعات أو أشهر.
         الفارق الزمني يُعرض فقط لمساعدة المراجع على الفهم.

المستويات:
  مؤكد  >= 8 نقاط
  محتمل 5-7 نقاط
  ضعيف 2-4 نقاط

التجميع: Union-Find يدمج الأزواج المتداخلة في مجموعات.

الاستخدام:
  python scripts/find_duplicates.py البلاغات.xlsx
  python scripts/find_duplicates.py البلاغات.xlsx --output تقرير.xlsx --min-score 2
"""

import argparse
import difflib
import io
import re
import itertools
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd

# Reference data lives under <repo>/data/ in this layout. The legacy
# behaviour kept the files at the repo root; we redirect here.
_REF_DIR = Path(__file__).resolve().parents[3] / "data"
ASSET_MAP_FILE = _REF_DIR / "asset_description.xls"
LOC_MAP_FILE = _REF_DIR / "location_description.xls"


def _load_ref(path: Path, key_col: str, val_col: str) -> dict:
    """يحمّل ملف مرجعي HTML-XLS ويرجع قاموس {كود → وصف عربي}.

    ملفات ماكسيمو تُعلن UTF-8 في الـ meta tag لكن البيانات الفعلية مشفّرة
    بـ cp1256 (ويندوز عربي). نمرر BytesIO مباشرة لـ pd.read_html مع تحديد
    الترميز الصحيح بدلاً من فكّ التشفير يدوياً.
    """
    if not path.exists():
        return {}
    try:
        import io as _io

        raw = path.read_bytes()
        # نجرب الترميزات بالترتيب — utf-8 أولاً (ملفات Maximo HTML-XLS)
        # ثم cp1256 كـ fallback للملفات القديمة
        tables = None
        for enc in ("utf-8", "utf-8-sig", "cp1256", "windows-1256"):
            try:
                tbls = pd.read_html(_io.BytesIO(raw), encoding=enc)
                if tbls:
                    # نتحقق أن البيانات تحتوي أحرفاً عربية فعلية (U+0600–U+06FF)
                    sample = str(tbls[0].iloc[:5].to_string())
                    if any("؀" <= c <= "ۿ" for c in sample):
                        tables = tbls
                        break
            except Exception:
                continue
        if tables is None:
            # آخر محاولة بدون تحقق
            tables = pd.read_html(_io.BytesIO(raw), encoding="utf-8")
        if not tables:
            return {}
        df = tables[0]
        if all(isinstance(c, (int, float)) for c in df.columns):
            df.columns = df.iloc[0].astype(str).str.strip()
            df = df.iloc[1:].reset_index(drop=True)
        for col in df.select_dtypes("object").columns:
            df[col] = (
                df[col]
                .astype(str)
                .str.replace("\xa0", " ", regex=False)  # non-breaking space
                .str.replace("\xad", "", regex=False)  # soft hyphen
                .str.replace("‏", "", regex=False)  # RLM
                .str.replace("‎", "", regex=False)  # LRM
                .str.strip()
            )
        cols = {c.strip().lower(): c for c in df.columns}
        k = cols.get(key_col.lower(), key_col)
        v = cols.get(val_col.lower(), val_col)
        if k not in df.columns or v not in df.columns:
            return {}
        return dict(zip(df[k].astype(str).str.strip(), df[v].astype(str)))
    except Exception:
        return {}


def _load_ref_json(json_gz_path: Path) -> dict:
    """يحمّل قاموس {كود → وصف} من ملف JSON مضغوط (.json.gz) كـ fallback."""
    if not json_gz_path.exists():
        return {}
    try:
        import gzip as _gz, json as _json

        with _gz.open(json_gz_path, "rt", encoding="utf-8") as f:
            return _json.load(f)
    except Exception:
        return {}


def _load_ref_auto(xls_path: Path, json_gz_path: Path, key_col: str, val_col: str) -> dict:
    """يجرب XLS أولاً، ثم JSON.gz كـ fallback (للـ Cloud حين XLS غير موجود)."""
    if xls_path.exists():
        m = _load_ref(xls_path, key_col, val_col)
        if m:
            return m
    return _load_ref_json(json_gz_path)


# يُحمَّل مرة واحدة عند استيراد الموديول
_ASSET_MAP: dict = _load_ref_auto(
    ASSET_MAP_FILE,
    _REF_DIR / "ref_asset.json.gz",
    "Asset",
    "Description",
)
_LOC_MAP: dict = _load_ref_auto(
    LOC_MAP_FILE,
    _REF_DIR / "ref_loc.json.gz",
    "Location",
    "Description",
)


def load_ref_from_bytes(asset_bytes: bytes | None, loc_bytes: bytes | None) -> None:
    """يُحدِّث الخرائط المرجعية من bytes مباشرة (للـ Streamlit Cloud حين الملفات غير موجودة).

    استخدام:
        import find_duplicates as fd
        fd.load_ref_from_bytes(asset_xls_bytes, loc_xls_bytes)
    """
    import tempfile, io as _io

    global _ASSET_MAP, _LOC_MAP

    def _from_bytes(data: bytes, key_col: str, val_col: str) -> dict:
        suf = ".xls"
        with tempfile.NamedTemporaryFile(suffix=suf, delete=False) as tf:
            tf.write(data)
            tmp = Path(tf.name)
        try:
            return _load_ref(tmp, key_col, val_col)
        finally:
            tmp.unlink(missing_ok=True)

    if asset_bytes:
        m = _from_bytes(asset_bytes, "Asset", "Description")
        if m:
            _ASSET_MAP = m
    if loc_bytes:
        m = _from_bytes(loc_bytes, "Location", "Description")
        if m:
            _LOC_MAP = m


# openpyxl يُستورد فقط عند الحاجة (داخل write_report) لتفادي مشاكل التبعيات
# في FastAPI server الذي لا يحتاج openpyxl.


# ─────────────────────────────────────────────────────────────────────────────
# helpers
# ─────────────────────────────────────────────────────────────────────────────

_MAXIMO_MARKERS = {
    "Service Request",
    "Site",
    "Status",
    "Summary",
    "LOCATION",
    "Details",
    "Work Zone",
    "Asset",
    "History",
    "Source",
}


def _str(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() in ("nan", "none", "null", "<na>") else s


_HTML_TAG_RE = re.compile(r"<[^>]+>")
_HTML_CMT_RE = re.compile(r"<!--.*?-->", re.DOTALL)


def _strip_html(s: str) -> str:
    """يزيل وسوم HTML والتعليقات من نص التفاصيل (Maximo Rich Text)."""
    if not s:
        return s
    s = _HTML_CMT_RE.sub(" ", s)
    s = _HTML_TAG_RE.sub(" ", s)
    s = s.replace("&nbsp;", " ").replace("&amp;", "&").replace("&lt;", "<").replace("&gt;", ">")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _normalize_ar(s: str) -> str:
    """توحيد النص العربي: حذف \\xa0، توحيد الهمزات، تصغير الإنجليزية."""
    if not s:
        return ""
    s = s.replace("\xa0", " ").strip()
    # توحيد الهمزات
    s = re.sub(r"[إأآا]", "ا", s)
    s = s.replace("ى", "ي").replace("ة", "ه")
    # تصغير الحروف الإنجليزية
    s = re.sub(r"\s+", " ", s).lower()
    return s


def _fix_double_header(df):
    if len(df) == 0:
        return df
    first = {str(v).strip() for v in df.iloc[0].values}
    if len(first & _MAXIMO_MARKERS) >= 2:
        new_cols = []
        for i, v in enumerate(df.iloc[0].values):
            s = str(v).strip()
            new_cols.append(s if s and s.lower() != "nan" else str(df.columns[i]))
        df.columns = new_cols
        df = df.iloc[1:].reset_index(drop=True)
    return df


def read_file(path: str) -> pd.DataFrame:
    data = Path(path).read_bytes()
    for enc in ("utf-8", "windows-1256", "utf-8-sig"):
        try:
            tables = pd.read_html(io.BytesIO(data), encoding=enc, flavor="lxml")
            if tables and len(tables[0]) > 0:
                return _fix_double_header(tables[0])
        except Exception:
            pass
    for engine in ("openpyxl", "xlrd"):
        try:
            df = pd.read_excel(io.BytesIO(data), engine=engine, dtype=str)
            return _fix_double_header(df)
        except Exception:
            pass
    raise SystemExit(f"لا يمكن قراءة الملف: {path}")


def _find_col(df, *names):
    low = {str(c).strip().lower(): str(c) for c in df.columns}
    for n in names:
        if n in df.columns:
            return n
        if n.lower() in low:
            return low[n.lower()]
    return None


def _fault(summary: str) -> str:
    """يأخذ آخر جزء من Summary (نوع العطل)."""
    parts = [p.strip() for p in summary.split(",") if p.strip()]
    return _normalize_ar(parts[-1]) if parts else _normalize_ar(summary)


# ─────────────────────────────────────────────────────────────────────────────
# مقارنة النص الذكية: قالب + أرقام
# ─────────────────────────────────────────────────────────────────────────────

_NUM_RE = re.compile(r"\d+(?:/\d+)?")  # أرقام عادية + شواخص X/Y


def _smart_text_compare(a: str, b: str) -> tuple[str, int, int]:
    """
    يقارن نصّين بطريقة ذكية تفصل بين قالب الجملة والأرقام داخلها.
    يرجّع: (التصنيف، النقاط، نسبة تشابه القالب %)

    التصنيف:
      'identical'     = قالب 90%+ AND أرقام متطابقة ≥ 50%   → +2 (مكرر فعلي)
      'similar'       = قالب 90%+ بدون شرط الأرقام         → +3
      'template_only' = قالب 90%+ لكن أرقام مختلفة (< 30%) → 0 ⚠️
      'different'     = قالب < 90%                          → 0
    """
    # ١) قالب النص بدون أرقام (نستبدل كل رقم بـ #)
    a_norm = _normalize_ar(_strip_html(a or ""))
    b_norm = _normalize_ar(_strip_html(b or ""))
    a_template = _NUM_RE.sub("#", a_norm)
    b_template = _NUM_RE.sub("#", b_norm)

    sm = difflib.SequenceMatcher(None, a_template, b_template, autojunk=False)
    if sm.quick_ratio() < 0.70:
        tpl_pct = int(sm.ratio() * 100)
    else:
        tpl_pct = int(sm.ratio() * 100)
    # token_pct is computed on the original normalized text (with numbers
    # preserved) so that same-template / different-number pairs trip the
    # template_only guard instead of being mis-classified as similar.
    token_pct = _token_similarity_pct(a_norm, b_norm)
    final_pct = max(tpl_pct, token_pct)

    # ٢) الأرقام المذكورة
    a_nums = set(_NUM_RE.findall(a or ""))
    b_nums = set(_NUM_RE.findall(b or ""))
    if a_nums or b_nums:
        nums_overlap = len(a_nums & b_nums) / max(len(a_nums | b_nums), 1)
    else:
        nums_overlap = 1.0  # لا أرقام في الطرفين = اعتبرها متطابقة

    # ٣) الحكم
    if final_pct >= 90 and nums_overlap >= 0.5:
        return ("identical", 5, final_pct)  # نص متطابق حقيقي → +5
    if tpl_pct >= 90 and nums_overlap < 0.3:
        return ("template_only", 0, final_pct)  # قالب موحّد بأرقام مختلفة → +0
    if final_pct >= 90:
        return ("similar", 3, final_pct)  # نص متشابه ≥90% → +3
    return ("different", 0, final_pct)


_DETAIL_STOPWORDS = {
    "من",
    "في",
    "عن",
    "على",
    "الى",
    "الي",
    "رقم",
    "عدد",
    "داخل",
    "بداخل",
    "اقرب",
    "معلم",
    "الدوره",
    "الدورة",
    "دوره",
    "دورة",
    "المياه",
    "مياه",
    "القسم",
    "النسائي",
    "حمامات",
    "حسب",
    "افادة",
    "المبلغ",
}


def _detail_tokens(s: str) -> set[str]:
    # Numbers are preserved so same-template / different-number pairs
    # diverge at the token level (feeds the template_only guard).
    toks = re.findall(r"[\u0600-\u06FFa-zA-Z0-9/]+", s or "")
    return {t for t in toks if len(t) > 1 and t not in _DETAIL_STOPWORDS}


def _token_similarity_pct(a: str, b: str) -> int:
    ta, tb = _detail_tokens(a), _detail_tokens(b)
    if not ta or not tb:
        return 0
    inter = len(ta & tb)
    dice = (2 * inter) / max(len(ta) + len(tb), 1)
    containment = inter / max(min(len(ta), len(tb)), 1)
    fuzzy = _fuzzy_token_containment(ta, tb)
    # Short field reports often say the same issue with extra context in one side.
    # Containment catches that without weakening the hard asset/location/fault gates.
    pct = max(dice, containment * 0.85, fuzzy)
    return int(round(pct * 100))


def _fuzzy_token_containment(a: set[str], b: set[str]) -> float:
    small, large = (a, b) if len(a) <= len(b) else (b, a)
    if not small:
        return 0.0
    matched = 0
    for tok in small:
        if tok in large:
            matched += 1
            continue
        if any(
            difflib.SequenceMatcher(None, tok, other, autojunk=False).ratio() >= 0.72
            for other in large
        ):
            matched += 1
    return matched / len(small)


# ─────────────────────────────────────────────────────────────────────────────
# مساعدات الزمن
# ─────────────────────────────────────────────────────────────────────────────


def _parse_date(s: str):
    """يحاول تحويل نص لتاريخ. يرجّع None لو فشل."""
    if not s:
        return None
    s = _str(s)
    if not s:
        return None
    # صيغ متوقعة من ماكسيمو
    formats = [
        "%m/%d/%y %I:%M %p",
        "%m/%d/%Y %I:%M %p",
        "%m/%d/%y %H:%M",
        "%m/%d/%Y %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%d/%m/%Y %H:%M",
        "%d/%m/%y %I:%M %p",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    # محاولة pandas الأخيرة
    try:
        return pd.to_datetime(s, errors="coerce", dayfirst=False).to_pydatetime()
    except Exception:
        return None


def _fmt_arabic_gap(td: timedelta) -> str:
    """يحوّل timedelta لنص عربي مثل '+12 دقيقة' أو '+2 يوم'."""
    if td is None:
        return "—"
    total = abs(int(td.total_seconds()))
    if total < 60:
        return f"+{total} ثانية"
    if total < 3600:
        m = total // 60
        return f"+{m} دقيقة"
    if total < 86400:
        h = total // 3600
        m = (total % 3600) // 60
        return f"+{h} ساعة {m} دقيقة" if m else f"+{h} ساعة"
    days = total // 86400
    if days <= 14:
        return f"+{days} يوم"
    weeks = days // 7
    if weeks <= 8:
        return f"+{weeks} أسبوع"
    months = days // 30
    if months <= 12:
        return f"+{months} شهر"
    years = days // 365
    return f"+{years} سنة"


def _fmt_date(d) -> str:
    if d is None:
        return ""
    return d.strftime("%Y-%m-%d %H:%M")


# ─────────────────────────────────────────────────────────────────────────────
# Union-Find للتجميع
# ─────────────────────────────────────────────────────────────────────────────


class DSU:
    def __init__(self):
        self.parent = {}

    def find(self, x):
        if x not in self.parent:
            self.parent[x] = x
        while self.parent[x] != x:
            self.parent[x] = self.parent[self.parent[x]]
            x = self.parent[x]
        return x

    def union(self, a, b):
        ra, rb = self.find(a), self.find(b)
        if ra != rb:
            self.parent[ra] = rb


# ─────────────────────────────────────────────────────────────────────────────
# scoring
# ─────────────────────────────────────────────────────────────────────────────


def _score(r1: dict, r2: dict) -> tuple[int, list[str], dict]:
    """
    احسب نقاط التشابه بين بلاغين. يرجّع (المجموع، الأسباب، metadata).

    ملاحظة: العطل والموقع مضمونان من الـ blocking ولا يُحتسبان هنا.
    """
    score = 0
    reasons = []
    meta = {"tpl_pct": 0, "txt_class": "different"}

    # ١) الموقع مضمون (blocking) — +4 للموقع
    if r1["loc"]:
        reasons.append(f"نفس الموقع ({r1['loc']})")
        score += 4

    # ١ب) العطل مضمون (blocking) — +3 للعطل
    if r1["fault"]:
        reasons.append(f"نفس العطل ({r1['fault']})")
        score += 3

    # ٢) نفس الأصل +4 — فقط إذا يختلف عن الموقع (تفادياً للاحتساب المزدوج)
    if r1["asset"] and r2["asset"] and r1["asset"] == r2["asset"]:
        loc_n = _normalize_ar(r1["loc"])
        ast_n = _normalize_ar(r1["asset"])
        if ast_n != loc_n:
            score += 4
            reasons.append(f"نفس الأصل ({r1['asset']})")
        else:
            reasons.append(f"نفس الأصل/الموقع ({r1['asset']})")  # مذكور مرة واحدة فقط

    # ٣) مقارنة التفاصيل الذكية
    cls, pts, tpl_pct = _smart_text_compare(r1["detail"], r2["detail"])
    meta["tpl_pct"] = tpl_pct
    meta["txt_class"] = cls
    if pts > 0:
        score += pts
        label = {
            "identical": f"تفاصيل متطابقة ({tpl_pct}%)",
            "similar": f"تفاصيل متشابهة ({tpl_pct}%)",
        }.get(cls, "")
        if label:
            reasons.append(label)
    elif cls == "template_only":
        reasons.append(f"تنبيه: قالب موحّد بأرقام مختلفة ({tpl_pct}%)")

    # ٤) نفس رقم المبلّغ +2
    if (
        r1.get("requestor_no")
        and r2.get("requestor_no")
        and r1["requestor_no"] == r2["requestor_no"]
    ):
        score += 2
        reasons.append(f"نفس رقم المبلّغ ({r1['requestor_no']})")

    # ٥) الفارق الزمني (يُحتسب في النقاط)
    dt1, dt2 = r1.get("reported_dt"), r2.get("reported_dt")
    if dt1 and dt2:
        gap_days = abs((dt2 - dt1).total_seconds()) / 86400
        if gap_days < 1:
            score += 3
            reasons.append("نفس اليوم (+3)")
        elif gap_days < 2:
            score += 2
            reasons.append("فارق يوم واحد (+2)")
        elif gap_days <= 2:
            score += 1
            reasons.append("فارق يومان (+1)")

    return score, reasons, meta


# ─────────────────────────────────────────────────────────────────────────────
# detection
# ─────────────────────────────────────────────────────────────────────────────


def detect(df: pd.DataFrame, min_score: int = 5, max_days: int = 2) -> dict:
    """يكتشف المكررات ويرجّع {"pairs":[...], "groups":[...], "rows_by_sr":{}}."""
    col_sr = _find_col(df, "Service Request", "SR")
    col_loc = _find_col(df, "LOCATION", "Location")
    col_asset = _find_col(df, "Asset", "ASSET")
    col_summ = _find_col(df, "Summary", "SUMMARY")
    col_detail = _find_col(df, "Details", "DETAILS")
    col_reporter = _find_col(df, "REPORTED NAME", "Reported By", "Reporter")
    col_status = _find_col(df, "Status", "STATUS")
    col_site = _find_col(df, "Site", "SITE")
    col_workzone = _find_col(df, "Work Zone", "WorkZone")
    col_reported = _find_col(df, "تاريخ فتح البلاغ", "Reported Date", "Report Date")
    col_started = _find_col(df, "تاريخ المباشره", "تاريخ المباشرة", "Start Date")
    col_resolved = _find_col(df, "تاريخ المعالجة", "Resolved Date")
    col_statusdt = _find_col(df, "Status Date")
    # أعمدة إضافية
    col_requestor = _find_col(df, "REQUESTOR NO.", "Requestor No.", "REQUESTOR")
    col_source = _find_col(df, "Source", "SOURCE")
    col_history = _find_col(df, "History", "HISTORY")
    col_status_desc = _find_col(df, "Status Description", "StatusDescription")
    col_block = _find_col(df, "المربع")
    col_region = _find_col(df, "المنطقة")
    col_asset_ar = _find_col(df, "الأصل")
    col_reported_by = _find_col(df, "Reported By")
    col_resp_time = _find_col(df, "زمن الاستجابه", "Response Time")
    col_resp_esc = _find_col(df, "Response Esclation", "Response Escalation")
    col_resol_time = _find_col(df, "Resolution Time")
    col_resol_esc = _find_col(df, "Resolution Escalation")
    col_party = _find_col(df, "الجهة")
    col_lat = _find_col(df, "Latitude(Y)", "Latitude", "LAT")
    col_lon = _find_col(df, "Longitude(X)", "Longitude", "LON")
    col_priority = _find_col(df, "Internal Priority", "Priority")
    col_ticket_other = _find_col(df, "Ticket in Other Party")
    col_contract = _find_col(df, "Contract", "CONTRACT")
    col_contractor = _find_col(df, "Contractor", "CONTRACTOR")

    if not col_sr:
        raise SystemExit("عمود 'Service Request' غير موجود")

    rows = []
    rows_by_sr = {}
    records = df.to_dict("records")
    for raw in records:
        rd = {str(k): _str(v) for k, v in raw.items()}
        sr = rd.get(col_sr, "") if col_sr else ""
        if not sr:
            continue
        detail = _strip_html(rd.get(col_detail, "") if col_detail else "")
        loc = rd.get(col_loc, "") if col_loc else ""
        asset = rd.get(col_asset, "") if col_asset else ""
        asset = re.sub(r"\.0+$", "", asset)
        fault_full_o = rd.get(col_summ, "") if col_summ else ""
        # ──────────────────────────────────────────────────────────────────
        # التصنيف عبارة عن 4 أقسام مفصولة بفواصل:
        #   "مجمعات دورات المياه,الدورات العامة,غرفة التفتيش,تلف احد المكونات"
        #
        # المشكلة: بعض البلاغات تُدخَل بقسمين فقط:
        #   "مجمعات دورات المياه,تلف احد المكونات"
        # آخر جزئَين هنا يختلفان عن آخر جزئَين في البلاغ ذي الأربعة أقسام.
        #
        # الحل: نستخدم آخر قسم واحد فقط كمفتاح الـ blocking
        # (الأكثر تحديداً ومشتركاً بين جميع مستويات الإدخال).
        # أما fault_o (للعرض والتسمية) فنبقيه آخر قسمَين.
        # ──────────────────────────────────────────────────────────────────
        _parts = [p.strip() for p in fault_full_o.split(",") if p.strip()]
        fault_o = ",".join(_parts[-2:]) if len(_parts) >= 2 else fault_full_o
        fault_n = _normalize_ar(fault_o)
        fault_last = _normalize_ar(_parts[-1]) if _parts else fault_n  # آخر قسم واحد للـ blocking

        row = {
            "sr": sr,
            "loc": loc,
            "asset": asset,
            "fault": fault_n,
            "fault_last": fault_last,  # آخر قسم واحد — مفتاح الـ blocking
            "fault_orig": fault_o,
            "fault_full": fault_full_o,
            "detail": detail,
            "reporter": rd.get(col_reporter, "") if col_reporter else "",
            "status": rd.get(col_status, "") if col_status else "",
            "site": rd.get(col_site, "") if col_site else "",
            "workzone": rd.get(col_workzone, "") if col_workzone else "",
            "reported": rd.get(col_reported, "") if col_reported else "",
            "started": rd.get(col_started, "") if col_started else "",
            "resolved": rd.get(col_resolved, "") if col_resolved else "",
            "status_dt": rd.get(col_statusdt, "") if col_statusdt else "",
            "reported_dt": _parse_date(rd.get(col_reported, "") if col_reported else ""),
            # أعمدة إضافية
            "requestor_no": rd.get(col_requestor, "") if col_requestor else "",
            "source": rd.get(col_source, "") if col_source else "",
            "history": rd.get(col_history, "") if col_history else "",
            "status_desc": rd.get(col_status_desc, "") if col_status_desc else "",
            "block": rd.get(col_block, "") if col_block else "",
            "region": rd.get(col_region, "") if col_region else "",
            "asset_ar": _ASSET_MAP.get(asset, ""),
            "loc_ar": _LOC_MAP.get(loc, ""),
            "reported_by": rd.get(col_reported_by, "") if col_reported_by else "",
            "resp_time": rd.get(col_resp_time, "") if col_resp_time else "",
            "resp_esc": rd.get(col_resp_esc, "") if col_resp_esc else "",
            "resol_time": rd.get(col_resol_time, "") if col_resol_time else "",
            "resol_esc": rd.get(col_resol_esc, "") if col_resol_esc else "",
            "party": rd.get(col_party, "") if col_party else "",
            "lat": rd.get(col_lat, "") if col_lat else "",
            "lon": rd.get(col_lon, "") if col_lon else "",
            "priority": rd.get(col_priority, "") if col_priority else "",
            "ticket_other": rd.get(col_ticket_other, "") if col_ticket_other else "",
            "contract": rd.get(col_contract, "") if col_contract else "",
            "contractor": rd.get(col_contractor, "") if col_contractor else "",
            "_row_orig": rd,
        }
        rows.append(row)
        rows_by_sr[sr] = row

    # ─── Blocking: نجمّع حسب (fault, location) ───
    # هذا يضمن أن كل زوج في الـ block يشترك بنفس العطل والموقع (100%)
    # البلاغات التي موقعها فارغ لا تدخل في أي block (الموقع بوابة صلبة)
    fault_loc_groups = defaultdict(list)
    for r in rows:
        loc_n = _normalize_ar(r["loc"])
        if not r["fault"] or not loc_n:
            continue
        fault_loc_groups[(r["fault"], loc_n)].append(r)

    pairs_raw = []
    pair_keys = set()

    def _try_pair(r1, r2):
        if r1["sr"] == r2["sr"]:
            return
        key = (min(r1["sr"], r2["sr"]), max(r1["sr"], r2["sr"]))
        if key in pair_keys:
            return
        pair_keys.add(key)

        # ── بوابة الأصل: إذا الطرفان عندهما أصل فيجب أن يكون متطابقاً ──
        a1, a2 = r1["asset"], r2["asset"]
        if a1 and a2 and a1 != a2:
            return

        # ── بوابة الزمن: الفارق يجب أن يكون ≤ max_days (إذا التاريخ موجود) ──
        dt1, dt2 = r1["reported_dt"], r2["reported_dt"]
        if dt1 and dt2:
            gap_days = abs((dt2 - dt1).total_seconds()) / 86400
            if gap_days > max_days:
                return

        # ── بوابة التفاصيل: النص يجب أن يكون متشابهاً 90% أو أكثر ──
        # استثناء: لو التصنيف الكامل متطابق تماماً (بما فيه الناقص)،
        # فالموقع + التصنيف كافيَين → نتجاوز بوابة التفاصيل
        full1 = _normalize_ar(r1.get("fault_full", ""))
        full2 = _normalize_ar(r2.get("fault_full", ""))
        exact_fault_match = bool(full1 and full1 == full2)
        txt_cls, _, tpl_pct = _smart_text_compare(r1["detail"], r2["detail"])
        if not exact_fault_match:
            if txt_cls in ("different", "template_only") or tpl_pct < 90:
                return

        sc, reasons, meta = _score(r1, r2)
        if sc < min_score:
            return
        pairs_raw.append(
            {
                "r1": r1,
                "r2": r2,
                "score": sc,
                "reasons": reasons,
                "meta": meta,
            }
        )

    for (fault, loc_n), grp in fault_loc_groups.items():
        n = len(grp)
        if n <= 400:
            for i in range(n):
                for j in range(i + 1, n):
                    _try_pair(grp[i], grp[j])
        else:
            # block كبير جداً → نقسّم بالأصل (asset)
            sub = defaultdict(list)
            for r in grp:
                sub[r["asset"] or "__no_asset__"].append(r)
            for sub_grp in sub.values():
                if len(sub_grp) > 500:
                    continue
                m = len(sub_grp)
                for i in range(m):
                    for j in range(i + 1, m):
                        _try_pair(sub_grp[i], sub_grp[j])

    # ─── Union-Find لتجميع الأزواج في مجموعات ───
    # كل زوج اجتاز البوابات الصارمة → ندمجه في مجموعة
    dsu = DSU()
    pair_scores = {}
    for p in pairs_raw:
        a, b = p["r1"]["sr"], p["r2"]["sr"]
        pair_scores[(min(a, b), max(a, b))] = p
        dsu.union(a, b)
    all_pair_srs = {p["r1"]["sr"] for p in pairs_raw} | {p["r2"]["sr"] for p in pairs_raw}
    clusters_map = defaultdict(list)
    for sr in all_pair_srs:
        clusters_map[dsu.find(sr)].append(sr)

    # بناء المجموعات النهائية
    groups = []
    for root, members in clusters_map.items():
        if len(members) < 2:
            continue
        member_rows = [rows_by_sr[sr] for sr in members]
        # ترتيب زمني
        member_rows.sort(key=lambda r: r["reported_dt"] or datetime.max)

        # أعلى نقطة بين الأزواج الداخلية
        inner_pairs = []
        for a, b in itertools.combinations(members, 2):
            key = (min(a, b), max(a, b))
            if key in pair_scores:
                inner_pairs.append(pair_scores[key])
        base = max((p["score"] for p in inner_pairs), default=0)
        size_bonus = min(max(0, len(members) - 2), 3)  # حد أقصى +3
        group_score = base + size_bonus

        # تجميع الأسباب من كل الأزواج الداخلية (uniq)
        all_reasons = []
        seen_reasons = set()
        for p in inner_pairs:
            for r in p["reasons"]:
                if r not in seen_reasons:
                    all_reasons.append(r)
                    seen_reasons.add(r)

        conf = "مراجعة"

        # تحليل زمني
        dates = [r["reported_dt"] for r in member_rows if r["reported_dt"]]
        time_stats = {}
        if len(dates) >= 2:
            dates.sort()
            time_stats = {
                "first": dates[0],
                "last": dates[-1],
                "span": dates[-1] - dates[0],
                "max_gap": max(dates[i + 1] - dates[i] for i in range(len(dates) - 1)),
                "same_day": all(d.date() == dates[0].date() for d in dates),
                "same_week": (dates[-1] - dates[0]).days <= 7,
            }

        groups.append(
            {
                "members": member_rows,
                "score": group_score,
                "base_score": base,
                "size": len(members),
                "confidence": conf,
                "reasons": all_reasons,
                "inner_pairs": inner_pairs,
                "time_stats": time_stats,
            }
        )

    groups.sort(key=lambda g: (-g["score"], -g["size"]))

    return {
        "pairs": pairs_raw,
        "groups": groups,
        "rows_by_sr": rows_by_sr,
        "all_rows": rows,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Excel styling helpers
# ─────────────────────────────────────────────────────────────────────────────

C_DARK = "1C2E4A"
C_GOLD = "B9975B"
C_GOLD_BG = "FDF8EE"
C_WHITE = "FFFFFF"
C_GREY = "F0EDE6"
C_GREEN_BG = "D6EFD8"  # أخضر ناعم — نفس اليوم
C_AMBER = "FEF3C7"
C_BLUE = "DBEAFE"
C_GREEN = "D4EDDA"

# لون موحّد لجميع المجموعات (لا تصنيف ثقة)
TIER_COLOR = ("EAF0FB", "1E3A8A")  # أزرق فاتح / أزرق داكن

# ألوان الفارق الزمني (بدون أحمر)
DATE_COLORS = {
    "same_day": "D6EFD8",  # أخضر ناعم — نفس اليوم (تحذير إيجابي)
    "same_week": "FEF3C7",  # عنبري — نفس الأسبوع
    "far": "EAF0FB",  # أزرق خفيف — بعيد
}

# الـ helpers (_fill, _font, _thin, _al, _hrow, _cell) تُعرّف داخل write_report
# لأنها تحتاج openpyxl الذي نُحمّله lazy.


# ─────────────────────────────────────────────────────────────────────────────
# Excel output: المجموعات
# ─────────────────────────────────────────────────────────────────────────────

# الأعمدة في Sheet 1 — مجموعة ثم كل أعمدة ماكسيمو الأصلية
GROUP_COLS = [
    # أعمدة المجموعة
    "رقم المجموعة",
    "ترتيب البلاغ",
    "الفارق عن السابق",
    "أسباب الكشف",
    # أعمدة ماكسيمو الأصلية (بنفس ترتيب ملف التصدير)
    "Service Request",
    "Site",
    "History",
    "Source",
    "Status",
    "Status Date",
    "Status Description",
    "Summary",
    "Work Zone",
    "LOCATION",
    "وصف المعلم",
    "المربع",
    "المنطقة",
    "Asset",
    "وصف الأصل",
    "REQUESTOR NO.",
    "Contract",
    "Contractor",
    "REPORTED NAME",
    "Reported By",
    "زمن الاستجابه",
    "Response Esclation",
    "Resolution Time",
    "Resolution Escalation",
    "الجهة",
    "Details",
    "Latitude(Y)",
    "Longitude(X)",
    "Internal Priority",
    "Ticket in Other Party",
    "تاريخ فتح البلاغ",
    "تاريخ المباشره",
    "تاريخ المعالجة",
]


def write_report(detection: dict, output_path: str, total_rows: int):
    # استيراد openpyxl محلياً (lazy) — لا يلزم FastAPI server
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    # styling helpers (تستخدم openpyxl)
    def _fill(c):
        return PatternFill("solid", fgColor=c)

    def _font(bold=False, size=10, color="000000"):
        return Font(bold=bold, size=size, color=color, name="Arial")

    def _thin():
        s = Side(style="thin", color="D0D0D0")
        return Border(left=s, right=s, top=s, bottom=s)

    def _al(h="right", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrapText=wrap, readingOrder=2)

    def _hrow(ws, row, vals, bg=C_DARK, fg="FFFFFF"):
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.fill = _fill(bg)
            c.font = _font(True, 10, fg)
            c.alignment = _al("center")
            c.border = _thin()
        ws.row_dimensions[row].height = 24

    def _cell(
        ws,
        row,
        col,
        val,
        bg=C_WHITE,
        bold=False,
        wrap=False,
        align="right",
        color="000000",
        size=10,
    ):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = _fill(bg)
        c.font = _font(bold, size, color)
        c.alignment = _al(align, wrap=wrap)
        c.border = _thin()
        return c

    groups = detection["groups"]

    # تدرّجات اللون بين المجموعات (نفس المجموعة لون موحّد، المجموعات المتجاورة مختلفة)
    GROUP_SHADES = ["F8FAFC", "EFF6FF", "F0FDF4", "FEF3C7", "FAF5FF"]

    wb = openpyxl.Workbook()

    # ════════ Sheet 1: مجموعات المكررات ════════
    ws1 = wb.active
    ws1.title = "مجموعات المكررات"
    ws1.sheet_view.rightToLeft = True
    ws1.freeze_panes = "A3"

    # عنوان
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(GROUP_COLS))
    t = ws1.cell(
        1,
        1,
        f"مجموعات المكررات المحتملة — {len(groups)} مجموعة | "
        f"{sum(g['size'] for g in groups)} بلاغ متأثر",
    )
    t.fill = _fill(C_DARK)
    t.font = _font(True, 13, C_GOLD)
    t.alignment = _al("center")
    ws1.row_dimensions[1].height = 30

    _hrow(ws1, 2, GROUP_COLS)

    ri = 3
    for gi, g in enumerate(groups, 1):
        members = g["members"]
        shade = GROUP_SHADES[gi % len(GROUP_SHADES)]
        tier_bg, tier_fg = TIER_COLOR

        prev_dt = None
        for mi, row in enumerate(members, 1):
            # حساب الفارق عن السابق
            cur_dt = row["reported_dt"]
            if mi == 1 or prev_dt is None or cur_dt is None:
                gap_str = "—"
                gap_color = C_WHITE
            else:
                gap_td = cur_dt - prev_dt
                gap_str = _fmt_arabic_gap(gap_td)
                # لون الفارق
                days = gap_td.days
                if days == 0:
                    gap_color = DATE_COLORS["same_day"]
                elif days <= 7:
                    gap_color = DATE_COLORS["same_week"]
                else:
                    gap_color = DATE_COLORS["far"]
            prev_dt = cur_dt or prev_dt

            # تلوين الصف الأول (الأصل)
            is_origin = mi == 1
            bg = shade

            # ── أعمدة المجموعة (1–4) ──
            _cell(ws1, ri, 1, f"#{gi}", bg=bg, bold=True, align="center")
            _cell(
                ws1,
                ri,
                2,
                f"{mi}" + (" (الأصل)" if is_origin else ""),
                bg=bg,
                bold=is_origin,
                align="center",
            )
            _cell(ws1, ri, 3, gap_str, bg=gap_color, bold=True, align="center")
            _cell(
                ws1,
                ri,
                4,
                " | ".join(g["reasons"]) if is_origin else "",
                bg=bg,
                wrap=True,
                color=tier_fg,
            )
            # ── أعمدة ماكسيمو الأصلية (5–36) ──
            _cell(ws1, ri, 5, row["sr"], bg=bg, bold=True, align="center")
            _cell(ws1, ri, 6, row["site"], bg=bg, align="center")
            _cell(ws1, ri, 7, row["history"], bg=bg, wrap=True)
            _cell(ws1, ri, 8, row["source"], bg=bg, align="center")
            _cell(ws1, ri, 9, row["status"], bg=bg, align="center")
            _cell(ws1, ri, 10, row["status_dt"], bg=bg, align="center")
            _cell(ws1, ri, 11, row["status_desc"], bg=bg, wrap=True)
            _cell(ws1, ri, 12, row["fault_full"], bg=bg)
            _cell(ws1, ri, 13, row["workzone"], bg=bg, align="center")
            _cell(ws1, ri, 14, row["loc"], bg=bg, align="center")
            _cell(ws1, ri, 15, row["loc_ar"], bg=bg)  # وصف المعلم
            _cell(ws1, ri, 16, row["block"], bg=bg, align="center")
            _cell(ws1, ri, 17, row["region"], bg=bg, align="center")
            _cell(ws1, ri, 18, row["asset"], bg=bg, align="center")
            _cell(ws1, ri, 19, row["asset_ar"], bg=bg)  # وصف الأصل
            _cell(ws1, ri, 20, row["requestor_no"], bg=bg, align="center")
            _cell(ws1, ri, 21, row["contract"], bg=bg, align="center")
            _cell(ws1, ri, 22, row["contractor"], bg=bg)
            _cell(ws1, ri, 23, row["reporter"], bg=bg)
            _cell(ws1, ri, 24, row["reported_by"], bg=bg)
            _cell(ws1, ri, 25, row["resp_time"], bg=bg, align="center")
            _cell(ws1, ri, 26, row["resp_esc"], bg=bg, align="center")
            _cell(ws1, ri, 27, row["resol_time"], bg=bg, align="center")
            _cell(ws1, ri, 28, row["resol_esc"], bg=bg, align="center")
            _cell(ws1, ri, 29, row["party"], bg=bg, align="center")
            _cell(ws1, ri, 30, row["detail"], bg=bg, wrap=True)
            _cell(ws1, ri, 31, row["lat"], bg=bg, align="center")
            _cell(ws1, ri, 32, row["lon"], bg=bg, align="center")
            _cell(ws1, ri, 33, row["priority"], bg=bg, align="center")
            _cell(ws1, ri, 34, row["ticket_other"], bg=bg, align="center")
            _cell(
                ws1, ri, 35, _fmt_date(cur_dt) if cur_dt else row["reported"], bg=bg, align="center"
            )
            _cell(ws1, ri, 36, row["started"], bg=bg, align="center")
            _cell(ws1, ri, 37, row["resolved"], bg=bg, align="center")

            # ارتفاع الصف
            ws1.row_dimensions[ri].height = 45
            ri += 1

        # فاصل بين المجموعات
        ws1.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=len(GROUP_COLS))
        sep = ws1.cell(ri, 1, "")
        sep.fill = _fill(C_GREY)
        ws1.row_dimensions[ri].height = 6
        ri += 1

    # عرض الأعمدة (1-4 أعمدة مجموعة، 5-37 أعمدة ماكسيمو)
    col_widths = {
        1: 9,
        2: 11,
        3: 16,
        4: 40,  # أعمدة المجموعة
        5: 14,
        6: 8,
        7: 18,
        8: 12,
        9: 12,
        10: 18,
        11: 22,  # SR…Status Description
        12: 32,
        13: 14,
        14: 14,
        15: 28,
        16: 12,
        17: 12,  # Summary…المنطقة (15=وصف المعلم)
        18: 14,
        19: 28,
        20: 16,
        21: 14,
        22: 20,  # Asset…Contractor (19=وصف الأصل)
        23: 20,
        24: 20,  # REPORTED NAME, Reported By
        25: 14,
        26: 14,
        27: 14,
        28: 14,
        29: 14,  # زمن الاستجابه…الجهة
        30: 45,  # Details
        31: 14,
        32: 14,
        33: 14,
        34: 18,  # Lat, Lon, Priority, Ticket
        35: 18,
        36: 18,
        37: 18,  # تواريخ
    }
    for ci, w in col_widths.items():
        ws1.column_dimensions[get_column_letter(ci)].width = w

    # ════════ Sheet 2: ملخص المجموعات ════════
    ws2 = wb.create_sheet("ملخص المجموعات")
    ws2.sheet_view.rightToLeft = True
    ws2.freeze_panes = "A3"

    SUMMARY_COLS = [
        "رقم المجموعة",
        "حجم المجموعة",
        "الثقة",
        "النقاط",
        "أعضاء المجموعة",
        "تاريخ أول بلاغ",
        "تاريخ آخر بلاغ",
        "المدى الزمني",
        "أسباب الكشف",
    ]
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(SUMMARY_COLS))
    t2 = ws2.cell(1, 1, f"ملخص المجموعات ({len(groups)})")
    t2.fill = _fill(C_DARK)
    t2.font = _font(True, 13, C_GOLD)
    t2.alignment = _al("center")
    ws2.row_dimensions[1].height = 28

    _hrow(ws2, 2, SUMMARY_COLS)

    for gi, g in enumerate(groups, 1):
        ri = 2 + gi
        tier_bg, tier_fg = TIER_COLOR
        bg = C_GOLD_BG if gi % 2 == 0 else C_WHITE
        ts = g["time_stats"]
        members_str = " ، ".join(m["sr"] for m in g["members"])

        _cell(ws2, ri, 1, f"#{gi}", bg=bg, bold=True, align="center")
        _cell(ws2, ri, 2, g["size"], bg=bg, align="center")
        _cell(ws2, ri, 3, "مراجعة", bg=tier_bg, bold=True, align="center", color=tier_fg)
        _cell(ws2, ri, 4, g["score"], bg=bg, bold=True, align="center")
        _cell(ws2, ri, 5, members_str, bg=bg, align="center")
        _cell(ws2, ri, 6, _fmt_date(ts.get("first")) if ts else "", bg=bg, align="center")
        _cell(ws2, ri, 7, _fmt_date(ts.get("last")) if ts else "", bg=bg, align="center")
        _cell(
            ws2,
            ri,
            8,
            _fmt_arabic_gap(ts.get("span")) if ts.get("span") else "—",
            bg=bg,
            align="center",
        )
        _cell(ws2, ri, 9, " | ".join(g["reasons"]), bg=bg, wrap=True)
        ws2.row_dimensions[ri].height = 30

    for ci, w in {1: 10, 2: 10, 3: 10, 4: 10, 5: 30, 6: 18, 7: 18, 8: 16, 9: 50}.items():
        ws2.column_dimensions[get_column_letter(ci)].width = w

    # ════════ Sheet 3: إحصاءات ════════
    ws3 = wb.create_sheet("إحصاءات")
    ws3.sheet_view.rightToLeft = True
    ws3.column_dimensions["A"].width = 42
    ws3.column_dimensions["B"].width = 16

    ws3.merge_cells("A1:B1")
    th = ws3.cell(1, 1, "إحصاءات المجموعات")
    th.fill = _fill(C_DARK)
    th.font = _font(True, 13, C_GOLD)
    th.alignment = _al("center")
    ws3.row_dimensions[1].height = 32

    same_day_groups = [g for g in groups if g["time_stats"].get("same_day")]
    same_week_groups = [g for g in groups if g["time_stats"].get("same_week")]
    big_groups = [g for g in groups if g["size"] >= 3]
    affected_srs = sum(g["size"] for g in groups)

    # توزيع العطل الأكثر تكراراً
    fault_counts = defaultdict(int)
    for g in groups:
        for m in g["members"]:
            fault_counts[m["fault_orig"]] += 1
    top_faults = sorted(fault_counts.items(), key=lambda x: -x[1])[:10]

    # توزيع حسب المصدر (SCADA وغيره)
    source_counts = defaultdict(int)
    for g in groups:
        for m in g["members"]:
            src = m.get("source", "") or "غير محدد"
            source_counts[src] += 1
    top_sources = sorted(source_counts.items(), key=lambda x: -x[1])[:5]

    stats = [
        ("إجمالي البلاغات في الملف", total_rows),
        ("إجمالي مجموعات المكررات المحتملة", len(groups)),
        ("البلاغات المتأثرة (داخل مجموعات)", affected_srs),
        None,
        ("مجموعات تحوي 3 بلاغات أو أكثر", len(big_groups)),
        ("مجموعات بلاغاتها في نفس اليوم", len(same_day_groups)),
        ("مجموعات بلاغاتها خلال أسبوع", len(same_week_groups)),
        None,
        ("— أكثر 10 أعطال تكراراً —", ""),
    ]
    stats.extend(top_faults)
    stats.append(None)
    stats.append(("— توزيع حسب المصدر —", ""))
    stats.extend(top_sources)

    for ri2, item in enumerate(stats, 2):
        if item is None:
            ws3.row_dimensions[ri2].height = 8
            continue
        lbl, val = item
        bg = C_GOLD_BG if ri2 % 2 == 0 else C_WHITE
        ca = ws3.cell(ri2, 1, lbl)
        ca.fill = _fill(bg)
        ca.font = _font(True)
        ca.alignment = _al()
        ca.border = _thin()
        cb = ws3.cell(ri2, 2, val)
        cb.fill = _fill(bg)
        cb.font = _font(True, 12, C_GOLD)
        cb.alignment = _al("center")
        cb.border = _thin()
        ws3.row_dimensions[ri2].height = 22

    wb.active = ws1
    wb.save(output_path)


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────


def main():
    import sys

    # تفادي خطأ تشفير الـ console على Windows (cp1256)
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

    p = argparse.ArgumentParser(
        description="كشف البلاغات المكررة v2 — مجموعات + مقارنة ذكية + زمن",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    p.add_argument("input", help="ملف البلاغات (XLS أو XLSX)")
    p.add_argument("--output", default=None)
    p.add_argument(
        "--min-score", type=int, default=5, help="الحد الأدنى للنقاط (افتراضي: 5 = محتمل+مؤكد)"
    )
    p.add_argument(
        "--max-days",
        type=int,
        default=2,
        help="أقصى فارق زمني مسموح (بالأيام) للاعتبار زوجاً (افتراضي: 2)",
    )
    args = p.parse_args()

    out = args.output or (Path(args.input).stem + "_مكررات_v2.xlsx")

    print(f"\n>>> قراءة: {args.input}", flush=True)
    df = read_file(args.input)
    print(f"    {len(df):,} بلاغ  |  {len(df.columns)} عمود", flush=True)

    print(">>> فحص المكررات (Dropdowns + Smart Text + Time)...", flush=True)
    detection = detect(df, min_score=args.min_score, max_days=args.max_days)
    groups = detection["groups"]

    big = sum(1 for g in groups if g["size"] >= 3)
    affected = sum(g["size"] for g in groups)

    print(f"    مجموعات مكررات محتملة: {len(groups)}", flush=True)
    print(f"    مجموعات 3+ بلاغات:     {big}", flush=True)
    print(f"    البلاغات المتأثرة:      {affected} من {len(df)}", flush=True)

    if not groups:
        print("    لا توجد مكررات", flush=True)
        return

    print(f"\n>>> حفظ: {out}", flush=True)
    write_report(detection, out, total_rows=len(df))
    print("    تم Excel", flush=True)

    # حفظ pickle للـ Streamlit والـ live_monitor (للوصول لكامل البيانات بسرعة)
    import pickle

    pkl_path = Path(out).with_suffix(".pkl")
    # all_rows: كل البلاغات (بدون _row_orig لتوفير الحجم) — يحتاجها live_monitor
    safe_all_rows = [
        {k: v for k, v in r.items() if k != "_row_orig"} for r in detection["all_rows"]
    ]
    serializable = {
        "groups": [
            {
                "members": g["members"],
                "score": g["score"],
                "base_score": g["base_score"],
                "size": g["size"],
                "confidence": g["confidence"],
                "reasons": g["reasons"],
                "time_stats": g["time_stats"],
            }
            for g in detection["groups"]
        ],
        "all_rows": safe_all_rows,
        "total_rows": len(df),
        "input_path": args.input,
    }
    with open(pkl_path, "wb") as f:
        pickle.dump(serializable, f)
    print(f"    تم pickle: {pkl_path}", flush=True)

    print("\n-- أبرز المجموعات (الأعلى نقاطاً) --", flush=True)
    for g in groups[:5]:
        members_str = " ، ".join(m["sr"] for m in g["members"])
        print(f"  [{g['score']} نقطة, {g['size']} بلاغات]  {members_str}", flush=True)
        for r in g["reasons"][:4]:
            print(f"    - {r}", flush=True)
        print(flush=True)


if __name__ == "__main__":
    main()
